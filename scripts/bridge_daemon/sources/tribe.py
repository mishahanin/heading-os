"""Real-data source for the /tribe endpoint.

Parses crm/contacts/*.md YAML frontmatter and returns Tribe members
(relationship_type 'tribe' or 'tribe-leadership') with role + days
since last touch.

Sort order is days-since-touch DESC (longest-overlooked first), so the
morning glance surfaces who needs attention. Note the CEO memory rule:
'ALL Tribe members - no cadence alerts; CEO talks daily, only track
recorded action items' - so days-since-touch is informational, not an
alert state. We DO NOT mark anyone stale or red based on touch age.
"""
import logging
import re
from datetime import date, datetime, timezone
from pathlib import Path

from scripts.utils.paths import get_data_root

logger = logging.getLogger(__name__)

TRIBE_TYPES = {"tribe", "tribe-leadership"}

# Phase 1.37: org roster spreadsheet. The Tribe page shows the full
# roster from this file, enriched with CRM data where a person matches.
TRIBE_ROSTER_XLSX = "datastore/operations/tribe/31C_Tribe.xlsx"

# Parses minimal YAML frontmatter (between --- markers at file start).
# We only need a few keys; full YAML parsing is overkill and adds a dep.
_FRONTMATTER_RE = re.compile(
    r"\A---\s*\n(.*?)\n---\s*\n",
    re.DOTALL,
)
_KEY_RE = re.compile(r"^([A-Za-z_][A-Za-z0-9_-]*):\s*(.*)$")

# The contact's display name is on a `# Name (...)` H1 line after the
# frontmatter. Example: "# Omar Said (misha-hanin)".
_H1_RE = re.compile(r"^#\s+(?P<name>.+?)(?:\s+\([^)]*\))?\s*$")


def _parse_frontmatter(text: str) -> dict:
    """Best-effort line-level frontmatter parser. Returns {} on no FM block."""
    m = _FRONTMATTER_RE.match(text)
    if not m:
        return {}
    result: dict = {}
    for line in m.group(1).splitlines():
        line = line.rstrip()
        if not line or line.startswith("#"):
            continue
        km = _KEY_RE.match(line)
        if not km:
            continue
        key = km.group(1).strip()
        val = km.group(2).strip()
        # Strip surrounding quotes if present.
        if (val.startswith('"') and val.endswith('"')) or (val.startswith("'") and val.endswith("'")):
            val = val[1:-1]
        result[key] = val
    return result


def _display_name(text: str, slug: str) -> str:
    """Find the H1 line right after the frontmatter and extract the name."""
    after_fm = _FRONTMATTER_RE.sub("", text, count=1)
    for line in after_fm.splitlines():
        if not line.strip():
            continue
        m = _H1_RE.match(line)
        if m:
            return m.group("name").strip()
        # Stop after the first non-empty line - if it's not an H1, we'll
        # fall back to the slug.
        break
    return slug.replace("-", " ").title()


def _days_since(date_str: str | None, today: date | None = None) -> int | None:
    """Parse YYYY-MM-DD and return days from then to today. None if unparseable."""
    if not date_str:
        return None
    try:
        d = date.fromisoformat(date_str)
    except ValueError:
        return None
    return ((today or date.today()) - d).days


def _norm_name(s: str | None) -> str:
    """Normalised name for matching: lowercased, whitespace-collapsed."""
    return " ".join((s or "").strip().lower().split())


def _load_tribe_roster(workspace_root: Path) -> list[dict]:
    """Read the org roster from datastore/operations/tribe/31C_Tribe.xlsx.

    Returns a list of {name, email, title, department, reports_to,
    telegram}. Returns [] on any failure (file missing, openpyxl not
    installed, unreadable) so the Tribe page degrades to CRM-only.
    """
    path = workspace_root / TRIBE_ROSTER_XLSX
    if not path.exists():
        return []
    try:
        import openpyxl
    except ImportError:
        return []
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
    except Exception:  # noqa: BLE001 - any load failure -> degrade to CRM-only
        return []
    try:
        sheet = "Tribe Roster"
        ws = wb[sheet] if sheet in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
    except Exception:  # noqa: BLE001 - any read failure -> degrade to CRM-only
        return []
    finally:
        try:
            wb.close()
        except Exception as exc:  # noqa: BLE001 - close is best-effort
            logger.debug("tribe: workbook close failed: %s", exc)

    # Locate the header row (the one carrying both 'name' and 'email').
    header = None
    header_idx = -1
    for i, r in enumerate(rows):
        cells = [str(c).strip().lower() if c is not None else "" for c in r]
        if "name" in cells and "email" in cells:
            header, header_idx = cells, i
            break
    if header is None:
        return []

    def col(*needles):
        for j, c in enumerate(header):
            if c in needles:
                return j
        for j, c in enumerate(header):
            if any(n in c for n in needles):
                return j
        return None

    ci = {
        "name": col("name"),
        "email": col("email"),
        "title": col("title"),
        "department": col("function / department", "department"),
        "reports_to": col("reports to"),
        "telegram": col("telegram"),
    }
    if ci["name"] is None:
        return []

    def cell(r, idx):
        if idx is None or idx >= len(r) or r[idx] is None:
            return ""
        return str(r[idx]).strip()

    roster = []
    for r in rows[header_idx + 1:]:
        if not r:
            continue
        name = cell(r, ci["name"])
        if not name:
            continue
        roster.append({
            "name": name,
            "email": cell(r, ci["email"]).lower(),
            "title": cell(r, ci["title"]),
            "department": cell(r, ci["department"]),
            "reports_to": cell(r, ci["reports_to"]),
            "telegram": cell(r, ci["telegram"]),
        })
    return roster


def _merge_tribe(crm_members: list[dict], roster: list[dict]) -> list[dict]:
    """Union the CRM tribe members with the xlsx roster.

    The roster drives the listing - every roster person appears, carrying
    org fields (title, department, reports_to, telegram). A roster person
    who is also a CRM tribe contact also carries the CRM fields (slug,
    role, last_touch, ...). CRM tribe members absent from the roster are
    kept too, so nothing is lost. Match is by email, then normalised name.
    """
    by_email: dict = {}
    by_name: dict = {}
    for m in crm_members:
        if m.get("email"):
            by_email[m["email"].lower()] = m
        by_name[_norm_name(m["name"])] = m

    merged = []
    matched_slugs = set()
    for rr in roster:
        crm = None
        email = (rr.get("email") or "").lower()
        if email and email in by_email:
            crm = by_email[email]
        elif _norm_name(rr["name"]) in by_name:
            crm = by_name[_norm_name(rr["name"])]
        if crm is not None:
            matched_slugs.add(crm["slug"])
        merged.append({
            "slug": crm["slug"] if crm else None,
            "name": (crm["name"] if crm else "") or rr["name"],
            "email": email or (crm.get("email", "") if crm else ""),
            "role": (crm["role"] if crm else None) or "tribe",
            "last_touch": crm["last_touch"] if crm else None,
            "days_since_touch": crm["days_since_touch"] if crm else None,
            "status": crm["status"] if crm else None,
            "title": rr.get("title", ""),
            "department": rr.get("department", ""),
            "reports_to": rr.get("reports_to", ""),
            "telegram": rr.get("telegram", ""),
            "in_roster": True,
        })

    # CRM tribe members not in the roster - keep them (no org fields).
    for m in crm_members:
        if m["slug"] in matched_slugs:
            continue
        merged.append({
            "slug": m["slug"], "name": m["name"], "email": m.get("email", ""),
            "role": m["role"], "last_touch": m["last_touch"],
            "days_since_touch": m["days_since_touch"], "status": m["status"],
            "title": "", "department": "", "reports_to": "", "telegram": "",
            "in_roster": False,
        })
    return merged


def list_tribe(workspace_root: Path, today: date | None = None,
               data_root: "Path | None" = None) -> dict:
    """Return the Tribe roster: the full org roster from 31C_Tribe.xlsx,
    enriched with CRM data (slug, role, last_touch, ...) where a person
    matches a crm/contacts/ tribe contact.

    Returns:
        {
            "members": [
                {
                    "slug": str or None,         # None for roster-only
                    "name": str,
                    "email": str,
                    "role": "tribe-leadership" | "tribe",
                    "last_touch": "YYYY-MM-DD" or None,
                    "days_since_touch": int or None,
                    "status": str or None,
                    "title": str, "department": str,
                    "reports_to": str, "telegram": str,
                    "in_roster": bool,
                },
                ...
            ] sorted by days_since_touch DESC (oldest first, None last),
            "counts": {"tribe-leadership": int, "tribe": int},
            "data_time": ISO 8601 UTC, most-recent of contact + xlsx mtime,
        }

    Degrades to CRM-only when the roster xlsx is missing/unreadable.

    HEADING OS engine/data split: crm/contacts/ and the roster xlsx are
    DATA, so they resolve under ``data_root`` (falls back to
    ``workspace_root`` when not supplied; identical on transitional ceo-main).
    """
    if data_root is None:
        data_root = get_data_root()
    contacts_dir = data_root / "crm" / "contacts"
    crm_members = []
    most_recent_mtime: float = 0.0
    if contacts_dir.exists():
        for p in contacts_dir.glob("*.md"):
            try:
                text = p.read_text(encoding="utf-8")
                mtime = p.stat().st_mtime
            except OSError:
                continue
            fm = _parse_frontmatter(text)
            rel_type = fm.get("relationship_type", "")
            if rel_type not in TRIBE_TYPES:
                continue
            slug = p.stem
            crm_members.append({
                "slug": slug,
                "name": _display_name(text, slug),
                "email": (fm.get("email") or "").strip(),
                "role": rel_type,
                "last_touch": fm.get("last_touch") or None,
                "days_since_touch": _days_since(fm.get("last_touch"), today=today),
                "status": fm.get("status") or None,
            })
            if mtime > most_recent_mtime:
                most_recent_mtime = mtime

    # Phase 1.37: enrich from / drive by the org roster xlsx (DATA).
    roster = _load_tribe_roster(data_root)
    if roster:
        try:
            xm = (data_root / TRIBE_ROSTER_XLSX).stat().st_mtime
            if xm > most_recent_mtime:
                most_recent_mtime = xm
        except OSError:
            pass
    members = _merge_tribe(crm_members, roster)

    # Sort by days_since_touch DESC (oldest first); None last.
    # Tuple: (none_flag, -days). none_flag=1 for None pushes them after
    # all real values (which carry none_flag=0).
    def sort_key(m):
        d = m["days_since_touch"]
        return (1 if d is None else 0, -(d or 0))
    members.sort(key=sort_key)

    counts: dict = {}
    for m in members:
        counts[m["role"]] = counts.get(m["role"], 0) + 1

    data_time = (
        datetime.fromtimestamp(most_recent_mtime, tz=timezone.utc).isoformat()
        if most_recent_mtime
        else None
    )
    return {"members": members, "counts": counts, "data_time": data_time}


# Contact file path safety: slug must be lowercase + hyphens.
CONTACT_SLUG_RE = re.compile(r"^[a-z0-9][a-z0-9-]{0,63}$")


def _extract_section(text: str, heading: str) -> str:
    """Extract a markdown H2 section by name (case-insensitive).

    Returns the section body (after the heading) up to the next H2 or EOF.
    Returns empty string if heading not found.
    """
    lines = text.splitlines()
    heading_lower = heading.lower().strip()
    in_section = False
    body_lines = []
    for line in lines:
        # Detect H2 heading.
        if line.startswith("## "):
            if in_section:
                # End of our section.
                break
            if line[3:].strip().lower() == heading_lower:
                in_section = True
                continue
        elif in_section:
            body_lines.append(line)
    return "\n".join(body_lines).strip()


def read_contact(workspace_root: Path, slug: str, data_root: "Path | None" = None) -> dict:
    """Read a CRM contact file safely. Returns parsed metadata + sections.

    Path safety:
    - slug must match CONTACT_SLUG_RE
    - file must be under crm/contacts/
    - must be a regular file, not a symlink
    - size <= 500 KB

    Returns:
        {"ok": True, "slug": str, "name": str, "frontmatter": dict,
         "active_commitments": str, "interaction_log": str, "size": int}
        OR {"ok": False, "error": str}

    HEADING OS engine/data split: crm/contacts/ is DATA, so it resolves
    under ``data_root`` (falls back to ``workspace_root`` when not supplied).
    """
    if data_root is None:
        data_root = get_data_root()
    if not slug or not isinstance(slug, str):
        return {"ok": False, "error": "missing slug"}
    if not CONTACT_SLUG_RE.match(slug):
        return {"ok": False, "error": "invalid slug"}
    contacts_dir = (data_root / "crm" / "contacts").resolve()
    target = (data_root / "crm" / "contacts" / f"{slug}.md").resolve()
    try:
        target.relative_to(contacts_dir)
    except ValueError:
        return {"ok": False, "error": "path escapes crm/contacts/"}
    if not target.exists():
        return {"ok": False, "error": "not found"}
    try:
        if target.is_symlink():
            return {"ok": False, "error": "symlinks not allowed"}
    except OSError:
        return {"ok": False, "error": "stat failed"}
    if not target.is_file():
        return {"ok": False, "error": "not a file"}
    try:
        size = target.stat().st_size
    except OSError:
        return {"ok": False, "error": "stat failed"}
    if size > 500_000:
        return {"ok": False, "error": f"file too large ({size} bytes)"}
    try:
        text = target.read_text(encoding="utf-8")
    except (OSError, UnicodeDecodeError) as e:
        return {"ok": False, "error": f"read failed: {e}"}
    frontmatter = _parse_frontmatter(text)
    name = _display_name(text, slug)
    return {
        "ok": True,
        "slug": slug,
        "name": name,
        "frontmatter": frontmatter,
        "active_commitments": _extract_section(text, "Active Commitments"),
        "interaction_log": _extract_section(text, "Interaction Log"),
        "size": size,
    }
