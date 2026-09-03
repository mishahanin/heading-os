#!/usr/bin/env python3
"""Tribe Fireside Bot - coordinates Mon + Wed firesides via Telegram.

Implementation per:
  - Spec: docs/superpowers/specs/2026-05-03-tribe-fireside-bot-design.md
  - Plan: docs/superpowers/plans/2026-05-03-tribe-fireside-bot-implementation.md (v1.3)
  - Operating model: runs on Misha's workstation, state in datastore/operations/tribe/fireside-state/

Subcommands. Every entry in `main()`'s `handlers` dict appears here, and
`tests/test_fireside_helmsman.py` fails when the two drift apart. Until
2026-09-02 this list ended at `init-state` and omitted nine implemented
commands, all of Phase 4 among them, while its own header claimed to state
"current implementation status" and so read as the authoritative surface.

  bootstrap                 - One-time: enumerate Telegram group, build initial roster
  poll                      - Process Telegram updates, every 5 min
  speaker-dms               - Send 2-week + 3-day speaker reminders
  sunday-preview            - Post pinned weekly preview to 31C Tribe
  dayof-reminders           - DM speakers Zoom link 3h before session
  helmsman-brief            - Brief next week's Helmsman 7 days ahead
  helmsman set|list|gaps    - Assign / list / audit Helmsman coverage
  speaker-gaps              - List members with no slot this cycle; exits 1 if any
  weekly-discrepancy-report - Report Telegram-vs-xlsx mismatches
  email-backup              - Email reminder for unresponsive Tribe
  stats                     - Generate stats markdown report
  health-check              - Alert if poll hasn't run in 30 min
  unpin-weekly              - Unpin Sunday preview after Wed session
  log-session               - Log session result, manual command
  topic-nudge               - Nudge the Tribe for topic ideas mid-cycle
  topic-digest              - Post the running topic backlog to the group
  cycle-end-invite          - Draft the cycle-end topic invite to the CEO for approval
  cycle-rollover            - Roll the schedule into the next cycle
  topic-ideas               - List the ideas collected for a cycle
  set-webhook               - Register the Telegram webhook URL
  delete-webhook            - Remove the Telegram webhook registration
  webhook-info              - Print Telegram's view of the current webhook
  heartbeat                 - Liveness ping for the daemon's healthcheck
  test-telegram             - Smoke test: send DM to Misha
  xlsx-check                - Print xlsx loader summary
  init-state                - Initialise state directory + files

Usage:
  python scripts/fireside-bot.py <subcommand> [args]
  python scripts/fireside-bot.py --help

Tests: tests/test_a_sweep_that_reported_the_letters_it_never_read.py

Tests: tests/test_a_promise_that_misha_would_help.py
"""

import io
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

# Force UTF-8 stdout/stderr on Windows so emoji and non-ASCII names print correctly.
# Guard against pythonw.exe where sys.stdout/stderr are None (no console attached).
if sys.platform == "win32":
    if sys.stdout is not None and hasattr(sys.stdout, "buffer"):
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    if sys.stderr is not None and hasattr(sys.stderr, "buffer"):
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding="utf-8", errors="replace")

import argparse
import contextlib
import json
import os
import re
import socket
import tempfile
from datetime import datetime
from typing import Any, Optional
from zoneinfo import ZoneInfo

import requests
import urllib3.util.connection as _urllib3_connection

# Force IPv4 for all outbound Telegram calls. The service host
# has an AAAA record for api.telegram.org but no working IPv6 route - urllib3
# picks the IPv6 result first and stalls 30s before falling back to IPv4. On
# 2026-05-25 every webhook handler was taking 30s+ per sendMessage because of
# this. Forcing AF_INET via urllib3's allowed_gai_family hook cuts the call
# from 30.09s to 0.07s (measured on the VM).
#
# NOT a no-op on a host where IPv6 works, which this comment used to claim. The
# hook is PROCESS-GLOBAL urllib3 state: every urllib3 user in the interpreter,
# including any future `requests` caller that has nothing to do with Telegram,
# resolves IPv4-only from here on. That is a real constraint accepted for a
# 400x latency win on the one host this bot runs on, not an absence of one.
# Anything that needs IPv6 in this process has to undo it explicitly.
_urllib3_connection.allowed_gai_family = lambda: socket.AF_INET

from scripts.utils.colors import GREEN, YELLOW, RED, GRAY, CYAN, BOLD, RESET
from scripts.utils.healthchecks import ping as hc_ping
from scripts.utils.operator_identity import admin_email, get_operator
from scripts.utils.telegram_bot import TelegramAPIError, TelegramBot
from scripts.utils.workspace import get_datastore_dir, get_default_tz, get_default_tz_name, get_outputs_dir, get_workspace_root, load_env, resolve_config_with_example
from scripts import fireside_topics as ft

# ============================================================
# Configuration
# ============================================================

WORKSPACE_ROOT = get_workspace_root()


def state_dir() -> Path:
    """Resolved at call time, never at import.

    `get_datastore_dir()` reads `HEADING_OS_DATA` on every call, so it follows
    the environment for a caller that asks after the environment moved. As a
    module-level constant it asked once, during its own import, and stored the
    answer, so a test that imported this module and then repointed the root
    still wrote into the operator's real overlay. The `mkdir` in
    `require_writable_state_dir` is not among the primitives `tests/conftest.py`
    wraps, so that write drew no refusal.
    """
    return get_datastore_dir() / "operations" / "tribe" / "fireside-state"


def stats_dir() -> Path:
    return get_outputs_dir() / "operations" / "tribe-fireside" / "stats"


def tribe_xlsx() -> Path:
    return get_datastore_dir() / "operations" / "tribe" / "31C_Tribe.xlsx"


TELEGRAM_API_BASE = "https://api.telegram.org"

# State file names (relative to state_dir())
TRIBE_ROSTER = "tribe-roster.json"
SCHEDULE = "schedule.json"
HELMSMEN = "helmsmen.json"
OPT_INS = "opt-ins.json"
DM_LOG = "dm-log.jsonl"
SESSIONS_LOG = "sessions.jsonl"
LAST_UPDATE_ID = "last-update-id.json"
LAST_PINNED = "last-pinned.json"
ERRORS_LOG = "errors.log"
EXCLUSIONS = "exclusions.json"  # CEO-managed list of Tribe members excluded from fireside rotations
OUTSIDER_RATE = "outsider-forward-rate.json"  # rate-limit state for outsider DM forwards to Misha
SWAP_REQUESTS_LOG = "swap-requests.jsonl"  # append-only event log for /swap state machine
TOPIC_IDEAS = "topic-ideas.jsonl"  # append-only topic backlog (see fireside_topics)
TOPIC_STATE = "topic-collection-state.json"  # digest cursor + pending cycle invite

# /swap interactive flow tuning
SWAP_HORIZON_WEEKS = 4  # how far ahead to scan for candidate sessions
SWAP_B_RESPONSE_TTL_HOURS = 24  # how long B has to accept/decline before request expires
SWAP_CANDIDATES_LIMIT = 2  # how many buttons to show A

# Senior-leader title fragments for VP detection in xlsx.
#
# "vice president" is spelled out in the sheet more often than it is abbreviated,
# and until 2026-08-25 only the abbreviations were listed: "VP of Engineering"
# matched, "Vice President of Engineering" and "Senior Vice President" did not.
# The fragment carries a space, so it covers the senior/executive prefixes too.
VP_TITLE_FRAGMENTS = (
    "ceo", "cfo", "cto", "csto", "cso", "cmo", "chro", "cio", "clo",
    "chief ", "vp ", "svp ", "vp,", "svp,", "vice president",
    "founder", "co-founder",
)


def _vp_title_pattern(fragments=VP_TITLE_FRAGMENTS):
    """One case-insensitive pattern, each fragment held to word boundaries.

    The fragments were matched as bare substrings, which is safe for the ones
    carrying a space or a comma and wrong for the nine bare trigrams: "cio"
    matched "precious" and "ex officio", "clo" matched "clothing", "cso" matched
    any compound containing those three letters. A member titled "Clothing and
    Apparel Lead" came back is_vp. The 2026-08-25 pass on this list added the
    spelled-out "vice president" and never touched the boundary problem.

    A boundary is only added on a side where the fragment itself ends in a word
    character, so "vp " keeps its trailing space as the separator it was written
    to be and "co-founder" is not split at its hyphen.
    """
    parts = []
    for frag in fragments:
        lead = r"\b" if frag[:1].isalnum() else ""
        tail = r"\b" if frag[-1:].isalnum() else ""
        parts.append(f"{lead}{re.escape(frag)}{tail}")
    return re.compile("|".join(parts), re.IGNORECASE)


_VP_TITLE_RE = _vp_title_pattern()


def is_vp_title(title: str) -> bool:
    """True when a job title names a senior leader."""
    return bool(_VP_TITLE_RE.search(title or ""))

# Cycle-1 speaker schedule is per-instance DATA: real names, themes, and the cycle
# start date live in the data overlay at <data-root>/config/fireside-schedule.json
# (resolved via get_data_config_dir()). The engine ships
# scripts/fireside-schedule.example.json as the generic template/fallback, so a
# data-less clone bootstraps cleanly. Speakers are identified by full name (the
# join to telegram_username happens at bootstrap).
def _fireside_schedule_file() -> Path:
    return resolve_config_with_example(
        "fireside-schedule.json", WORKSPACE_ROOT / "scripts" / "fireside-schedule.example.json"
    )


# Guarded, and non-fatal. A missing or malformed `fireside-schedule.json` used
# to kill every invocation of the file before argparse ran -- `--help` included,
# and every subcommand that never touches the cycle config. An operator whose
# config was broken got a traceback instead of the usage text that would have
# told them which file to fix. The accessors fall back to empty and the commands
# that need them fail with a sentence.
def _fireside_config() -> tuple:
    """(cycle_1_start_monday, weeks, error) read from disk on every call.

    This parse used to run at IMPORT, off a path `resolve_config_with_example()`
    had already frozen at import too, so a test that repointed the data root
    afterwards still read the operator's real config.
    """
    path = _fireside_schedule_file()
    try:
        schedule = json.loads(path.read_text(encoding="utf-8"))
        return (datetime.fromisoformat(schedule["cycle_1_start_monday"]).date(),
                schedule["weeks"], None)
    except (OSError, ValueError, KeyError, TypeError) as exc:
        return None, [], f"{path}: {type(exc).__name__}: {exc}"


def cycle_1_start_monday():
    return _fireside_config()[0]


def week_1_to_9_schedule():
    return _fireside_config()[1]


def _fireside_config_error():
    return _fireside_config()[2]


def require_fireside_config() -> None:
    """Refuse a command that needs the cycle config when it did not load.

    Called by the subcommands that read `cycle_1_start_monday()` or
    `week_1_to_9_schedule()`. Everything else -- `--help`, `health-check`,
    `xlsx-check` -- runs without it, which is the point of not raising at import.
    """
    error = _fireside_config_error()
    if error:
        raise SystemExit(
            f"fireside schedule config could not be read: {error}"
        )


def _load_fireside_config_fresh() -> tuple:
    """Re-read the cycle config from disk, returning (start_monday, weeks).

    Distinct from `_fireside_config()` above only in raising rather than
    reporting: cycle rollover wants the failure, not an empty schedule it would
    silently roll over on top of.
    """
    path = _fireside_schedule_file()
    cfg = json.loads(path.read_text(encoding="utf-8"))
    start = datetime.fromisoformat(cfg["cycle_1_start_monday"]).date()
    return start, cfg["weeks"]


# ============================================================
# Time helper (the configured timezone)
# ============================================================

def local_now() -> datetime:
    """Return current time as a timezone-aware datetime in the configured timezone."""
    return datetime.now(get_default_tz())


# ============================================================
# State file helpers (atomic writes, JSONL append, error log)
# ============================================================

def state_path(filename: str) -> Path:
    """Return absolute path to a state file under `state_dir()`."""
    return state_dir() / filename


def require_writable_state_dir() -> Path:
    """`state_dir()`, but only when a private data overlay actually backs it.

    Operator law, 2026-08-26: no data from the DATA repository may ever sit in
    the engine. `state_dir()` resolves through `get_datastore_dir()`,
    and with no overlay `get_data_root()` falls to its documented last resort
    `<workspace_root>/examples`. So on a public clone every writer below resolved
    to `examples/datastore/operations/tribe/fireside-state/` inside the clone and
    its `mkdir(parents=True)` created the tree. Measured that day on a worktree
    with no sibling overlay: one suite run left `errors.log` there, inside the
    repository that gets pushed.

    One funnel for the five writers, on purpose. Four of them had their own
    `path.parent.mkdir(parents=True, exist_ok=True)` line, and a guard added to
    some of them is a guard the next writer will not have.

    Calls `state_dir()` fresh on every call rather than closing over one
    import-time value, so a caller that redirects the resolver is honoured. The
    whole fireside suite does exactly that, and an earlier version of this guard
    asked whether an overlay existed instead of where the write was going: it
    refused fifty writes aimed at a `tmp_path` and never near the clone.

    Raises rather than redirects: fireside state is the bot's memory of who is in
    the Tribe and which session ran, so a write that silently lands somewhere
    else is worse than one that fails loudly. `log_error` is the single exception
    and says why at its own site.
    """
    from scripts.utils.paths import require_outside_engine_clone

    require_outside_engine_clone(state_dir(), "the fireside state directory")
    state_dir().mkdir(parents=True, exist_ok=True)
    return state_dir()


def require_writable_stats_dir() -> Path:
    """`stats_dir()`, under the same law as the state funnel above.

    `stats_dir()` resolves through `get_outputs_dir()`, which reaches
    `get_data_root()` the same way `state_dir()` does and falls to
    `<workspace_root>/examples` when no overlay backs it. `cmd_stats` carried
    the bare `mkdir(parents=True, exist_ok=True)` that the funnel above exists
    to replace, so `python scripts/fireside-bot.py stats` on a clone with no
    overlay created `examples/outputs/operations/tribe-fireside/stats/` inside
    the repository that gets pushed - and on a host where the overlay backs the
    state but not the outputs, it wrote real Tribe speaker names into it.

    The docstring above says why this is a second funnel and not a fifth inline
    guard: "a guard added to some of them is a guard the next writer will not
    have". `cmd_stats` was the next writer.
    """
    from scripts.utils.paths import require_outside_engine_clone

    require_outside_engine_clone(stats_dir(), "the fireside stats directory")
    stats_dir().mkdir(parents=True, exist_ok=True)
    return stats_dir()


def _unreadable_sheet_errors() -> tuple:
    """Every exception a file that is not a readable workbook can raise.

    openpyxl is a heavy local import everywhere else in this file, and it is
    optional on a host that never touches the sheet, so the concrete classes are
    resolved lazily and the tuple degrades to the two the reader itself raises.
    """
    import zipfile
    errors: list = [ValueError, zipfile.BadZipFile]
    try:
        from openpyxl.utils.exceptions import InvalidFileException
    except ImportError:
        pass
    else:
        errors.append(InvalidFileException)
    return tuple(errors)


_UNREADABLE_SHEET = _unreadable_sheet_errors()


def ensure_state_dir() -> None:
    """Create state directory and initialise empty state files if missing.

    Files initialised (only if they don't exist):
      - tribe-roster.json: rebuilt from 31C_Tribe.xlsx if xlsx is reachable,
        else an empty {} placeholder. Without this self-heal a VM rebuild or
        state-loss leaves the bot rejecting every DM as outsider.
      - Other JSON files with sensible empty defaults
      - JSONL files as empty text files
      - errors.log as empty text file
    """
    require_writable_state_dir()

    # Self-heal tribe-roster.json from xlsx (the source of truth for membership).
    # load_tribe_metadata already returns entries shaped like roster records
    # (active=True, telegram_user_id=None). user_ids are NOT populated here:
    # cross_reference is skipped because it only emits the intersection with the
    # live Telegram membership, which Telethon would have to enumerate. After a
    # state-loss rebuild every entry is unbound, so the operator must re-run the
    # trusted `bootstrap` to bind user_ids. DM handlers deliberately refuse to
    # bind a user_id from a self-reported username (handle-takeover guard), so
    # the bot rejects DMs from unbound members (forwarding them to Misha) until
    # bootstrap runs -- a safe failure, not a silent self-enrollment hole.
    if not state_path(TRIBE_ROSTER).exists():
        try:
            xlsx_roster = load_tribe_metadata()
            exclusions = load_state(EXCLUSIONS) or {}
            excluded = {k.lower(): v for k, v in exclusions.get("excluded", {}).items()}
            roster: dict = {}
            for username, data in xlsx_roster.items():
                entry = dict(data)
                if username.lower() in excluded:
                    entry["active"] = False
                    entry["excluded_from_fireside"] = True
                    entry["exclusion_reason"] = excluded[username.lower()].get("reason", "")
                    entry["excluded_at"] = excluded[username.lower()].get("excluded_at", "")
                roster[username] = entry
            save_state(TRIBE_ROSTER, roster)
        except FileNotFoundError as exc:
            # No sheet to heal from. An empty placeholder is the honest state.
            log_error(f"self-heal wrote an EMPTY tribe-roster.json: {exc}")
            print(f"{YELLOW}init-state: tribe-roster.json is EMPTY - {tribe_xlsx()} "
                  f"was not found. Every DM will be refused as an outsider until "
                  f"`bootstrap` runs.{RESET}", file=sys.stderr)
            save_state(TRIBE_ROSTER, {})
        except _UNREADABLE_SHEET as exc:
            # `load_tribe_metadata` raises ValueError ONLY for a sheet the
            # operator has to fix -- a Telegram handle claimed by two members, or
            # a missing column. Swallowing it wrote `{}` in silence, and an empty
            # roster is precisely the state this self-heal exists to prevent: the
            # bot then refuses every DM as an outsider, exactly as it would with
            # no file at all, while `init-state` printed OK and listed the file's
            # size.
            #
            # The tuple also names what OPENPYXL raises from under it. The
            # handler list had been written from what `load_tribe_metadata`
            # raises deliberately, and a file that is not a workbook never
            # reaches that code: `zipfile.BadZipFile` for a truncated or 0-byte
            # sync, `InvalidFileException` for an HTML error page saved over the
            # sheet. Both are plain `Exception` subclasses, NOT OSError and NOT
            # ValueError, so they have to be named. Each one crashed
            # `init-state` with a traceback instead of writing the placeholder
            # and the warning below -- during exactly the VM rebuild or state
            # loss this self-heal exists for.
            log_error(f"self-heal wrote an EMPTY tribe-roster.json: {exc}")
            print(f"{RED}init-state: tribe-roster.json is EMPTY because the Tribe "
                  f"sheet could not be read: {exc}{RESET}", file=sys.stderr)
            print(f"{YELLOW}      Until that is fixed and `bootstrap` re-run, the "
                  f"bot refuses every DM as an outsider.{RESET}", file=sys.stderr)
            save_state(TRIBE_ROSTER, {})

    initial: dict[str, Any] = {
        SCHEDULE: [],
        HELMSMEN: {},
        OPT_INS: {"helmsman": [], "wildcard": []},
        LAST_UPDATE_ID: {"offset": 0},
        LAST_PINNED: {"message_id": None},
    }
    text_files = [DM_LOG, SESSIONS_LOG, ERRORS_LOG]

    for name, default in initial.items():
        path = state_path(name)
        if path.exists():
            continue
        save_state(name, default)

    for name in text_files:
        path = state_path(name)
        if not path.exists():
            path.write_text("", encoding="utf-8")

    # Topic-collection files (see fireside_topics). Empty backlog + default state.
    topic_ideas = state_path(TOPIC_IDEAS)
    if not topic_ideas.exists():
        topic_ideas.write_text("", encoding="utf-8")
    if not state_path(TOPIC_STATE).exists():
        save_state(TOPIC_STATE, {"last_digest_idea_id": None, "pending_cycle_invite": None})


def load_state(filename: str) -> Any:
    """Load a JSON state file. Returns None if file does not exist."""
    path = state_path(filename)
    if not path.exists():
        return None
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def save_state(filename: str, data: Any) -> None:
    """Atomically write a JSON state file (write-to-tmp + os.replace).

    Prevents corruption on crash mid-write. The temp file lives in the same
    directory as the target so os.replace is atomic on Windows + POSIX.
    """
    require_writable_state_dir()
    path = state_path(filename)
    tmp_fd, tmp_path = tempfile.mkstemp(prefix=path.name + ".", dir=path.parent, suffix=".tmp")
    try:
        with os.fdopen(tmp_fd, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)
            f.write("\n")
        os.replace(tmp_path, path)
    except Exception:
        if os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except OSError:
                pass
        raise


def append_jsonl(filename: str, event: dict) -> None:
    """Append one JSON event as a single line to a JSONL file."""
    require_writable_state_dir()
    path = state_path(filename)
    with open(path, "a", encoding="utf-8") as f:
        f.write(json.dumps(event, ensure_ascii=False) + "\n")


@contextlib.contextmanager
def locked_state(filename: str, default):
    """Read-modify-write ONE fireside state file under a cross-process lock.

    Yields the loaded value; whatever the block leaves in it is saved when the
    block exits without raising. On an exception nothing is written.

    `save_state` makes each WRITE atomic, which is a different guarantee from
    making a read and its following write atomic. Cron subcommands and update
    handlers both load-modify-save `schedule.json` and `helmsmen.json`, and
    overlapping runs simply lost the earlier change: `helmsman set` writes a
    record while `helmsman-brief` rewrites the dict it loaded before that, and
    two swap callbacks can both pass their status checks before either terminal
    event is on disk.

    NOT the shared `checkpoint_paths.locked_state`: that one reads through
    `read_json` and hands back a dict, and the fireside schedule is a LIST.
    Same lock primitive underneath, different state shape on top.

    Bounded, never blocking, exactly like the primitive it wraps: on expiry the
    block runs UNLOCKED with a line on stderr, because a Telegram handler that
    hangs is worse than one that races.
    """
    from scripts.utils.checkpoint_paths import file_lock

    require_writable_state_dir()
    path = state_path(filename)
    with file_lock(path.with_name(path.name + ".lock"), label="fireside"):
        value = load_state(filename)
        if value is None:
            value = default
        yield value
        save_state(filename, value)


def misha_user_id() -> int:
    """The operator's Telegram id from the environment, or 0 when unusable.

    Sixteen call sites did `misha_user_id()`
    raw. A typo in that variable therefore raised ValueError out of whatever was
    running -- a cron subcommand mid-job, or an update handler after its side
    effects had already landed -- instead of behaving like the "not configured"
    case every one of those sites already handles by checking for 0.
    """
    raw = os.environ.get("MISHA_TELEGRAM_USER_ID", "0")
    try:
        return int(raw)
    except (TypeError, ValueError):
        log_error(f"MISHA_TELEGRAM_USER_ID is {raw!r}, not an integer; "
                  f"treating the operator as unconfigured")
        return 0


def _read_jsonl_rows(path: Path) -> list[dict]:
    """Every parseable JSON object in a JSONL file. Missing file -> empty list.

    One reader, so a caller cannot forget the `exists()` guard the way
    `cmd_email_backup` did on its second pass over dm-log.jsonl.

    Decoding is per LINE, through `jsonl_lines`, and that settles two separate
    defects in the one expression this used to be.

    `read_text(encoding="utf-8")` raises `UnicodeDecodeError` on undecodable
    bytes, and that is a `ValueError`: neither an `OSError` nor the
    `json.JSONDecodeError` the per-line handler below catches. So one torn
    append to `dm-log.jsonl` or `sessions.jsonl` raised out of here, out of
    `cmd_email_backup`, and no speaker got a backup email. Widening the
    whole-file read to `except UnicodeError` stopped the raise by returning the
    file EMPTY, which trades a crash for losing every intact row -- and this
    reader's own contract is that one bad LINE costs one row.

    `str.splitlines()` was the second. It breaks on eight characters a JSONL
    record does not end at, three of which (U+0085, U+2028, U+2029)
    `append_jsonl` writes raw because it uses `ensure_ascii=False`. A record
    carrying one was written as a single valid line and then shredded into
    halves that no longer parsed, and the `JSONDecodeError` clause dropped both
    in silence. MEASURED 2026-09-01: an `idea_submitted` row whose text held one
    U+2028 (what a browser paste produces for a line separator) vanished from
    `responded_user_ids` in `cmd_email_backup`, so the member who HAD answered
    was mailed "haven't seen a response". `bytes.splitlines()` breaks on `\\n`
    and `\\r` only.
    """
    from scripts.utils.jsonl_lines import jsonl_lines

    rows: list[dict] = []
    if not path.exists():
        return rows
    try:
        lines = list(jsonl_lines(path))
    except OSError:
        log_error(f"could not read {path}")
        return rows
    undecodable = 0
    for line in lines:
        if line is None:
            undecodable += 1
            continue
        try:
            row = json.loads(line)
        except json.JSONDecodeError:
            continue
        if isinstance(row, dict):
            rows.append(row)
    if undecodable:
        log_error(f"{path}: skipped {undecodable} undecodable line(s)")
    return rows


def log_error(message: str, exception: Optional[BaseException] = None) -> None:
    """Append an error line to errors.log with ISO-8601 local timestamp.

    The one writer that must not raise. `main()` is wrapped so an uncaught
    exception lands here, so a refusal thrown from this function replaces the
    error being reported with a second error and the first one is lost. It goes
    to stderr instead, where a systemd unit still captures it.
    """
    from scripts.utils.paths import DataRootError

    ts = local_now().isoformat()
    if exception is not None:
        line = f"[{ts}] ERROR: {message} [{type(exception).__name__}: {exception}]\n"
    else:
        line = f"[{ts}] ERROR: {message}\n"
    try:
        require_writable_state_dir()
    except DataRootError as exc:
        print(f"fireside: {exc}; error not persisted", file=sys.stderr)
        print(line, end="", file=sys.stderr)
        return
    with open(state_path(ERRORS_LOG), "a", encoding="utf-8") as f:
        f.write(line)


# ============================================================
# Telegram Bot API wrapper (raw HTTPS via requests)
# ============================================================
#
# TelegramAPIError / TelegramBot live in scripts/utils/telegram_bot.py
# (imported above) so the dedicated system-notifications bot
# (scripts/utils/telegram_notify.py) can reuse the same production-hardened
# wrapper. get_bot() below wires Fireside's own error logging into it via
# the on_error callback, so behavior is unchanged from before the extraction.


def get_bot() -> TelegramBot:
    """Construct the bot from FIRESIDE_BOT_TOKEN in env."""
    load_env()
    token = os.environ.get("FIRESIDE_BOT_TOKEN")
    if not token:
        print(f"{RED}ERROR: FIRESIDE_BOT_TOKEN not set in .env{RESET}", file=sys.stderr)
        sys.exit(1)
    return TelegramBot(token, on_error=log_error)


# ============================================================
# xlsx reader
# ============================================================

def _first_present(headers: dict, *names: str):
    """The index of the first header name present, or None.

    Membership, not truthiness. `headers.get(a) or headers.get(b)` reads
    correctly right up until the preferred column sits at index 0.
    """
    for name in names:
        if name in headers:
            return headers[name]
    return None


def load_tribe_metadata() -> dict:
    """Read 31C_Tribe.xlsx and return metadata keyed by telegram_username.

    Returns:
        dict[telegram_username -> dict with keys: name, email, title, function,
        is_vp, languages, telegram_user_id, active]

    Notes:
        - Rows without a Telegram Username value are skipped silently here;
          the weekly-discrepancy-report subcommand surfaces them.
        - is_vp is heuristically derived from the Title (reconciled) column
          using VP_TITLE_FRAGMENTS.
        - telegram_user_id is initialised to None; populated later only via the
          trusted Telethon bootstrap (Phase 2 task 2.2). DM handlers never bind
          it from a self-reported username (handle-takeover guard).
    """
    import openpyxl  # local import - openpyxl is heavy

    if not tribe_xlsx().exists():
        msg = f"Tribe xlsx not found at {tribe_xlsx()}"
        log_error(msg)
        raise FileNotFoundError(msg)

    wb = openpyxl.load_workbook(tribe_xlsx(), data_only=True)
    ws = wb.active

    # Find header row by looking for "Name" column
    header_row_idx = None
    headers: dict[str, int] = {}
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if not row:
            continue
        for idx, cell in enumerate(row):
            if cell == "Name":
                header_row_idx = i
                headers = {str(c).strip(): j for j, c in enumerate(row) if c}
                break
        if header_row_idx is not None:
            break

    if header_row_idx is None:
        raise ValueError(f"Could not find header row (no 'Name' column) in {tribe_xlsx()}")

    name_col = headers.get("Name")
    email_col = headers.get("Email")
    tg_username_col = headers.get("Telegram Username")
    # `_first_present`, not `or`: these are COLUMN INDEXES and index 0 is falsy,
    # so a sheet whose first column is "Title (reconciled)" fell through to
    # "Title" and then to None. The preferred column being first is the normal
    # case, and it was the one that silently lost its metadata.
    title_col = _first_present(headers, "Title (reconciled)", "Title")
    function_col = _first_present(headers, "Function / Department", "Function")

    if name_col is None:
        raise ValueError(f"'Name' column not found in {tribe_xlsx()}")

    if tg_username_col is None:
        # Friendly error - this is Phase 0 task 0.5 (Misha-side)
        raise ValueError(
            f"'Telegram Username' column not found in {tribe_xlsx()}. "
            f"Add this column and populate it for all 54 Tribe members "
            f"(Phase 0 task 0.5) before running this command."
        )

    roster: dict[str, dict] = {}
    # Duplicate Telegram usernames used to overwrite in silence, so the roster
    # came back SHORTER than the sheet and nothing said which member had gone.
    # A membership count that quietly drops is the worst shape for this file:
    # every downstream count, schedule and discrepancy report agrees with it.
    seen_usernames: dict[str, str] = {}
    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or row[name_col] is None:
            continue
        name = str(row[name_col]).strip()
        if not name:
            continue

        username_raw = row[tg_username_col]
        if not username_raw:
            # Skip rows without telegram_username; discrepancy report will surface them
            continue
        username = str(username_raw).strip().lstrip("@")
        if not username:
            continue

        title = ""
        if title_col is not None and row[title_col] is not None:
            title = str(row[title_col]).strip()

        function = "Unknown"
        if function_col is not None and row[function_col] is not None:
            function = str(row[function_col]).strip()

        email = ""
        if email_col is not None and row[email_col] is not None:
            email = str(row[email_col]).strip()

        is_vp = is_vp_title(title)

        if username.lower() in seen_usernames:
            raise ValueError(
                f"{tribe_xlsx()}: Telegram username @{username} appears twice "
                f"({seen_usernames[username.lower()]} and {name}). One username "
                f"is one member; fix the sheet before the roster is rebuilt."
            )
        seen_usernames[username.lower()] = name

        roster[username] = {
            "name": name,
            "email": email,
            "title": title,
            "function": function,
            "is_vp": is_vp,
            "languages": ["en"],  # default; refine in v2 if needed
            "telegram_user_id": None,  # populated only via trusted Telethon bootstrap
            "active": True,
        }

    return roster


# ============================================================
# Schedule generator (Phase 2 task 2.3)
# ============================================================

def next_cycle_number(schedule: list) -> int:
    """The cycle a rebuild of `schedule` should carry: one past the highest in it.

    An entry with no `cycle` field is read as cycle 1, matching
    `fireside_topics.current_cycle`, so a schedule written before the field
    existed rolls over to 2 rather than to 1.
    """
    if not schedule:
        return 1
    return max(int(e.get("cycle", 1) or 1) for e in schedule) + 1


def build_schedule(roster_by_name: dict, start_monday=None,
                   weeks=None, cycle: int = 1) -> tuple[list, list[str]]:
    """Convert a week calendar into the schedule.json structure.

    Args:
        roster_by_name: dict[full_name -> {telegram_username, ...}] for username lookup.
        start_monday: Week-1 Monday date. Defaults to `cycle_1_start_monday()`.
            Pass a fresh value when a long-running daemon rebuilds from an
            updated config file.
        weeks: the week calendar (list of {week, theme, mon, wed}). Defaults to
            `week_1_to_9_schedule()`.
        cycle: the cycle number stamped on every entry. Until 2026-08-23 this
            was hardcoded to 1, so `cmd_cycle_rollover` produced a cycle-2
            calendar that called itself cycle 1 -- and since
            `fireside_topics.current_cycle` reads this field, every later
            `/idea` was filed under cycle 1, the cycle-end backlog never
            emptied, and `cycle_end_invite`'s idempotency key kept matching the
            invite already sent for cycle 1.

    Returns:
        (schedule_entries, missing_speakers) where:
          - schedule_entries: list of one dict per (session, slot) — 18 sessions x 3 slots = 54 entries
          - missing_speakers: list of names from the schedule that have no roster match
    """
    from datetime import timedelta

    if start_monday is None or weeks is None:
        # The config accessors are the fallback and they are empty when the
        # cycle config failed to load. Say which file, once, instead of building
        # a zero-week schedule and reporting success.
        require_fireside_config()
    if start_monday is None:
        start_monday = cycle_1_start_monday()
    if weeks is None:
        weeks = week_1_to_9_schedule()

    entries = []
    missing = []
    for week_data in weeks:
        week_num = week_data["week"]
        theme = week_data["theme"]
        # Mon = start_monday + (week-1)*7 days; Wed = Mon + 2
        mon_date = start_monday + timedelta(days=(week_num - 1) * 7)
        wed_date = mon_date + timedelta(days=2)

        for day_label, day_date, speaker_names in [
            ("Mon", mon_date, week_data["mon"]),
            ("Wed", wed_date, week_data["wed"]),
        ]:
            for slot_idx, name in enumerate(speaker_names, start=1):
                entry = roster_by_name.get(name)
                username = entry["telegram_username"] if entry else None
                if username is None:
                    missing.append(name)
                entries.append({
                    "cycle": cycle,
                    "week": week_num,
                    "session_date": day_date.isoformat(),
                    "day": day_label,
                    "theme": theme,
                    "slot": slot_idx,
                    "speaker_name": name,
                    "speaker_username": username,
                    "swapped_with": None,
                    "no_show": False,
                    "completed": False,
                })
    return entries, missing


# ============================================================
# Telethon-based bootstrap (Phase 2 tasks 2.1, 2.2, 2.5)
# ============================================================

def _telethon_session_path() -> Path:
    """Return path to the existing /telegram skill's Telethon session."""
    return WORKSPACE_ROOT / ".sessions" / "telegram" / "telegram"


def _telethon_credentials() -> tuple[int, str, str]:
    """Load TELEGRAM_API_ID, TELEGRAM_API_HASH, TELEGRAM_PHONE from env."""
    load_env()
    api_id_str = os.environ.get("TELEGRAM_API_ID")
    api_hash = os.environ.get("TELEGRAM_API_HASH")
    phone = os.environ.get("TELEGRAM_PHONE")
    missing = [k for k, v in
               [("TELEGRAM_API_ID", api_id_str), ("TELEGRAM_API_HASH", api_hash),
                ("TELEGRAM_PHONE", phone)] if not v]
    if missing:
        raise RuntimeError(f"Missing Telethon env vars: {', '.join(missing)}")
    try:
        api_id = int(api_id_str)
    except ValueError:
        raise RuntimeError(f"TELEGRAM_API_ID must be integer, got {api_id_str!r}")
    return api_id, api_hash, phone


async def _enumerate_tribe_members(client, chat_id: int) -> dict:
    """Enumerate participants of the 31C Tribe group via Telethon.

    Returns:
        dict[telegram_username -> {user_id, first_name, last_name, full_name}]
        plus a special key '_no_username' which is a list of dicts for users
        who have no username set on their Telegram account.
    """
    members_by_username: dict = {"_no_username": []}
    chat = await client.get_entity(chat_id)
    async for user in client.iter_participants(chat):
        first = (user.first_name or "").strip()
        last = (user.last_name or "").strip()
        full_name = f"{first} {last}".strip()
        record = {
            "user_id": user.id,
            "first_name": first,
            "last_name": last,
            "full_name": full_name,
            "is_bot": bool(user.bot),
        }
        if user.username:
            members_by_username[user.username.lower()] = record
        else:
            members_by_username["_no_username"].append(record)
    return members_by_username


async def _read_launch_reactions(client, chat_id: int, msg_id: int) -> dict:
    """Read 🧭 (helmsman) and 🌟 (wildcard) reactions on the launch announcement.

    Returns:
        dict with keys 'helmsman' and 'wildcard', each a list of dicts
        {user_id, username (or None)}. If the message can't be read, returns
        empty lists and logs the error.
    """
    from telethon.tl.functions.messages import GetMessageReactionsListRequest
    from telethon.tl.types import ReactionEmoji
    from telethon import errors as terrors

    out = {"helmsman": [], "wildcard": []}
    chat = await client.get_entity(chat_id)

    for emoji, key in [("🧭", "helmsman"), ("🌟", "wildcard")]:
        try:
            offset = ""
            collected = 0
            while True:
                result = await client(GetMessageReactionsListRequest(
                    peer=chat,
                    id=msg_id,
                    reaction=ReactionEmoji(emoticon=emoji),
                    offset=offset,
                    limit=100,
                ))
                for reaction in result.reactions:
                    peer = reaction.peer_id
                    user_id = getattr(peer, "user_id", None)
                    if user_id is None:
                        continue
                    # Resolve username via the users list returned alongside
                    username = None
                    for u in result.users:
                        if u.id == user_id:
                            username = u.username
                            break
                    out[key].append({
                        "user_id": user_id,
                        "username": (username.lower() if username else None),
                    })
                    collected += 1
                if not result.next_offset:
                    break
                offset = result.next_offset
                if collected > 1000:  # safety bound
                    break
        except terrors.MessageIdInvalidError:
            log_error(f"Launch announcement message_id={msg_id} not found in chat {chat_id}")
            continue
        except terrors.ReactionInvalidError:
            # The emoji isn't allowed in this chat - treat as zero reactions
            continue
        except Exception as e:
            log_error(f"Failed to read {emoji} reactions on msg {msg_id}", e)
            continue

    return out


async def _bootstrap_async() -> dict:
    """Async heart of the bootstrap subcommand. Returns a dict of results."""
    from telethon import TelegramClient

    api_id, api_hash, phone = _telethon_credentials()
    session_path = _telethon_session_path()

    chat_id_str = os.environ.get("FIRESIDE_TRIBE_CHAT_ID")
    if not chat_id_str:
        raise RuntimeError("FIRESIDE_TRIBE_CHAT_ID not set in .env")
    chat_id = int(chat_id_str)

    msg_id_str = os.environ.get("FIRESIDE_LAUNCH_ANNOUNCEMENT_MSG_ID")
    if not msg_id_str:
        raise RuntimeError("FIRESIDE_LAUNCH_ANNOUNCEMENT_MSG_ID not set in .env")
    msg_id = int(msg_id_str)

    client = TelegramClient(str(session_path), api_id, api_hash)
    await client.connect()
    if not await client.is_user_authorized():
        await client.disconnect()
        raise RuntimeError(
            "Telethon session not authorised. "
            "Run /telegram skill setup first to create the session."
        )

    try:
        tg_members = await _enumerate_tribe_members(client, chat_id)
        reactions = await _read_launch_reactions(client, chat_id, msg_id)
    finally:
        await client.disconnect()

    return {
        "telegram_members": tg_members,
        "reactions": reactions,
    }


def cross_reference(xlsx_roster: dict, telegram_members: dict) -> tuple[dict, dict]:
    """Build the operational tribe-roster.json and a discrepancy report.

    Args:
        xlsx_roster: dict[telegram_username -> metadata] from load_tribe_metadata()
        telegram_members: dict[telegram_username -> telegram record] from Telethon

    Returns:
        (roster, discrepancy):
            roster: dict[telegram_username -> merged record with telegram_user_id populated]
            discrepancy: dict with keys: in_tg_not_xlsx, in_xlsx_not_tg, no_username_in_tg
    """
    # Lowercase compare
    xlsx_lower = {u.lower(): (u, data) for u, data in xlsx_roster.items()}
    tg_lower = {u: m for u, m in telegram_members.items() if u != "_no_username"}

    # Load exclusions list (CEO-managed; persists across roster rebuilds)
    exclusions = load_state(EXCLUSIONS) or {}
    excluded_users = {k.lower(): v for k, v in exclusions.get("excluded", {}).items()}

    roster = {}
    in_tg_not_xlsx = []
    in_xlsx_not_tg = []

    for u_lower, tg_record in tg_lower.items():
        if tg_record.get("is_bot"):
            continue  # skip the bot itself and any other bots in the group
        if u_lower in xlsx_lower:
            original_username, xlsx_data = xlsx_lower[u_lower]
            merged = dict(xlsx_data)  # copy metadata
            merged["telegram_user_id"] = tg_record["user_id"]
            merged["telegram_full_name"] = tg_record["full_name"]
            # Apply CEO-managed exclusions: stays in Tribe (xlsx + Telegram group)
            # but bot excludes from speaker/helmsman/wildcard rotations.
            if u_lower in excluded_users:
                merged["active"] = False
                merged["excluded_from_fireside"] = True
                merged["exclusion_reason"] = excluded_users[u_lower].get("reason", "")
                merged["excluded_at"] = excluded_users[u_lower].get("excluded_at", "")
            roster[original_username] = merged
        else:
            in_tg_not_xlsx.append({
                "username": u_lower,
                "user_id": tg_record["user_id"],
                "full_name": tg_record["full_name"],
            })

    for u_lower, (original_username, xlsx_data) in xlsx_lower.items():
        if u_lower not in tg_lower:
            in_xlsx_not_tg.append({
                "username": original_username,
                "name": xlsx_data["name"],
            })

    discrepancy = {
        "in_telegram_not_in_xlsx": in_tg_not_xlsx,
        "in_xlsx_not_in_telegram": in_xlsx_not_tg,
        "no_username_in_telegram": telegram_members.get("_no_username", []),
    }
    return roster, discrepancy


def build_roster_by_name(roster: dict) -> dict:
    """Build a name -> roster_entry index for schedule lookups.

    An AMBIGUOUS name maps to nothing. Schedule rows join on display name, and
    two active members called "Alex Kim" used to collapse into whichever entry
    the dict iteration reached last -- so every "Alex Kim" session silently
    bound to one of them and DM'd the wrong person. Refusing to guess leaves
    the slot unbound, which the discrepancy report already surfaces; picking one
    is a wrong answer that looks like a right one.

    An EXCLUDED or DEPARTED member maps to nothing either, for the same reason
    and by the same convention `speaker_gaps` uses (`active` defaults True when
    the key is absent). The index was built over the whole roster, while
    `cross_reference` deliberately KEEPS excluded members in it with
    `active=False, excluded_from_fireside=True` and `_handle_chat_member` marks
    leavers `active=False`. Both `bootstrap` and `cycle-rollover` feed that
    roster to `build_schedule`, so a config naming an excluded or departed
    person bound their slot: they were DM'd the 2wk, 3day and day-of reminders
    (no DM loop checks `active`), and `speaker_gaps` stayed silent because it
    skips them as gap candidates AND counts their handle as scheduled. Nothing
    anywhere said the cycle config names someone who has gone.

    Leaving the slot unbound puts the name into `build_schedule`'s
    `missing_speakers`, which the discrepancy report and the rollover heads-up
    already print, and the reason is logged here so it is not read as a typo.
    """
    by_name: dict = {}
    ambiguous: dict[str, list[str]] = {}
    for username, data in roster.items():
        full_name = data.get("name", "").strip()
        if not full_name:
            continue
        if not data.get("active", True) or data.get("excluded_from_fireside"):
            reason = ("excluded from fireside" if data.get("excluded_from_fireside")
                      else "no longer active in the group")
            log_error(
                f"roster name {full_name!r} (@{username}) is {reason}; leaving any "
                f"schedule row naming them UNBOUND rather than scheduling and "
                f"DM-ing someone who has gone"
            )
            continue
        if full_name in by_name:
            ambiguous.setdefault(full_name, [by_name[full_name]["telegram_username"]])
            ambiguous[full_name].append(username)
            continue
        entry = dict(data)
        entry["telegram_username"] = username
        by_name[full_name] = entry
    for full_name, usernames in ambiguous.items():
        by_name.pop(full_name, None)
        log_error(
            f"roster name {full_name!r} maps to {len(usernames)} usernames "
            f"({', '.join('@' + u for u in usernames)}); leaving it UNBOUND "
            f"rather than guessing which member a schedule row means"
        )
    return by_name


def print_discrepancy_report(discrepancy: dict, missing_in_schedule: list) -> None:
    """Print a human-readable discrepancy summary."""
    in_tg = discrepancy["in_telegram_not_in_xlsx"]
    in_xlsx = discrepancy["in_xlsx_not_in_telegram"]
    no_un = discrepancy["no_username_in_telegram"]

    print()
    print(f"{BOLD}=== Discrepancy report ==={RESET}")

    if not in_tg and not in_xlsx and not no_un and not missing_in_schedule:
        print(f"{GREEN}OK{RESET}  No discrepancies. Telegram group + xlsx fully aligned.")
        return

    if in_tg:
        print(f"{YELLOW}In Telegram group but missing from xlsx ({len(in_tg)}):{RESET}")
        for r in in_tg:
            print(f"  @{r['username']} (id={r['user_id']}) - {r['full_name']}")
    if in_xlsx:
        print(f"{YELLOW}In xlsx but missing from Telegram group ({len(in_xlsx)}):{RESET}")
        for r in in_xlsx:
            print(f"  @{r['username']} - {r['name']}")
    if no_un:
        print(f"{YELLOW}In Telegram group but no username set ({len(no_un)}):{RESET}")
        print(f"      (these users cannot be matched via Telegram Username column)")
        for r in no_un:
            print(f"  user_id={r['user_id']} - {r['full_name']}")
    if missing_in_schedule:
        print(f"{RED}Names in 9-week schedule with no roster match ({len(missing_in_schedule)}):{RESET}")
        for n in sorted(set(missing_in_schedule)):
            print(f"  {n}")
        print(f"      (Verify name spelling in xlsx 'Name' column matches the schedule)")


def seed_opt_ins(reactions: dict, roster: dict) -> tuple[dict, list[str]]:
    """Turn launch-announcement reactions into opt-ins.json, through the roster.

    The SAME guard the live handler uses. `_handle_message_reaction` resolves a
    reactor's user_id to a canonical roster key and drops the opt-in when it
    cannot; bootstrap copied `r["username"]` straight off the reaction. Two ways
    that went wrong, both of which the roster already names:

      - A group member who is not in the xlsx - the `in_tg_not_xlsx` discrepancy
        class exists precisely because this happens - was seeded, and then
        appeared on the wildcard roster DM'd to every Helmsman.
      - A reactor with no Telegram handle was seeded as `username: None` and
        rendered by `cmd_helmsman_brief` as the literal `  - @None`.

    Resolution is by user_id against the roster passed in, not by re-reading
    state per reactor, and not by handle: a reclaimed handle would otherwise
    resolve to the previous owner's rows.

    Returns (opt_ins, dropped) where `dropped` names every reaction refused, so
    the refusal is reported rather than silent.
    """
    by_user_id = {v.get("telegram_user_id"): k for k, v in roster.items()
                  if v.get("telegram_user_id")}
    dropped: list[str] = []

    def _seed(bucket: str) -> list:
        out = []
        for r in reactions.get(bucket, []):
            canonical = by_user_id.get(r.get("user_id"))
            if canonical is None:
                dropped.append(f"{bucket}:@{r.get('username') or '(no handle)'} "
                               f"(id={r.get('user_id')})")
                continue
            out.append({"username": canonical, "user_id": r["user_id"]})
        return out

    return {"helmsman": _seed("helmsman"), "wildcard": _seed("wildcard")}, dropped


def cmd_bootstrap(args) -> None:
    """Bootstrap the bot: enumerate Telegram, cross-reference with xlsx, generate schedule + opt-ins."""
    import asyncio

    print(f"{BOLD}=== Phase 2 Bootstrap ==={RESET}")
    ensure_state_dir()

    print(f"{CYAN}1. Loading xlsx roster...{RESET}")
    try:
        xlsx_roster = load_tribe_metadata()
    except (FileNotFoundError, *_UNREADABLE_SHEET) as e:
        print(f"{RED}xlsx load failed:{RESET} {e}", file=sys.stderr)
        print(f"{YELLOW}Hint: complete Phase 0 task 0.5 (add Telegram Username column "
              f"to xlsx and populate for all 54 Tribe members).{RESET}", file=sys.stderr)
        sys.exit(1)
    print(f"     {len(xlsx_roster)} entries with telegram_username")

    print(f"{CYAN}2. Connecting to Telegram via Telethon (as Misha)...{RESET}")
    try:
        bootstrap_result = asyncio.run(_bootstrap_async())
    except Exception as e:
        print(f"{RED}Telethon bootstrap failed:{RESET} {e}", file=sys.stderr)
        sys.exit(1)

    tg_members = bootstrap_result["telegram_members"]
    reactions = bootstrap_result["reactions"]
    n_tg = len(tg_members) - 1  # exclude '_no_username' bucket
    n_no_un = len(tg_members.get("_no_username", []))
    print(f"     {n_tg} members with username + {n_no_un} without username")
    print(f"     🧭 helmsman reactors: {len(reactions['helmsman'])}")
    print(f"     🌟 wildcard reactors: {len(reactions['wildcard'])}")

    print(f"{CYAN}3. Cross-referencing xlsx + Telegram members...{RESET}")
    roster, discrepancy = cross_reference(xlsx_roster, tg_members)
    print(f"     {len(roster)} matched entries written to tribe-roster.json")
    save_state(TRIBE_ROSTER, roster)

    print(f"{CYAN}4. Building schedule.json from Week 1-9 calendar...{RESET}")
    roster_by_name = build_roster_by_name(roster)
    existing_schedule = load_state(SCHEDULE)
    if existing_schedule:
        # Don't clobber the live schedule on re-bootstrap. Manual swaps and
        # exclusions are recorded directly in schedule.json (e.g. a mid-cycle
        # speaker swap and a member exclusion); rebuilding from the
        # week_1_to_9_schedule() calendar would wipe both. Compute missing-
        # speaker stats for the discrepancy report but do not save.
        _, missing_in_schedule = build_schedule(roster_by_name)
        print(f"     {len(existing_schedule)} schedule entries already populated; not overwriting "
              f"(delete schedule.json manually for a clean rebuild)")
    else:
        schedule_entries, missing_in_schedule = build_schedule(roster_by_name)
        save_state(SCHEDULE, schedule_entries)
        print(f"     {len(schedule_entries)} schedule entries written ({len(missing_in_schedule)} unresolved)")

    print(f"{CYAN}5. Seeding helmsmen.json (empty - Misha selects Week 1 from reactors)...{RESET}")
    if load_state(HELMSMEN) is None or load_state(HELMSMEN) == {}:
        save_state(HELMSMEN, {})
        print(f"     helmsmen.json initialised empty")
    else:
        print(f"     helmsmen.json already populated; not overwriting")

    print(f"{CYAN}6. Writing opt-ins.json from launch-announcement reactions...{RESET}")
    opt_ins, dropped = seed_opt_ins(reactions, roster)
    save_state(OPT_INS, opt_ins)
    print(f"     {len(opt_ins['helmsman'])} helmsman opt-ins, {len(opt_ins['wildcard'])} wildcard opt-ins")
    if dropped:
        print(f"     {YELLOW}{len(dropped)} reaction(s) NOT seeded - reactor is not "
              f"a bound roster member:{RESET}")
        for d in sorted(dropped):
            print(f"       - {d}")

    print_discrepancy_report(discrepancy, missing_in_schedule)

    print()
    print(f"{GREEN}OK{RESET}  Bootstrap complete. State files written to {state_dir()}")


# ============================================================
# Subcommand: test-telegram (Phase 1 DoD)
# ============================================================

def cmd_test_telegram(args) -> None:
    """Send a smoke-test DM to Misha. Confirms bot can authenticate and DM.

    This is the Phase 1 Definition-of-Done check.
    """
    load_env()
    bot = get_bot()

    misha_id_str = os.environ.get("MISHA_TELEGRAM_USER_ID")
    if not misha_id_str:
        print(f"{RED}ERROR: MISHA_TELEGRAM_USER_ID not set in .env{RESET}", file=sys.stderr)
        sys.exit(1)
    try:
        misha_id = int(misha_id_str)
    except ValueError:
        print(f"{RED}ERROR: MISHA_TELEGRAM_USER_ID is not an integer: {misha_id_str!r}{RESET}",
              file=sys.stderr)
        sys.exit(1)

    # Confirm bot identity first
    me = bot.get_me()
    bot_handle = me.get("username", "unknown")
    bot_id = me.get("id", "?")

    hostname = socket.gethostname()
    now_str = local_now().strftime("%Y-%m-%d %H:%M:%S %Z")
    text = (
        f"\U0001F916 *Bot online — first contact*\n\n"
        f"Host: `{hostname}`\n"
        f"Time: `{now_str}`\n"
        f"Bot: `@{bot_handle}` (id={bot_id})\n\n"
        f"This is the Phase 1 smoke test from `scripts/fireside-bot.py test-telegram`. "
        f"If you're reading this, the bot can DM you. Phase 1 DoD ✅"
    )
    try:
        result = bot.send_dm(misha_id, text)
    except TelegramAPIError as e:
        print(f"{RED}FAIL{RESET}  Could not DM Misha (user_id={misha_id})", file=sys.stderr)
        print(f"      {e}", file=sys.stderr)
        sys.exit(1)

    msg_id = result.get("message_id") if isinstance(result, dict) else "?"
    print(f"{GREEN}OK{RESET}  DM sent to Misha (user_id={misha_id})")
    print(f"     Telegram message_id = {msg_id}")
    print(f"     Bot: @{bot_handle} (id={bot_id})")
    print(f"     Time: {now_str}")


# ============================================================
# Subcommand: xlsx-check (Phase 1 DoD helper)
# ============================================================

def cmd_xlsx_check(args) -> None:
    """Print xlsx loader output summary. Verifies load_tribe_metadata() works."""
    try:
        roster = load_tribe_metadata()
    except (FileNotFoundError, *_UNREADABLE_SHEET) as e:
        print(f"{RED}xlsx load failed:{RESET} {e}", file=sys.stderr)
        sys.exit(1)

    n = len(roster)
    print(f"{CYAN}xlsx loaded: {n} entries with telegram_username{RESET}")
    if n == 0:
        print(f"{YELLOW}Note: 0 entries means the Telegram Username column exists but is empty.{RESET}")
        print(f"{YELLOW}      Populate it for all 54 Tribe members (Phase 0 task 0.5).{RESET}")
        return

    sample = list(roster.items())[:3]
    print(f"{GRAY}Sample (first {len(sample)}):{RESET}")
    for username, data in sample:
        vp_marker = " [VP]" if data["is_vp"] else ""
        print(f"  @{username}: {data['name']}{vp_marker}")
        print(f"      function={data['function']!r}, title={data['title']!r}")
        print(f"      email={data['email']}, user_id={data['telegram_user_id']}")

    vps = sum(1 for d in roster.values() if d["is_vp"])
    no_user_id = sum(1 for d in roster.values() if d["telegram_user_id"] is None)
    print(f"{GRAY}Totals: {n} entries, {vps} VPs, {no_user_id} awaiting telegram_user_id{RESET}")


# ============================================================
# Subcommand: init-state (Phase 1 DoD helper)
# ============================================================

def cmd_init_state(args) -> None:
    """Initialise state directory + empty state files. Idempotent."""
    ensure_state_dir()
    print(f"{GREEN}OK{RESET}  State directory ready: {state_dir()}")
    files = [TRIBE_ROSTER, SCHEDULE, HELMSMEN, OPT_INS,
             DM_LOG, SESSIONS_LOG, LAST_UPDATE_ID, LAST_PINNED, ERRORS_LOG]
    for name in files:
        path = state_path(name)
        if path.exists():
            size = path.stat().st_size
            print(f"     {path.name}: {size} bytes")
        else:
            print(f"     {RED}{path.name}: MISSING{RESET}")


# ============================================================
# Templates (Phase 3) - mirror outputs/operations/tribe-fireside/*-template.md
# ============================================================

# Note: keep these in sync with the .md template files. The .md is the human
# reference; this is what the bot actually renders.

SPEAKER_DM_2WK = """Hi {name},

You're on the speaker list for the Tribe fireside on {session_date} ({session_day}). The theme that week is **{theme}**.

You'll have 5 minutes to share + a couple of minutes for questions from the Tribe.

Some category ideas if helpful: a book, a place you've lived or want to live, a kid story, something you're proud of (outside work), the most interesting thing you learned this month, a question you want to ask the Tribe, your hometown, a skill you're learning, your last weekend - or something completely else if the theme sparks something.

Two weeks gives you time to think about what you want to share. No pressure to be polished.

If the date doesn't work, send /swap to the bot - it'll show you open dates you can move to (or arrange a swap with another speaker) on the spot.

— 31C Fireside Bot"""

SPEAKER_DM_3DAY = """Hi {name},

Your Tribe fireside slot is in 3 days — {session_day} {session_date}. Theme: **{theme}**.

5 minutes to share, then Q&A. Format is intentionally informal — no slides, just talk.

Drop a 1-paragraph preview into the 31C Tribe group on Sunday evening if you'd like — it warms the room. Optional.

— 31C Fireside Bot"""

SPEAKER_DM_DAYOF = """Hi {name},

Your Tribe fireside is at 18:30 local time today.

Zoom: {zoom_link}

5 min share + Q&A. Helmsman this week is {helmsman_name}.

— 31C Fireside Bot"""

HELMSMAN_BRIEF = """Hi {name},

You're the Helmsman for the week starting {week_starting}. Your job: open and close two firesides — Mon {monday_date} and Wed {wednesday_date} at 18:30 local time — and hold the line on the format.

**This week's speakers:**
- Mon: {monday_speakers}
- Wed: {wednesday_speakers}
- Theme: **{theme}**

**Wildcard roster** (in case of no-show, in priority order):
{wildcard_list}

**Pop-up rules to read at session open:**
"Welcome. For the next thirty minutes — no work talk except things you're proud of. No laptops. No Slack. We're here to actually meet each other. Theme this week: {theme}. Speakers today are [names]. Five minutes each plus a couple of minutes for questions. Let's start."

**Closing go-around (last 5 min):** pick 4 random Tribe members from the audience. "One thing you're taking from this session — 30 seconds each."

**If a speaker no-shows:** DM the wildcard roster in order. First to respond within 90 sec takes the slot. If no response, run a group prompt: "Quick — [theme-relevant question]. Anyone share for two minutes."

**If somebody slips into work talk:** gentle redirect. "Let's hold that for the standup, this is the fireside."

**Time discipline:** 30 min default. Flex to 35 max if room is in flow. Beyond 35 — close cleanly.

You've got this. The role rotates - somebody else next week. DM Misha if anything goes sideways.

— 31C Fireside Bot"""

EMAIL_BACKUP_SUBJECT = "Your Tribe fireside slot — {session_date}"
EMAIL_BACKUP_BODY = """Hi {name},

Quick note from the Tribe fireside system — you're on the speaker list for {session_date} ({session_day}). I've sent you a few Telegram DMs about it but haven't seen a response, so wanted to make sure you saw this here too.

Theme that week: {theme}

5 minutes to share, plus a couple of minutes for Q&A. Format is informal.

If the date doesn't work, reply to this email and we'll find a swap.

— 31C Fireside Bot{human_contact}"""

SUNDAY_PREVIEW = """🔥 **Tribe fireside this week**

**Theme:** {theme}

**Monday {monday_date}:** {monday_speakers}
**Wednesday {wednesday_date}:** {wednesday_speakers}

Speakers — drop a 1-paragraph preview + a photo as a reply to this message if you'd like. Warms the room.
Tribe — react with 🔥 📚 🌍 ❤️ to whatever resonates.

**Helmsman this week:** {helmsman_name}

Zoom: {zoom_link} · Same recurring link every week.

See you Monday at 18:30 local time."""


UNAUTHORIZED_REPLY = (
    "This bot is private to the 31C Tribe Fireside. "
    "If you think you should have access, message Misha."
)

OUTSIDER_FORWARD_COOLDOWN_S = 3600  # 1 forward to Misha per outsider per hour


WELCOME_DM = """Welcome to the Tribe Fireside Bot.

You're subscribed to automatic reminders:
  - 2 weeks before you speak
  - 3 days before
  - Day-of, with the Zoom link

Format: Tribe Fireside runs every Mon and Wed at 18:30 local time. Three speakers per session (~7 min each). Theme rotates weekly.

Commands you can use anytime:
  /me        - when am I scheduled to speak?
  /next      - who's at the next session?
  /who       - speakers and theme for this week
  /theme     - this week's theme only
  /schedule  - full 9-week schedule
  /zoom      - Zoom link
  /swap      - show open dates or counterparty swaps you can move to
  /idea      - propose a topic for a future fireside
  /help      - this menu

Questions or feedback? Just reply here. Misha reads DMs."""


HELP_DM = """Tribe Fireside Bot - commands:

  /me        - when am I scheduled to speak?
  /next      - who's at the next session?
  /who       - speakers and theme for this week
  /theme     - this week's theme only
  /schedule  - full 9-week schedule
  /zoom      - Zoom link
  /swap      - show open dates or counterparty swaps you can move to
  /idea      - propose a topic for a future fireside
  /help      - this menu

Sessions run Mon and Wed at 18:30 local time. Reply here with any question."""


# ============================================================
# Phase 3 helpers
# ============================================================

def _zoom_url() -> str:
    """Return the recurring Zoom URL from env, or a placeholder warning string."""
    load_env()
    url = os.environ.get("FIRESIDE_ZOOM_URL", "").strip()
    if not url:
        return "[FIRESIDE_ZOOM_URL not set in .env - Misha needs to add it]"
    return url


def _week_speakers(schedule: list, week: int, day: str) -> list[dict]:
    """Return speaker entries for a given week + day (Mon/Wed) sorted by slot."""
    return sorted(
        [s for s in schedule if s["week"] == week and s["day"] == day],
        key=lambda s: s["slot"],
    )


def _today_local_date():
    """Return today's date in local timezone as a date object."""
    return local_now().date()


def _current_or_upcoming_week(schedule: list, today=None) -> Optional[int]:
    """Return the week number of the current or next upcoming session.

    If today's date matches a session date, return that week.
    Otherwise return the week of the next future session.
    Returns None if no future sessions remain.
    """
    from datetime import date as _date
    if today is None:
        today = _today_local_date()
    upcoming = [s for s in schedule if _date.fromisoformat(s["session_date"]) >= today]
    if not upcoming:
        return None
    upcoming.sort(key=lambda s: s["session_date"])
    return upcoming[0]["week"]


def _first_name(raw, fallback: str) -> str:
    """The greeting name, or `fallback` when there is nothing to greet with.

    One implementation for the four send loops that need it. Three of them
    carried their own copy of `raw.split()[0] if raw else <fallback>`, each
    above a comment recording that a blank name had already crashed a live send
    loop mid-flight. The fourth, `cmd_helmsman_brief`, was written without the
    guard and read `entry.get("name", "Helmsman").split()[0]`, where the `.get`
    default never applies because the key is PRESENT and empty: `""` reaches
    `[0]` and raises IndexError, `None` raises AttributeError. That one runs on
    a daily cron and raises before the entry is stamped `briefed`, so the job
    picks the same candidate again the next day, and the next, and no Helmsman
    is ever briefed. Its healthcheck goes red by silence at the same time.
    """
    text = str(raw or "").strip()
    return text.split()[0] if text else fallback


def _roster_entry(roster: dict, username: Optional[str]) -> Optional[dict]:
    """The roster record for `username`, matched exactly then case-insensitively.

    Roster keys keep whatever case the xlsx column held; a schedule row carries
    whatever case the cycle config or the swap machinery produced, and nothing
    normalises either side at write time. Two of the three call sites already
    fell back to a lowercase comparison and the third, `cmd_email_backup`, did a
    bare `roster.get(username)`, so a case mismatch there classified the member
    as `not-in-roster` and the one person that command exists to reach, an
    unresponsive speaker, was named in a summary line instead of emailed.
    """
    if not username:
        return None
    entry = roster.get(username)
    if entry is not None:
        return entry
    lowered = str(username).lower()
    for key, data in roster.items():
        if str(key).lower() == lowered:
            return data
    return None


def _resolve_speaker_user_id(roster: dict, speaker_username: Optional[str]) -> Optional[int]:
    """Look up a speaker's telegram_user_id from the roster by username."""
    entry = _roster_entry(roster, speaker_username)
    return entry.get("telegram_user_id") if entry else None


def _dm_already_sent(dm_log_path: Path, speaker_username: str, dm_type: str,
                     session_date: str, on_date: Optional[str] = None) -> bool:
    """Check dm-log.jsonl to see if a specific DM type has already been sent
    for a specific speaker + session.

    `on_date` (an ISO `YYYY-MM-DD`) narrows the question to "already sent TODAY"
    rather than "already sent at all". `email-backup` needs the narrow form: it
    mails on BOTH Sundays inside a session's 1..14 day window on purpose,
    mirroring the two DM nudges, so the whole-window form would silence the
    second Sunday. Without either form, a rerun or a cron double-fire on one
    Sunday sent every addressee a second identical email, and the
    `_log_dm("email-backup", ...)` rows that exist to prevent it were written
    and never read.
    """
    if not dm_log_path.exists():
        return False
    # Through `jsonl_lines`, so a torn append costs that line rather than the
    # run. The decode used to happen inside `for line in f`, outside the handler
    # below, and `UnicodeDecodeError` is a `ValueError` that
    # `except json.JSONDecodeError` cannot see. A raise here takes the whole
    # email-backup job down; the alternative failure, a silent False, mails a
    # second identical letter to a real person. Skipping the one line is neither.
    for entry in _read_jsonl_rows(dm_log_path):
        if (entry.get("dm_type") == dm_type
                and entry.get("speaker_username") == speaker_username
                and entry.get("session_date") == session_date
                and entry.get("delivered")):
            if on_date is not None and str(entry.get("ts", ""))[:10] != on_date:
                continue
            return True
    return False


def _log_dm(dm_type: str, speaker_username: str, session_date: str,
            user_id: Optional[int], delivered: bool, error: Optional[str] = None) -> None:
    """Append a DM event to dm-log.jsonl."""
    append_jsonl(DM_LOG, {
        "ts": local_now().isoformat(),
        "dm_type": dm_type,
        "speaker_username": speaker_username,
        "session_date": session_date,
        "user_id": user_id,
        "delivered": delivered,
        "error": error,
    })


def _log_event(event_type: str, **fields) -> None:
    """Append a generic event to sessions.jsonl."""
    payload = {"ts": local_now().isoformat(), "event_type": event_type}
    payload.update(fields)
    append_jsonl(SESSIONS_LOG, payload)


# ============================================================
# Subcommand: poll (Phase 3 task 3.1)
# ============================================================

def cmd_poll(args) -> None:
    """Process Telegram updates: /start, /swap, message_reaction, chat_member events.

    Cron: every 5 min.
    Drains queue if 100 updates returned (cap-hit) to prevent 24h retention loss.
    """
    # Guarded on THIS subcommand rather than in main(), because only this one is
    # a daemon entry point: it sits in `while True: bot.get_updates(timeout=25)`
    # and is a long-lived poller whether the fireside daemon drives it or a
    # human runs it by hand. The other subcommands are ordinary CLI work and
    # stay open from a worktree. Same shape as the `memory-index build` guard,
    # and the same reason: a whole-script guard here would refuse reads that
    # nothing objects to.
    require_main_clone(__file__)
    bot = get_bot()
    last = load_state(LAST_UPDATE_ID) or {"offset": 0}
    offset = int(last.get("offset", 0))
    total_processed = 0

    # The liveness marker is written AFTER the first getUpdates succeeds, not
    # before it. Written first, a host with a revoked token or no route to
    # Telegram still stamped a fresh tick every five minutes, and health-check
    # read "alive" for the whole outage. The marker now means "this process
    # reached Telegram", which is the thing health-check is being asked.
    ticked = False

    while True:
        try:
            updates = bot.get_updates(offset=offset, timeout=25, limit=100)
        except TelegramAPIError as e:
            print(f"{RED}poll: getUpdates failed: {e}{RESET}", file=sys.stderr)
            return
        if not ticked:
            append_jsonl(DM_LOG, {
                "ts": local_now().isoformat(),
                "dm_type": "poll-tick",
            })
            ticked = True

        if not updates:
            break

        for update in updates:
            try:
                _handle_update(bot, update)
            except Exception as e:
                log_error(f"poll: failed to handle update {update.get('update_id')}", e)
            new_offset = update.get("update_id", 0) + 1
            if new_offset > offset:
                offset = new_offset
            total_processed += 1

        save_state(LAST_UPDATE_ID, {"offset": offset})

        if len(updates) < 100:
            break  # not at cap, queue is drained

    if total_processed:
        print(f"{GRAY}poll: processed {total_processed} update(s){RESET}")
    hc_ping("FIRESIDE_HC_POLL")


def _handle_update(bot: TelegramBot, update: dict) -> None:
    """Route a single Telegram update by type."""
    # Lazy expiration: walk swap-requests.jsonl, expire any past-deadline `proposed_to_b`.
    # Cheap while the log stays small; revisit if it grows past ~10k events.
    try:
        _sweep_expired_swap_requests(bot)
    except Exception as e:
        log_error("swap-expiration sweep failed", e)

    if "message" in update:
        _handle_message(bot, update["message"])
    elif "callback_query" in update:
        _handle_callback_query(bot, update["callback_query"])
    elif "message_reaction" in update:
        _handle_message_reaction(update["message_reaction"])
    elif "chat_member" in update:
        _handle_chat_member(update["chat_member"])
    elif "my_chat_member" in update:
        # The bot's own membership changed. A routine event, recorded and not
        # acted on -- so it belongs in sessions.jsonl, not errors.log. It used to
        # be appended there under an "ERROR:" prefix, which is the file an
        # operator opens to find what broke; a Telegram event that broke nothing
        # diluted it.
        _log_event("my_chat_member",
                   detail=json.dumps(update["my_chat_member"])[:200])


def _resolve_my_username(user_id: int) -> Optional[str]:
    """Return the canonical roster-key @username bound to this Telegram user_id.

    Roster keys are the @usernames used as join keys with schedule.json /
    helmsmen.json. Resolution is by telegram_user_id ONLY -- the immutable,
    Telegram-assigned id. We deliberately do NOT fall back to a self-reported
    @username: that would let someone who has claimed a former member's dropped
    handle resolve to that member's schedule/helmsman rows (handle takeover).
    telegram_user_id is bound only by the trusted bootstrap, so any authorized
    caller already has a user_id match here.
    """
    roster = load_state(TRIBE_ROSTER) or {}
    for k, v in roster.items():
        if v.get("telegram_user_id") == user_id:
            return k
    return None


def _cmd_me_text(user_id: int) -> str:
    """Personalised schedule view for the calling user (identity by user_id)."""
    from datetime import date as _date
    schedule = load_state(SCHEDULE) or []
    helmsmen = load_state(HELMSMEN) or {}
    my_username = _resolve_my_username(user_id)
    if not my_username:
        return ("Couldn't find you in the Tribe roster. "
                "If you think this is wrong, reply here and Misha will check.")
    today = _today_local_date()
    upcoming, past = [], []
    for e in schedule:
        if (e.get("speaker_username") or "").lower() != my_username.lower():
            continue
        d = _date.fromisoformat(e["session_date"])
        line = (f"  - {e['session_date']} ({e['day']}) - "
                f"Week {e['week']}, slot {e['slot']}, theme: {e['theme']}")
        (upcoming if d >= today else past).append(line)
    lines = []
    if upcoming:
        lines.append("Upcoming speaker slots:")
        lines.extend(upcoming)
    else:
        lines.append("No upcoming speaker slots in this cycle.")
    if past:
        lines.append("")
        lines.append(f"Completed: {len(past)} session(s).")
    upcoming_helmsman = []
    for week_start, entry in helmsmen.items():
        if (entry.get("username") or "").lower() != my_username.lower():
            continue
        try:
            d = _date.fromisoformat(week_start)
        except ValueError:
            continue
        if d >= today:
            tag = " (already briefed)" if entry.get("briefed") else ""
            upcoming_helmsman.append((d, week_start, tag))
    if upcoming_helmsman:
        upcoming_helmsman.sort()
        lines.append("")
        lines.append("Helmsman weeks (you run the sessions):")
        for d, ws, tag in upcoming_helmsman:
            lines.append(f"  - Week starting {ws}{tag}")
    return "\n".join(lines)


def _cmd_next_text() -> str:
    """The very next future session: date, speakers, helmsman.

    Treats today as "past" after 19:30 local (sessions are 18:30-19:00 + buffer).
    """
    from datetime import date as _date, timedelta
    schedule = load_state(SCHEDULE) or []
    helmsmen = load_state(HELMSMEN) or {}
    now = local_now()
    today = now.date()
    cutoff = today
    if now.hour > 19 or (now.hour == 19 and now.minute >= 30):
        cutoff = today + timedelta(days=1)
    upcoming = [e for e in schedule if _date.fromisoformat(e["session_date"]) >= cutoff]
    if not upcoming:
        return "No upcoming sessions in the current cycle."
    upcoming.sort(key=lambda e: (e["session_date"], e["slot"]))
    next_date = upcoming[0]["session_date"]
    next_speakers = sorted([e for e in upcoming if e["session_date"] == next_date],
                           key=lambda e: e["slot"])
    d = _date.fromisoformat(next_date)
    mon_of_week = d - timedelta(days=d.weekday())
    h_entry = helmsmen.get(mon_of_week.isoformat()) or {}
    lines = [
        f"Next session: {d.strftime('%A')} {next_date} at 18:30 local time",
        f"Week {next_speakers[0]['week']}, theme: {next_speakers[0]['theme']}",
        "",
        "Speakers:",
    ]
    for e in next_speakers:
        lines.append(f"  {e['slot']}. {e['speaker_name']} (@{e['speaker_username']})")
    lines.append("")
    lines.append(f"Helmsman: {h_entry.get('name', 'TBD')}")
    return "\n".join(lines)


def _cmd_who_text() -> str:
    """This week's both sessions (Mon + Wed)."""
    schedule = load_state(SCHEDULE) or []
    helmsmen = load_state(HELMSMEN) or {}
    week_num = _current_or_upcoming_week(schedule)
    if week_num is None:
        return "No active week in the schedule."
    mon = _week_speakers(schedule, week_num, "Mon")
    wed = _week_speakers(schedule, week_num, "Wed")
    if not mon and not wed:
        return f"Week {week_num} has no scheduled speakers."
    theme = (mon[0] if mon else wed[0])["theme"]
    lines = [f"Week {week_num} - theme: {theme}", ""]
    if mon:
        lines.append(f"Mon {mon[0]['session_date']}:")
        for e in mon:
            lines.append(f"  {e['slot']}. {e['speaker_name']} (@{e['speaker_username']})")
    if wed:
        if mon:
            lines.append("")
        lines.append(f"Wed {wed[0]['session_date']}:")
        for e in wed:
            lines.append(f"  {e['slot']}. {e['speaker_name']} (@{e['speaker_username']})")
    week_start = mon[0]["session_date"] if mon else None
    if week_start:
        h_entry = helmsmen.get(week_start) or {}
        if h_entry.get("name"):
            lines.append("")
            lines.append(f"Helmsman: {h_entry['name']}")
    return "\n".join(lines)


def _cmd_theme_text() -> str:
    schedule = load_state(SCHEDULE) or []
    week_num = _current_or_upcoming_week(schedule)
    if week_num is None:
        return "No active week in the schedule."
    entries = [e for e in schedule if e["week"] == week_num]
    if not entries:
        return f"Week {week_num} has no entries."
    return f"Week {week_num} theme: {entries[0]['theme']}"


def _cmd_schedule_text() -> str:
    schedule = load_state(SCHEDULE) or []
    if not schedule:
        return "No schedule loaded."
    weeks = sorted(set(e["week"] for e in schedule))
    lines = ["Tribe Fireside - full 9-week schedule (Mon and Wed at 18:30 local time)"]
    for w in weeks:
        wk = [e for e in schedule if e["week"] == w]
        if not wk:
            continue
        lines.append("")
        lines.append(f"Week {w} - {wk[0]['theme']}")
        mon = sorted([e for e in wk if e["day"] == "Mon"], key=lambda e: e["slot"])
        wed = sorted([e for e in wk if e["day"] == "Wed"], key=lambda e: e["slot"])
        if mon:
            names = ", ".join(e["speaker_name"] for e in mon)
            lines.append(f"  Mon {mon[0]['session_date']}: {names}")
        if wed:
            names = ", ".join(e["speaker_name"] for e in wed)
            lines.append(f"  Wed {wed[0]['session_date']}: {names}")
    return "\n".join(lines)


def _cmd_zoom_text() -> str:
    return (f"Zoom: {_zoom_url()}\n"
            "Same recurring link every Mon and Wed at 18:30 local time.")


def _is_authorized_user(user_id: int, username: Optional[str] = None) -> bool:
    """True iff user_id maps to an active, non-excluded Tribe roster member.

    Authorization is by Telegram user_id ONLY. A user_id is immutable and
    assigned by Telegram; a @username is mutable and reclaimable, so a username
    match is NOT proof of identity. Binding a user_id to a roster entry happens
    exclusively through the trusted `bootstrap` subcommand, which enumerates the
    real Tribe group via Misha's authenticated Telethon session (see
    cross_reference()). DM handlers never persist telegram_user_id from a
    self-reported username -- doing so would let anyone who claims a former
    member's dropped @handle take over that member's authorization (handle
    takeover). After a state-loss that leaves entries with telegram_user_id=None,
    recovery is a `bootstrap` run, not a self-reported DM.

    `username` is accepted only so callers can log it; it is deliberately ignored
    for the authorization decision.
    """
    if not user_id:
        return False
    roster = load_state(TRIBE_ROSTER) or {}
    for v in roster.values():
        if v.get("telegram_user_id") == user_id:
            return bool(v.get("active")) and not v.get("excluded_from_fireside", False)
    return False


def _maybe_forward_outsider(bot: "TelegramBot", user_id: int,
                             username: str, text: str) -> None:
    """Forward an outsider DM to Misha at most once per hour per user_id.

    Rate-limited via OUTSIDER_RATE state file. Failures to forward are silent;
    callers still log to sessions.jsonl regardless.
    """
    # No try/except: `misha_user_id()` already returns 0 rather than raising.
    misha_id = misha_user_id()
    if not misha_id:
        return
    rate = load_state(OUTSIDER_RATE) or {}
    last_iso = rate.get(str(user_id))
    if last_iso:
        try:
            last_dt = datetime.fromisoformat(last_iso)
            if (local_now() - last_dt).total_seconds() < OUTSIDER_FORWARD_COOLDOWN_S:
                return
        except ValueError:
            pass  # corrupt timestamp - forward and rewrite
    preview = (text or "(no text)")[:300]
    if text and len(text) > 300:
        preview += "..."
    try:
        bot.send_message(
            misha_id,
            f"Outsider DM to Fireside bot from @{username or '(no username)'} "
            f"(id={user_id}):\n{preview}",
            parse_mode="",
        )
    except TelegramAPIError as e:
        # The cooldown was written whatever happened, so a forward that FAILED
        # still silenced that outsider for the next hour: the operator never saw
        # the message and never saw a second chance at it either. A cooldown
        # suppresses a REPEAT of something delivered; there was no delivery.
        log_error(f"outsider forward to Misha failed for user {user_id}", e)
        return
    rate[str(user_id)] = local_now().isoformat()
    save_state(OUTSIDER_RATE, rate)


# ============================================================
# /swap interactive flow (Phase 3.5: self-serve swap state machine)
# ============================================================
#
# Event log lives at SWAP_REQUESTS_LOG (append-only JSONL). Each event row
# carries `rid` (request id, 8 hex chars). Latest event per rid defines
# current status: initiated -> a_tapped_vacancy|a_tapped_counterparty|cancelled_by_a;
# a_tapped_counterparty -> b_accepted|b_declined|expired. Terminal events are
# *_completed / *_declined / cancelled_by_a / expired.
#
# Callback_data schemas (Telegram limit: 64 bytes):
#   sw:a:<rid>:<idx>     - A tapped candidate idx 0/1
#   sw:a:<rid>:x         - A tapped Cancel
#   sw:b:<rid>:y         - B accepted
#   sw:b:<rid>:n         - B declined

import secrets as _secrets  # noqa: E402  (module-level import already present elsewhere)
from scripts.utils.clone_guard import require_main_clone


def _tell(bot: TelegramBot, user_id, text: str) -> None:
    """Best-effort DM: no recipient, or a Telegram failure, is not a hard error.

    The `if user_id:` + `try/except TelegramAPIError: pass` pair this replaces
    is written out about twenty times in the swap handlers. The new notification
    paths use this instead of a twenty-first copy, and `contextlib.suppress`
    rather than a bare `pass` because ruff reads the latter as a smell and the
    lint ratchet counts it.
    """
    if not user_id:
        return
    with contextlib.suppress(TelegramAPIError):
        bot.send_message(int(user_id), text, parse_mode="")


def _new_request_id() -> str:
    """Return an 8-char hex request id, collision risk negligible per session."""
    return _secrets.token_hex(4)


def _format_dm_date(date_iso: str, day: str | None) -> str:
    """'2026-06-08' + 'Mon' -> 'Mon, 8 Jun'. A blank day is read off the date.

    `day` is `str | None` because `or` already accepted both and one caller can
    supply either. The annotation said `str` while the test beside it passed
    `day or ""`, so neither declared nor measured the case the code handles.

    Where a None comes from, corrected 2026-09-02. This paragraph first said
    `c["day"]` "comes straight out of `config/fireside-schedule.json`, where
    `"day": null` is legal", which is false in both halves: that file carries
    `week`/`theme`/`mon`/`wed` and no `day` key at all, and every `day` in a
    schedule entry is the literal "Mon" or "Wed" that `build_schedule` writes.
    The value reaches here through the PERSISTED schedule state, read back as
    `entries[0].get("day", "")` in `find_swap_candidates` and carried into the
    candidate dict the tap handlers store and re-read. A missing key yields ""
    from that default; a key present and null yields None, which is what a
    hand-edited or half-written state file produces. That is the case the
    annotation now declares, and it is the reason `or` is load-bearing rather
    than decorative.

    Three call sites pass no day at all. `_handle_a_tap` reads `ctx["a_day"]`,
    which `_swap_kickoff_for_a` has never written, and `_handle_b_tap` passes ""
    outright, twice. Formatting a blank straight into the template produced
    ", 8 Jun" -- a label opening on a comma with the weekday silently gone --
    inside the swap proposal B reads and both confirmations after B accepts.
    Those three are the highest-stakes DMs this bot sends: two real people are
    agreeing to trade slots on the strength of the date in them.
    """
    from datetime import date as _date
    d = _date.fromisoformat(date_iso)
    return f"{day or d.strftime('%a')}, {d.day} {d.strftime('%b')}"


def _user_current_slot(schedule: list, username: str, today) -> Optional[dict]:
    """Return A's nearest future schedule entry, or None.

    'Future' means session_date >= today. If A has multiple future slots,
    returns the closest one by date.
    """
    from datetime import date as _date
    candidates = []
    uname_lc = username.lower()
    for e in schedule:
        if (e.get("speaker_username") or "").lower() != uname_lc:
            continue
        try:
            d = _date.fromisoformat(e["session_date"])
        except (ValueError, KeyError):
            continue
        if d >= today:
            candidates.append((d, e))
    if not candidates:
        return None
    candidates.sort(key=lambda x: x[0])
    return candidates[0][1]


def find_swap_candidates(schedule: list, current_username: str, today,
                          horizon_weeks: int = SWAP_HORIZON_WEEKS,
                          limit: int = SWAP_CANDIDATES_LIMIT) -> list[dict]:
    """Return up to `limit` swap candidates for `current_username`.

    Candidate shape:
      {"kind": "vacancy"|"counterparty", "date": "YYYY-MM-DD", "day": "Mon"|"Wed",
       "slot": int, "b_username": str|None, "b_user_id": int|None, "b_name": str|None}

    Logic:
      - Horizon: sessions where session_date in [today, today+horizon_weeks*7].
      - Exclude A's own session DATES - every date A already speaks on, not
        A's whole week. The line here read "A's own current week", which the
        code has never done: if A speaks Monday of week 3, Wednesday of week 3
        is offered, and a Mon->Wed move inside one week is a real thing to
        want. The docstring was the liar, so it is the thing corrected. If
        week-level exclusion is wanted instead, that is a behaviour change to
        the live /swap flow and the operator's call, not a docstring fix.
      - Vacancies first (sessions with <3 filled slots OR entries with null speaker),
        ordered by date proximity.
      - Then counterparties (other speakers in future sessions, one per session,
        slot 1 preferred), ordered by date proximity.
      - Returns at most `limit` candidates total.
    """
    from datetime import date as _date, timedelta
    from collections import defaultdict

    horizon = today + timedelta(weeks=horizon_weeks)
    uname_lc = current_username.lower()
    by_date = defaultdict(list)
    for e in schedule:
        try:
            d = _date.fromisoformat(e["session_date"])
        except (ValueError, KeyError):
            continue
        if today <= d <= horizon:
            by_date[e["session_date"]].append(e)

    # Identify A's own session dates so we don't propose swapping into them.
    a_dates = {e["session_date"] for e in schedule
               if (e.get("speaker_username") or "").lower() == uname_lc}

    vacancies = []
    counterparties = []
    for date_iso in sorted(by_date.keys()):
        if date_iso in a_dates:
            continue
        entries = by_date[date_iso]
        day = entries[0].get("day", "")
        # A slot is filled when it has a SPEAKER, bound or not. Keying on
        # `speaker_username` treated every unresolved name -- exactly what
        # `build_schedule()` writes when a roster lookup misses -- as an open
        # slot, so /swap offered a session that already had a speaker.
        filled_slots = {e["slot"] for e in entries
                        if e.get("speaker_username") or e.get("speaker_name")}
        # Vacancy: missing slot number, or entry with null speaker
        for slot_num in (1, 2, 3):
            if slot_num not in filled_slots:
                vacancies.append({
                    "kind": "vacancy",
                    "date": date_iso,
                    "day": day,
                    "slot": slot_num,
                    "b_username": None,
                    "b_user_id": None,
                    "b_name": None,
                })
                break  # one vacancy per session is enough for picker
        else:
            # Session is full - pick a counterparty (lowest slot number)
            counterparty_entries = sorted(
                [e for e in entries if (e.get("speaker_username") or "").lower() != uname_lc],
                key=lambda e: e["slot"],
            )
            if counterparty_entries:
                b = counterparty_entries[0]
                counterparties.append({
                    "kind": "counterparty",
                    "date": date_iso,
                    "day": day,
                    "slot": b["slot"],
                    "b_username": b.get("speaker_username"),
                    "b_user_id": None,  # resolved later via roster
                    "b_name": b.get("speaker_name"),
                })

    combined = vacancies + counterparties
    return combined[:limit]


def _resolve_user_id(username: str) -> Optional[int]:
    """Look up a Tribe member's Telegram user_id by username (case-insensitive)."""
    if not username:
        return None
    roster = load_state(TRIBE_ROSTER) or {}
    uname_lc = username.lower()
    for k, v in roster.items():
        if k.lower() == uname_lc:
            uid = v.get("telegram_user_id")
            return int(uid) if uid else None
    return None


def _append_swap_event(payload: dict) -> None:
    """Append one event to swap-requests.jsonl with ts auto-set."""
    enriched = {"ts": local_now().isoformat()}
    enriched.update(payload)
    append_jsonl(SWAP_REQUESTS_LOG, enriched)


def _load_swap_requests() -> dict:
    """Return {rid: list[event]} from the swap-requests JSONL."""
    from collections import defaultdict
    by_rid = defaultdict(list)
    path = state_path(SWAP_REQUESTS_LOG)
    if not path.exists():
        return {}
    # `_read_jsonl_rows`, not a fourth hand-rolled copy of the same walk. The
    # copies each decoded the whole file inside `for line in f`, outside their
    # own `except json.JSONDecodeError`, so a torn append raised out of the swap
    # state machine entirely.
    for e in _read_jsonl_rows(path):
        rid = e.get("rid")
        if rid:
            by_rid[rid].append(e)
    return dict(by_rid)


def _swap_request_status(events: list) -> str:
    """Derive current status from an ordered list of events for one rid."""
    if not events:
        return "unknown"
    return events[-1].get("event", "unknown")


def _swap_request_context(events: list) -> dict:
    """Merge all events for one rid into a single dict (last-wins). Use for read-only lookup."""
    ctx: dict = {}
    for e in events:
        ctx.update(e)
    return ctx


def _apply_vacancy_swap(a_username: str, a_current_date: str, a_current_slot: int,
                        target_date: str, target_slot: int, target_day: str,
                        theme_target: str, week_target: int) -> bool:
    """Move A from (a_current_date, a_current_slot) into the vacancy (target_date, target_slot).

    Mutates schedule.json atomically via save_state. Returns True on success.

    Side effects:
      - The entry at A's old slot is removed (creates a vacancy at the old date).
      - A new entry is appended for the target slot with A's name + swapped_with metadata.
    """
    # Under the lock, and the revalidation below is why: the read, the
    # occupancy check and the write have to be one indivisible step or the
    # check is answering a question about a schedule that no longer exists.
    with locked_state(SCHEDULE, []) as schedule:
        ok = _vacancy_swap_locked(schedule, a_username, a_current_date, a_current_slot,
                                  target_date, target_slot, target_day,
                                  theme_target, week_target)
    return ok


def _vacancy_swap_locked(schedule: list, a_username: str, a_current_date: str,
                         a_current_slot: int, target_date: str, target_slot: int,
                         target_day: str, theme_target: str, week_target: int) -> bool:
    """The body of `_apply_vacancy_swap`, mutating `schedule` IN PLACE.

    Split out so the lock in the caller spans the whole read-modify-write.
    """
    uname_lc = a_username.lower()
    # Find A's entry to lift name + display name
    a_entry = None
    new_schedule = []
    for e in schedule:
        is_a_current = (
            (e.get("speaker_username") or "").lower() == uname_lc
            and e.get("session_date") == a_current_date
            and e.get("slot") == a_current_slot
        )
        if is_a_current:
            a_entry = e
            continue  # drop A from old slot
        new_schedule.append(e)
    if a_entry is None:
        log_error(f"_apply_vacancy_swap: A's slot not found ({a_username} @ {a_current_date} #{a_current_slot})")
        return False
    # Revalidate the TARGET against the schedule as it is now, not as it was
    # when the keyboard was built. The candidate list is a snapshot: A can open
    # /swap, someone else fills the vacancy, and A taps a button that was
    # correct a minute ago. Appending without this check double-booked the slot
    # and neither speaker was told.
    occupant = next(
        (e for e in new_schedule
         if e.get("session_date") == target_date and e.get("slot") == target_slot
         and (e.get("speaker_username") or e.get("speaker_name"))),
        None,
    )
    if occupant is not None:
        log_error(
            f"_apply_vacancy_swap: {target_date} #{target_slot} was taken by "
            f"{occupant.get('speaker_name') or occupant.get('speaker_username')} "
            f"before @{a_username} tapped; refusing to double-book"
        )
        return False
    new_schedule.append({
        "cycle": a_entry.get("cycle", 1),
        "week": week_target,
        "session_date": target_date,
        "day": target_day,
        "theme": theme_target,
        "slot": target_slot,
        "speaker_name": a_entry.get("speaker_name"),
        "speaker_username": a_entry.get("speaker_username"),
        "swapped_with": {
            "with_username": None,  # vacancy fill, no counterparty
            "old_date": a_current_date,
            "old_slot": a_current_slot,
            "swapped_at": local_now().isoformat(),
        },
        "no_show": False,
        "completed": False,
    })
    # In place: the caller holds the lock and saves what this leaves behind.
    schedule[:] = new_schedule
    return True


def _apply_bilateral_swap(a_username: str, a_current_date: str, a_current_slot: int,
                          b_username: str, b_date: str, b_slot: int) -> bool:
    """Swap A's (date, slot) with B's (date, slot). Atomic.

    Both entries are updated in-place: speaker_name/username swap between them,
    and both get `swapped_with` metadata pointing to the other.
    """
    # Under the lock: two callbacks racing this used to both find their
    # entries, both mutate a private copy, and the later save silently undid the
    # earlier swap.
    with locked_state(SCHEDULE, []) as schedule:
        return _bilateral_swap_locked(schedule, a_username, a_current_date,
                                      a_current_slot, b_username, b_date, b_slot)


def _bilateral_swap_locked(schedule: list, a_username: str, a_current_date: str,
                           a_current_slot: int, b_username: str, b_date: str,
                           b_slot: int) -> bool:
    """The body of `_apply_bilateral_swap`, mutating `schedule` IN PLACE."""
    uname_lc_a = a_username.lower()
    uname_lc_b = (b_username or "").lower()
    a_idx = b_idx = None
    for i, e in enumerate(schedule):
        if (a_idx is None
                and (e.get("speaker_username") or "").lower() == uname_lc_a
                and e.get("session_date") == a_current_date
                and e.get("slot") == a_current_slot):
            a_idx = i
        elif (b_idx is None
                and (e.get("speaker_username") or "").lower() == uname_lc_b
                and e.get("session_date") == b_date
                and e.get("slot") == b_slot):
            b_idx = i
    if a_idx is None or b_idx is None:
        log_error(f"_apply_bilateral_swap: entries not found "
                  f"(a={a_username}@{a_current_date}#{a_current_slot} a_idx={a_idx}, "
                  f"b={b_username}@{b_date}#{b_slot} b_idx={b_idx})")
        return False

    swap_ts = local_now().isoformat()
    a_entry = schedule[a_idx]
    b_entry = schedule[b_idx]

    # Swap speaker identity between the two entries; preserve session metadata
    # (date, slot, week, theme) on each row.
    a_name, a_user = a_entry.get("speaker_name"), a_entry.get("speaker_username")
    b_name, b_user = b_entry.get("speaker_name"), b_entry.get("speaker_username")
    a_entry["speaker_name"] = b_name
    a_entry["speaker_username"] = b_user
    a_entry["swapped_with"] = {
        "with_username": a_user,  # B's row now records that the previous occupant was A
        "old_date": b_date,
        "old_slot": b_slot,
        "swapped_at": swap_ts,
    }
    b_entry["speaker_name"] = a_name
    b_entry["speaker_username"] = a_user
    b_entry["swapped_with"] = {
        "with_username": b_user,
        "old_date": a_current_date,
        "old_slot": a_current_slot,
        "swapped_at": swap_ts,
    }
    # Entries are mutated in place above; the caller holds the lock and saves.
    return True


def _sweep_expired_swap_requests(bot: TelegramBot) -> None:
    """Walk swap-requests.jsonl, expire any `proposed_to_b` past deadline.

    Lazy: runs at the top of every `_handle_update`. Performance is fine while
    the log stays small; if it grows past ~10k entries, switch to a tail-only
    read or move to a dedicated daemon refresher.
    """
    from datetime import datetime as _dt
    swap_requests = _load_swap_requests()
    now = local_now()
    for rid, events in swap_requests.items():
        status = _swap_request_status(events)
        if status != "proposed_to_b":
            continue
        ctx = _swap_request_context(events)
        deadline_iso = ctx.get("deadline")
        if not deadline_iso:
            continue
        try:
            deadline = _dt.fromisoformat(deadline_iso)
        except ValueError:
            continue
        if now <= deadline:
            continue
        # Expired - notify A and Misha, mark terminal
        _append_swap_event({"rid": rid, "event": "expired",
                            "expired_at": now.isoformat()})
        a_user_id = ctx.get("a_user_id")
        a_username = ctx.get("a_username", "")
        b_username = ctx.get("b_username", "")
        try:
            if a_user_id:
                bot.send_message(
                    int(a_user_id),
                    # The constant, not a literal "24h". The proposal message to
                    # B interpolates SWAP_B_RESPONSE_TTL_HOURS and these two did
                    # not, so changing the window would leave both expiry
                    # messages stating a deadline the code no longer uses.
                    f"Your swap request to @{b_username} expired (no response in "
                    f"{SWAP_B_RESPONSE_TTL_HOURS}h). "
                    f"I'll let Misha know - he'll arrange another date with you.",
                    parse_mode="",
                )
        except TelegramAPIError:
            pass
        misha_id = misha_user_id()
        if misha_id:
            try:
                bot.send_message(
                    misha_id,
                    f"Swap request expired: @{a_username} -> @{b_username} "
                    f"(no response in {SWAP_B_RESPONSE_TTL_HOURS}h). "
                    f"A will need manual help.",
                    parse_mode="",
                )
            except TelegramAPIError:
                pass


def _swap_kickoff_for_a(bot: TelegramBot, user_id: int, username: str) -> None:
    """Run the /swap command flow for user A. Either auto-fallback or present 2 buttons.

    `username` MUST be the canonical roster-key username resolved from the
    authorized user_id (see _resolve_my_username), never a self-reported Telegram
    handle. The schedule is keyed by roster username, so passing a spoofable
    handle here would let a caller operate on another member's slot.
    """
    from datetime import date as _date
    schedule = load_state(SCHEDULE) or []
    today = _today_local_date()
    a_slot = _user_current_slot(schedule, username, today)
    if not a_slot:
        bot.send_message(
            user_id,
            "You have no upcoming speaker slots in this cycle, so there's nothing to swap. "
            "If this looks wrong, message Misha.",
            parse_mode="",
        )
        return

    candidates = find_swap_candidates(schedule, username, today)
    if not candidates:
        # Fallback to legacy manual flow
        bot.send_message(
            user_id,
            "No open slots in the next 4 weeks. I'll let Misha know you'd like to swap - "
            "he'll reach out shortly to arrange a different date.",
            parse_mode="",
        )
        misha_id = misha_user_id()
        if misha_id:
            try:
                bot.send_message(
                    misha_id,
                    f"/swap from @{username} (id={user_id}) - "
                    f"no auto-candidates available. Currently scheduled "
                    f"{a_slot['session_date']} #{a_slot['slot']}. Reach out manually.",
                    parse_mode="",
                )
            except TelegramAPIError:
                pass
        _log_event("swap_requested", user_id=user_id, username=username,
                   outcome="no_candidates")
        return

    # Resolve B user_ids for counterparty candidates (best-effort; None means we can't DM)
    for c in candidates:
        if c["kind"] == "counterparty" and c.get("b_username"):
            c["b_user_id"] = _resolve_user_id(c["b_username"])

    rid = _new_request_id()
    # Build inline keyboard
    buttons = []
    for idx, c in enumerate(candidates):
        label_date = _format_dm_date(c["date"], c["day"])
        if c["kind"] == "vacancy":
            label = f"📅 {label_date} (open slot)"
        else:
            label = f"🔄 {label_date} - swap with @{c['b_username']}"
        buttons.append([{"text": label, "callback_data": f"sw:a:{rid}:{idx}"}])
    buttons.append([{"text": "Cancel", "callback_data": f"sw:a:{rid}:x"}])
    reply_markup = {"inline_keyboard": buttons}

    preamble = (f"You're currently on {_format_dm_date(a_slot['session_date'], a_slot['day'])}, "
                f"slot {a_slot['slot']}.\n\nPick a new date:")
    sent = bot.send_message(user_id, preamble, parse_mode="", reply_markup=reply_markup)
    a_message_id = sent.get("message_id") if isinstance(sent, dict) else None

    _append_swap_event({
        "rid": rid,
        "event": "initiated",
        "a_user_id": user_id,
        "a_username": username,
        "a_current_date": a_slot["session_date"],
        "a_current_slot": a_slot["slot"],
        "a_message_id": a_message_id,
        "candidates": candidates,
    })
    _log_event("swap_requested", user_id=user_id, username=username,
               rid=rid, candidates_count=len(candidates))


def _handle_callback_query(bot: TelegramBot, cq: dict) -> None:
    """Route inline-keyboard taps. Only `sw:*` payloads are handled here."""
    cq_id = cq.get("id")
    data = (cq.get("data") or "").strip()
    user = cq.get("from", {})
    tapper_user_id = user.get("id")
    tapper_username = (user.get("username") or "").lower()
    msg = cq.get("message", {}) or {}
    msg_chat_id = (msg.get("chat") or {}).get("id")
    msg_id = msg.get("message_id")

    # Topic feature: CEO approval of the cycle-end invite. Handled before the
    # sw:* gate. Only the CEO (MISHA_TELEGRAM_USER_ID) may approve/cancel.
    if data.startswith("cycle_invite:"):
        _handle_cycle_invite_tap(bot, cq_id, data, tapper_user_id, msg_chat_id, msg_id)
        return

    if not data.startswith("sw:"):
        # Not our domain - dismiss the spinner silently
        if cq_id:
            try:
                bot.answer_callback_query(cq_id)
            except TelegramAPIError:
                pass
        return

    # Authorization: only active Tribe members may tap swap buttons
    if not _is_authorized_user(tapper_user_id, username=tapper_username):
        if cq_id:
            try:
                bot.answer_callback_query(cq_id, text="Not authorized.")
            except TelegramAPIError:
                pass
        return

    parts = data.split(":")
    # Expected forms:
    #   sw:a:<rid>:<idx-or-x>
    #   sw:b:<rid>:<y|n>
    if len(parts) != 4:
        try:
            bot.answer_callback_query(cq_id, text="Malformed request.")
        except TelegramAPIError:
            pass
        return

    _, role, rid, choice = parts
    swap_requests = _load_swap_requests()
    events = swap_requests.get(rid, [])
    if not events:
        try:
            bot.answer_callback_query(cq_id, text="Request not found.")
            if msg_chat_id and msg_id:
                bot.edit_message_reply_markup(msg_chat_id, msg_id, None)
        except TelegramAPIError:
            pass
        return

    status = _swap_request_status(events)
    ctx = _swap_request_context(events)

    if role == "a":
        _handle_a_tap(bot, cq_id, rid, choice, ctx, status, msg_chat_id, msg_id, tapper_user_id)
    elif role == "b":
        _handle_b_tap(bot, cq_id, rid, choice, ctx, status, msg_chat_id, msg_id, tapper_user_id)
    else:
        try:
            bot.answer_callback_query(cq_id, text="Unknown action.")
        except TelegramAPIError:
            pass


def _handle_cycle_invite_tap(bot: TelegramBot, cq_id, data: str,
                             tapper_user_id, msg_chat_id, msg_id) -> None:
    """Process the CEO's tap on the cycle-end invite draft (send | cancel)."""
    misha_id = misha_user_id()
    if not misha_id or tapper_user_id != misha_id:
        if cq_id:
            try:
                bot.answer_callback_query(cq_id, text="Only Misha can approve this.")
            except TelegramAPIError:
                pass
        return

    state = ft.load_topic_state(state_dir())
    pending = state.get("pending_cycle_invite")
    if not pending:
        if cq_id:
            try:
                bot.answer_callback_query(cq_id, text="No pending invite.")
                if msg_chat_id and msg_id:
                    bot.edit_message_reply_markup(msg_chat_id, msg_id, None)
            except TelegramAPIError:
                pass
        return

    # The tapped card must be the card this draft was drafted onto. Telegram
    # keeps an old inline keyboard tappable forever, and `cycle-end-invite`
    # overwrites `pending_cycle_invite` when a new cycle comes round, so the
    # cycle-1 approval message stayed in the CEO's history showing cycle-1 text
    # while the pending draft underneath it had become cycle 2. Tapping "Send to
    # Tribe" on that card posted the cycle-2 invite: the text on screen and the
    # text sent were different, on the one flow whose whole purpose is the CEO
    # approving exact wording. Refused now, and the dead keyboard is cleared so
    # the card cannot be tapped again.
    approval_msg_id = pending.get("approval_msg_id")
    if approval_msg_id is not None and msg_id is not None and msg_id != approval_msg_id:
        _log_event("cycle_end_invite_stale_tap", cycle=pending.get("cycle"),
                   tapped_msg_id=msg_id, pending_msg_id=approval_msg_id)
        try:
            bot.answer_callback_query(
                cq_id, text="This draft was replaced by a newer one. Nothing was sent.")
            if msg_chat_id:
                bot.edit_message_reply_markup(msg_chat_id, msg_id, None)
        except TelegramAPIError:
            pass
        return

    choice = data.split(":", 1)[1]
    if choice == "cancel":
        state["pending_cycle_invite"] = None
        ft.save_topic_state(state_dir(), state)
        _log_event("cycle_end_invite_cancelled", cycle=pending.get("cycle"))
        try:
            bot.answer_callback_query(cq_id, text="Cancelled.")
            if msg_chat_id and msg_id:
                bot.edit_message_text(msg_chat_id, msg_id,
                                      "✖️ Cancelled — nothing was sent to the Tribe.",
                                      parse_mode="", reply_markup=None)
        except TelegramAPIError:
            pass
        return

    if choice == "send":
        try:
            chat_id = int(os.environ["FIRESIDE_TRIBE_CHAT_ID"])
            result = bot.send_message(chat_id, pending["text"])
        except (TelegramAPIError, KeyError, ValueError) as e:
            try:
                bot.answer_callback_query(cq_id, text="Send failed — see logs.")
            except TelegramAPIError:
                pass
            log_error(f"cycle-end-invite send failed: {e}")
            _log_event("cycle_end_invite_send_failed", cycle=pending.get("cycle"), error=str(e))
            return
        # Post succeeded — clear pending immediately so a re-tap cannot double-post.
        state["pending_cycle_invite"] = None
        ft.save_topic_state(state_dir(), state)
        # Pin is best-effort; a pin failure must not revert the cleared state.
        # But best-effort is not the same as unreported. The failure was
        # swallowed by a bare `pass` and the CEO's card was then rewritten to
        # say "Sent to the Tribe and pinned." whatever happened - a failure
        # reported as a success, on the one screen the operator reads to decide
        # whether anything is left to do. `cmd_sunday_preview` handles the
        # identical post-then-pin shape correctly: it tracks the outcome and
        # names the failure. The fix landed in one of the two copies.
        pinned = True
        try:
            bot.pin_chat_message(chat_id, result.get("message_id"), disable_notification=True)
        except TelegramAPIError as e:
            pinned = False
            log_error(f"cycle-end-invite posted but the pin failed: {e}")
        _log_event("cycle_end_invite_sent", cycle=pending.get("cycle"),
                   message_id=result.get("message_id"), pinned=pinned)
        card = ("✅ Sent to the Tribe and pinned." if pinned else
                "✅ Sent to the Tribe. ⚠️ The pin FAILED - pin it by hand. "
                "Do not re-tap: the message is already in the group.")
        try:
            bot.answer_callback_query(cq_id, text="Sent to the Tribe.")
            if msg_chat_id and msg_id:
                bot.edit_message_text(msg_chat_id, msg_id, card,
                                      parse_mode="", reply_markup=None)
        except TelegramAPIError:
            pass
        return

    # Unknown choice
    try:
        bot.answer_callback_query(cq_id, text="Unknown action.")
    except TelegramAPIError:
        pass


def _handle_a_tap(bot: TelegramBot, cq_id: str, rid: str, choice: str,
                  ctx: dict, status: str, msg_chat_id, msg_id,
                  tapper_user_id: int) -> None:
    """Process A's button tap. choice is '0', '1', ..., or 'x' (cancel)."""
    # Only the original A may tap
    if ctx.get("a_user_id") != tapper_user_id:
        try:
            bot.answer_callback_query(cq_id, text="This button is for someone else.")
        except TelegramAPIError:
            pass
        return

    if status != "initiated":
        try:
            bot.answer_callback_query(cq_id, text="This request is already closed.")
            if msg_chat_id and msg_id:
                bot.edit_message_reply_markup(msg_chat_id, msg_id, None)
        except TelegramAPIError:
            pass
        return

    if choice == "x":
        _append_swap_event({"rid": rid, "event": "cancelled_by_a"})
        try:
            bot.answer_callback_query(cq_id, text="Cancelled.")
            if msg_chat_id and msg_id:
                bot.edit_message_text(msg_chat_id, msg_id,
                                      "Swap request cancelled. No changes made.",
                                      parse_mode="", reply_markup=None)
        except TelegramAPIError:
            pass
        return

    try:
        idx = int(choice)
    except ValueError:
        try:
            bot.answer_callback_query(cq_id, text="Bad choice.")
        except TelegramAPIError:
            pass
        return

    candidates = ctx.get("candidates") or []
    if idx < 0 or idx >= len(candidates):
        try:
            bot.answer_callback_query(cq_id, text="Choice out of range.")
        except TelegramAPIError:
            pass
        return

    chosen = candidates[idx]
    a_username = ctx.get("a_username", "")
    a_current_date = ctx.get("a_current_date")
    a_current_slot = ctx.get("a_current_slot")

    if chosen["kind"] == "vacancy":
        # Auto-apply immediately; no counterparty consent needed
        # Look up theme/week for the target session by reading any existing entry there
        schedule = load_state(SCHEDULE) or []
        target_week = None
        target_theme = None
        for e in schedule:
            if e.get("session_date") == chosen["date"]:
                target_week = e.get("week")
                target_theme = e.get("theme")
                break
        if target_week is None or target_theme is None:
            try:
                bot.answer_callback_query(cq_id, text="Target session metadata missing.")
            except TelegramAPIError:
                pass
            log_error(f"swap rid={rid}: vacancy target {chosen['date']} has no metadata")
            return

        ok = _apply_vacancy_swap(
            a_username=a_username,
            a_current_date=a_current_date,
            a_current_slot=a_current_slot,
            target_date=chosen["date"],
            target_slot=chosen["slot"],
            target_day=chosen["day"],
            theme_target=target_theme,
            week_target=target_week,
        )
        if not ok:
            try:
                bot.answer_callback_query(cq_id, text="Could not apply swap. Try again or message Misha.")
            except TelegramAPIError:
                pass
            return

        _append_swap_event({
            "rid": rid, "event": "a_tapped_vacancy", "chosen_idx": idx,
            "target_date": chosen["date"], "target_slot": chosen["slot"],
        })
        _append_swap_event({
            "rid": rid, "event": "completed", "outcome": "vacancy_fill",
            "freed_date": a_current_date, "freed_slot": a_current_slot,
        })
        label_date = _format_dm_date(chosen["date"], chosen["day"])
        try:
            bot.answer_callback_query(cq_id, text="Done.")
            if msg_chat_id and msg_id:
                bot.edit_message_text(
                    msg_chat_id, msg_id,
                    f"Done. You're now on {label_date}, slot {chosen['slot']}. "
                    f"Your previous slot ({a_current_date}) is now open.",
                    parse_mode="", reply_markup=None,
                )
        except TelegramAPIError:
            pass
        misha_id = misha_user_id()
        if misha_id:
            try:
                bot.send_message(
                    misha_id,
                    f"/swap done (vacancy): @{a_username} moved "
                    f"{a_current_date} #{a_current_slot} -> {chosen['date']} #{chosen['slot']}. "
                    f"Freed slot at {a_current_date} #{a_current_slot}.",
                    parse_mode="",
                )
            except TelegramAPIError:
                pass
        _log_event("swap_completed", rid=rid, outcome="vacancy_fill",
                   a_username=a_username,
                   from_date=a_current_date, from_slot=a_current_slot,
                   to_date=chosen["date"], to_slot=chosen["slot"])
        return

    # Counterparty path: propose to B
    b_username = chosen.get("b_username")
    b_user_id = chosen.get("b_user_id")
    if not b_user_id:
        try:
            bot.answer_callback_query(cq_id,
                text=f"@{b_username} hasn't started the bot yet - I can't DM them. Misha will help.")
        except TelegramAPIError:
            pass
        # Fall back to manual
        misha_id = misha_user_id()
        if misha_id:
            try:
                bot.send_message(
                    misha_id,
                    f"/swap stuck: @{a_username} wants to swap with @{b_username} "
                    f"({chosen['date']} #{chosen['slot']}) but B has no telegram_user_id. "
                    f"Reach out manually.",
                    parse_mode="",
                )
            except TelegramAPIError:
                pass
        _append_swap_event({"rid": rid, "event": "b_unreachable", "b_username": b_username})
        if msg_chat_id and msg_id:
            try:
                bot.edit_message_text(
                    msg_chat_id, msg_id,
                    f"@{b_username} hasn't activated the bot yet. Misha will arrange this one manually.",
                    parse_mode="", reply_markup=None,
                )
            except TelegramAPIError:
                pass
        return

    from datetime import timedelta
    deadline = local_now() + timedelta(hours=SWAP_B_RESPONSE_TTL_HOURS)
    a_label = _format_dm_date(a_current_date, ctx.get("a_day") or "")
    b_label = _format_dm_date(chosen["date"], chosen["day"])
    b_text = (
        f"@{a_username} would like to swap fireside slots with you.\n\n"
        f"They're on {a_label} (slot {a_current_slot}); "
        f"you're on {b_label} (slot {chosen['slot']}).\n\n"
        f"If you accept, you'll move to {a_label} and they'll take your {b_label} slot. "
        f"This request expires in {SWAP_B_RESPONSE_TTL_HOURS}h."
    )
    b_buttons = {"inline_keyboard": [[
        {"text": "✅ Accept", "callback_data": f"sw:b:{rid}:y"},
        {"text": "❌ Decline", "callback_data": f"sw:b:{rid}:n"},
    ]]}
    try:
        sent_b = bot.send_message(int(b_user_id), b_text, parse_mode="", reply_markup=b_buttons)
        b_msg_id = sent_b.get("message_id") if isinstance(sent_b, dict) else None
    except TelegramAPIError as e:
        log_error(f"swap rid={rid}: failed to DM B (@{b_username}): {e}")
        try:
            bot.answer_callback_query(cq_id, text="Could not reach the other speaker. Misha will help.")
        except TelegramAPIError:
            pass
        # Misha is actually told, as the b_unreachable branch above already
        # does. This path promised A that "Misha will help" and then wrote one
        # errors.log line nobody is alerted to - the most likely cause is B
        # having blocked the bot, which no other surface reports.
        _tell(bot, misha_user_id(),
              f"/swap stuck: @{a_username} wants to swap with @{b_username} "
              f"({chosen['date']} #{chosen['slot']}) but the DM to B failed "
              f"({e}). Most likely B blocked the bot. Reach out manually.")
        # No event is appended and the message is NOT edited, deliberately: the
        # status stays where it was, so A's buttons remain live and A can retap.
        # Appending here would make this the last event and therefore the
        # status, ending a request that is still recoverable.
        return

    _append_swap_event({
        "rid": rid, "event": "a_tapped_counterparty", "chosen_idx": idx,
        "target_date": chosen["date"], "target_slot": chosen["slot"],
        "b_username": b_username, "b_user_id": int(b_user_id),
    })
    _append_swap_event({
        "rid": rid, "event": "proposed_to_b",
        "b_user_id": int(b_user_id), "b_message_id": b_msg_id,
        "deadline": deadline.isoformat(),
    })

    try:
        bot.answer_callback_query(cq_id, text="Sent to the other speaker.")
        if msg_chat_id and msg_id:
            bot.edit_message_text(
                msg_chat_id, msg_id,
                f"Request sent to @{b_username}. They have {SWAP_B_RESPONSE_TTL_HOURS}h "
                f"to accept or decline. I'll let you know.",
                parse_mode="", reply_markup=None,
            )
    except TelegramAPIError:
        pass


def _handle_b_tap(bot: TelegramBot, cq_id: str, rid: str, choice: str,
                  ctx: dict, status: str, msg_chat_id, msg_id,
                  tapper_user_id: int) -> None:
    """Process B's accept/decline tap. choice is 'y' or 'n'."""
    expected_b = ctx.get("b_user_id")
    if expected_b is None or int(expected_b) != tapper_user_id:
        try:
            bot.answer_callback_query(cq_id, text="This button is for someone else.")
        except TelegramAPIError:
            pass
        return

    if status != "proposed_to_b":
        try:
            bot.answer_callback_query(cq_id, text="This request is already closed.")
            if msg_chat_id and msg_id:
                bot.edit_message_reply_markup(msg_chat_id, msg_id, None)
        except TelegramAPIError:
            pass
        return

    a_username = ctx.get("a_username", "")
    a_user_id = ctx.get("a_user_id")
    a_current_date = ctx.get("a_current_date")
    a_current_slot = ctx.get("a_current_slot")
    b_username = ctx.get("b_username", "")
    target_date = ctx.get("target_date")
    target_slot = ctx.get("target_slot")

    if choice == "n":
        _append_swap_event({"rid": rid, "event": "b_declined"})
        try:
            bot.answer_callback_query(cq_id, text="Declined.")
            if msg_chat_id and msg_id:
                bot.edit_message_text(
                    msg_chat_id, msg_id,
                    "You declined the swap. No changes made.",
                    parse_mode="", reply_markup=None,
                )
        except TelegramAPIError:
            pass
        # Notify A and Misha
        if a_user_id:
            try:
                bot.send_message(
                    int(a_user_id),
                    f"@{b_username} declined the swap. Your slot stays at "
                    f"{a_current_date} #{a_current_slot}. Reply /swap to try another date, "
                    f"or message Misha for help.",
                    parse_mode="",
                )
            except TelegramAPIError:
                pass
        misha_id = misha_user_id()
        if misha_id:
            try:
                bot.send_message(
                    misha_id,
                    f"/swap declined: @{b_username} said no to @{a_username} "
                    f"({a_current_date} <-> {target_date}).",
                    parse_mode="",
                )
            except TelegramAPIError:
                pass
        return

    if choice != "y":
        try:
            bot.answer_callback_query(cq_id, text="Bad choice.")
        except TelegramAPIError:
            pass
        return

    # Accepted - apply bilateral swap atomically
    ok = _apply_bilateral_swap(
        a_username=a_username,
        a_current_date=a_current_date,
        a_current_slot=a_current_slot,
        b_username=b_username,
        b_date=target_date,
        b_slot=target_slot,
    )
    if not ok:
        try:
            bot.answer_callback_query(cq_id,
                text="Could not apply swap (entries changed since request). Misha will help.")
        except TelegramAPIError:
            pass
        _append_swap_event({"rid": rid, "event": "apply_failed"})
        # A and Misha are told, like every other terminal outcome of this flow.
        # They were not: B saw "Misha will help" and the only other effect was
        # this event. A's last message was "Request sent... I'll let you know",
        # so A waited forever; Misha never heard that a swap he was promised to
        # fix had failed; and `apply_failed` is terminal, so the expiry sweep -
        # which only looks at `proposed_to_b` - never reached it either. Three
        # silent paths out of the one branch that says a human will step in.
        #
        # The request is NOT re-opened. It failed because the schedule rows it
        # named are gone, so a fresh 24h window would re-propose a swap against
        # entries that no longer exist.
        _tell(bot, a_user_id,
              f"@{b_username} accepted, but the swap could not be applied: "
              f"the schedule changed since you asked. Your slot stays at "
              f"{a_current_date} #{a_current_slot}. Misha has been told and "
              f"will sort it out with you.")
        _tell(bot, misha_user_id(),
              f"/swap FAILED to apply: @{b_username} accepted @{a_username}'s "
              f"request ({a_current_date} #{a_current_slot} <-> {target_date} "
              f"#{target_slot}) but the schedule rows no longer match. Both were "
              f"told you would help. Nothing was changed.")
        _log_event("swap_failed", rid=rid, outcome="apply_failed",
                   a_username=a_username, b_username=b_username,
                   from_date=a_current_date, from_slot=a_current_slot,
                   to_date=target_date, to_slot=target_slot)
        return

    _append_swap_event({
        "rid": rid, "event": "b_accepted",
    })
    _append_swap_event({
        "rid": rid, "event": "completed", "outcome": "bilateral_swap",
    })
    a_label = _format_dm_date(a_current_date, "")
    b_label = _format_dm_date(target_date, "")
    try:
        bot.answer_callback_query(cq_id, text="Accepted. Swap applied.")
        if msg_chat_id and msg_id:
            bot.edit_message_text(
                msg_chat_id, msg_id,
                f"Accepted. You're now on {a_label}, slot {a_current_slot}. "
                f"@{a_username} takes {b_label}, slot {target_slot}.",
                parse_mode="", reply_markup=None,
            )
    except TelegramAPIError:
        pass
    if a_user_id:
        try:
            bot.send_message(
                int(a_user_id),
                f"@{b_username} accepted. You're now on {b_label}, slot {target_slot}. "
                f"Your old slot ({a_label}, slot {a_current_slot}) is now theirs.",
                parse_mode="",
            )
        except TelegramAPIError:
            pass
    misha_id = misha_user_id()
    if misha_id:
        try:
            bot.send_message(
                misha_id,
                f"/swap done (bilateral): @{a_username} <-> @{b_username}. "
                f"@{a_username}: {a_current_date} #{a_current_slot} -> {target_date} #{target_slot}. "
                f"@{b_username}: {target_date} #{target_slot} -> {a_current_date} #{a_current_slot}.",
                parse_mode="",
            )
        except TelegramAPIError:
            pass
    _log_event("swap_completed", rid=rid, outcome="bilateral_swap",
               a_username=a_username, b_username=b_username,
               from_date=a_current_date, from_slot=a_current_slot,
               to_date=target_date, to_slot=target_slot)


def _handle_message(bot: TelegramBot, message: dict) -> None:
    """Handle private DMs to the bot. Routes /start, /swap, and query commands.

    Authorization gate: only users whose Telegram user_id maps to an active,
    non-excluded entry in tribe-roster.json may interact. Outsiders get a
    generic 'private bot' reply, are logged to sessions.jsonl, and forwarded
    to Misha at most once per hour per user_id.
    """
    chat = message.get("chat", {})
    if chat.get("type") != "private":
        return  # ignore group messages
    text = (message.get("text") or "").strip()
    user = message.get("from", {})
    user_id = user.get("id")
    username = (user.get("username") or "").lower()

    # /start: greet members already bound to the roster by the trusted
    # `bootstrap` (Telethon enumeration of the real group). We deliberately do
    # NOT bind telegram_user_id from the self-reported @username here -- a handle
    # is reclaimable, so trusting it would allow handle takeover. Unbound or
    # unknown senders get the private-bot reply and are forwarded to Misha, who
    # re-runs `bootstrap` to enroll them.
    if text.startswith("/start"):
        if _is_authorized_user(user_id, username=username):
            bot.send_message(user_id, WELCOME_DM, parse_mode="")
            _log_event("start_received", user_id=user_id, username=username,
                       matched_in_roster=True)
            return
        # Unbound or outsider /start - unauthorized path
        _log_event("unauthorized_start", user_id=user_id, username=username)
        _maybe_forward_outsider(bot, user_id, username, text)
        bot.send_message(user_id, UNAUTHORIZED_REPLY, parse_mode="")
        return

    # All other commands: gate by user_id in active+non-excluded roster
    if not _is_authorized_user(user_id, username=username):
        _log_event("unauthorized_dm", user_id=user_id, username=username,
                   text_preview=text[:200])
        _maybe_forward_outsider(bot, user_id, username, text)
        bot.send_message(user_id, UNAUTHORIZED_REPLY, parse_mode="")
        return

    # User is authorized - existing command dispatch
    if text.startswith("/help"):
        bot.send_message(user_id, HELP_DM, parse_mode="")
        return

    if ft._is_idea_command(text):
        body = ft.parse_idea_command(text)
        if body is None:
            bot.send_message(
                user_id,
                "Send your idea after the command, e.g.\n`/idea a real DPI incident, start to finish`",
            )
            return
        schedule = load_state(SCHEDULE) or []
        roster = load_state(TRIBE_ROSTER) or {}
        name = (_roster_entry(roster, _resolve_my_username(user_id)) or {}).get("name", "")
        ft.append_idea(
            state_dir(),
            now_iso=local_now().isoformat(),
            user_id=user_id, username=username, name=name,
            text=body, cycle=ft.current_cycle(schedule, _today_local_date()),
        )
        _log_event("idea_submitted", user_id=user_id, username=username,
                   text_preview=body[:120])
        bot.send_message(
            user_id,
            "Logged ✓ — thank you. Your idea goes into the pool we draw the next fireside topics from.",
        )
        return

    if text.startswith("/me"):
        bot.send_message(user_id, _cmd_me_text(user_id), parse_mode="")
        return

    if text.startswith("/next"):
        bot.send_message(user_id, _cmd_next_text(), parse_mode="")
        return

    if text.startswith("/who"):
        bot.send_message(user_id, _cmd_who_text(), parse_mode="")
        return

    if text.startswith("/theme"):
        bot.send_message(user_id, _cmd_theme_text(), parse_mode="")
        return

    if text.startswith("/schedule"):
        bot.send_message(user_id, _cmd_schedule_text(), parse_mode="")
        return

    if text.startswith("/zoom"):
        bot.send_message(user_id, _cmd_zoom_text(), parse_mode="")
        return

    if text.startswith("/swap"):
        # Identity by user_id, not the self-reported @username: the schedule is
        # keyed by roster username, so resolving from user_id stops a caller who
        # has claimed a former member's handle from swapping that member's slot.
        my_username = _resolve_my_username(user_id)
        if not my_username:
            bot.send_message(
                user_id,
                "Couldn't find you in the Tribe roster, so I can't set up a swap. "
                "Reply here and Misha will sort it out.",
                parse_mode="",
            )
            return
        _swap_kickoff_for_a(bot, user_id, my_username)
        return

    # Unrecognised message from authorized user - forward to Misha so the Tribe member feels heard
    misha_id = misha_user_id()
    if misha_id and text:
        try:
            preview = text[:300] + ("..." if len(text) > 300 else "")
            bot.send_message(
                misha_id,
                f"DM to bot from @{username} (id={user_id}):\n{preview}",
                parse_mode="",
            )
        except TelegramAPIError:
            pass
    bot.send_message(
        user_id,
        "Got it - Misha will see this. Type /help for the command menu.",
        parse_mode="",
    )


def _handle_message_reaction(event: dict) -> None:
    """Update opt-ins.json when a Tribe member adds or removes 🧭/🌟 on the launch announcement."""
    msg_id = event.get("message_id")
    expected_msg = int(os.environ.get("FIRESIDE_LAUNCH_ANNOUNCEMENT_MSG_ID", "0"))
    if msg_id != expected_msg:
        return  # only track reactions on the launch announcement

    user = event.get("user", {})
    user_id = user.get("id")
    username = (user.get("username") or "").lower() if user else ""
    if not user_id:
        return

    # Identity by user_id: only a bound, authorized Tribe member may opt in, and
    # we store the canonical roster-key username (not the self-reported handle),
    # so an outsider reacting cannot pollute opt-ins and a reclaimed handle never
    # makes a stored opt-in stale. Removal is keyed by user_id, so a member who
    # later becomes unbound can still remove their own opt-in.
    #
    # AUTHORIZED means what `_is_authorized_user` says it means: active AND not
    # excluded from the fireside. This gated on `_resolve_my_username` alone,
    # which matches any roster row carrying the user_id and asks neither
    # question, so a member offboarded by setting `active: false` - who keeps
    # their telegram_user_id - could still react and be added to the Helmsman
    # rota, appear in the CEO's candidate list, and be counted in the stats.
    # The comment claimed the stronger gate; the code used the weaker one.
    my_username = _resolve_my_username(user_id)
    if my_username is not None and not _is_authorized_user(user_id):
        my_username = None

    new_reactions = event.get("new_reaction", []) or []
    emojis = {r.get("emoji") for r in new_reactions if r.get("type") == "emoji"}

    opt_ins = load_state(OPT_INS) or {"helmsman": [], "wildcard": []}
    changed = False
    for emoji, key in [("🧭", "helmsman"), ("🌟", "wildcard")]:
        existing = next((x for x in opt_ins[key] if x.get("user_id") == user_id), None)
        if emoji in emojis and not existing:
            if my_username is None:
                continue  # not an authorized/bound member -- ignore the opt-in
            opt_ins[key].append({"user_id": user_id, "username": my_username})
            changed = True
        elif emoji not in emojis and existing:
            opt_ins[key] = [x for x in opt_ins[key] if x.get("user_id") != user_id]
            changed = True
    if changed:
        save_state(OPT_INS, opt_ins)
        _log_event("opt_in_changed", user_id=user_id,
                   username=my_username or username, reactions=list(emojis))


def _handle_chat_member(event: dict) -> None:
    """Track joiners and leavers in the 31C Tribe group."""
    expected_chat = int(os.environ.get("FIRESIDE_TRIBE_CHAT_ID", "0"))
    chat_id = event.get("chat", {}).get("id")
    if chat_id != expected_chat:
        return

    new_member = event.get("new_chat_member", {})
    old_member = event.get("old_chat_member", {})
    user = new_member.get("user", {})
    user_id = user.get("id")
    username = (user.get("username") or "").lower()
    new_status = new_member.get("status", "")
    old_status = old_member.get("status", "")

    if new_status in ("member", "administrator") and old_status in ("left", "kicked", ""):
        _log_event("tribe_join", user_id=user_id, username=username)
        misha_id = misha_user_id()
        if misha_id:
            try:
                bot = get_bot()
                bot.send_message(misha_id,
                    f"📥 New 31C Tribe member: @{username or '<no username>'} "
                    f"(id={user_id}). Add to xlsx with their function/title for the next bootstrap.")
            except TelegramAPIError:
                pass
    elif new_status in ("left", "kicked") and old_status in ("member", "administrator"):
        _log_event("tribe_leave", user_id=user_id, username=username)
        # Mark inactive in roster
        roster = load_state(TRIBE_ROSTER) or {}
        for k, v in list(roster.items()):
            if v.get("telegram_user_id") == user_id:
                v["active"] = False
                save_state(TRIBE_ROSTER, roster)
                break


# ============================================================
# Subcommand: speaker-dms (Phase 3 task 3.2)
# ============================================================

def cmd_speaker_dms(args) -> None:
    """Send 2-week and 3-day speaker reminders. Cron: daily 09:00 local time."""
    from datetime import date as _date, timedelta

    bot = get_bot()
    schedule = load_state(SCHEDULE) or []
    roster = load_state(TRIBE_ROSTER) or {}
    today = _today_local_date()
    dm_log_path = state_path(DM_LOG)

    sent_2wk = 0
    sent_3day = 0
    skipped = 0
    failed = 0

    for entry in schedule:
        username = entry.get("speaker_username")
        if not username:
            continue
        session_date = entry["session_date"]
        d = _date.fromisoformat(session_date)
        days_until = (d - today).days
        user_id = _resolve_speaker_user_id(roster, username)
        # Guarded exactly as `cmd_email_backup` guards its own greeting, and for
        # the reason that file already records: a missing key raised KeyError
        # and a blank one raised IndexError out of the MIDDLE of a send loop.
        # Speakers already reached kept their DM, everyone later in the schedule
        # got nothing, no summary printed, and `hc_ping` never fired - so the
        # healthcheck went red-by-silence at best and the partial send was
        # invisible. A handle is a worse greeting than a first name and a much
        # better one than a dead job.
        name = _first_name(entry.get("speaker_name"), f"@{username}")
        session_day = d.strftime("%a")
        theme = entry["theme"]

        for window, dm_type, template in [
            ((10, 14), "2wk", SPEAKER_DM_2WK),
            ((1, 3), "3day", SPEAKER_DM_3DAY),
        ]:
            if not (window[0] <= days_until <= window[1]):
                continue
            if _dm_already_sent(dm_log_path, username, dm_type, session_date):
                skipped += 1
                continue
            if user_id is None:
                _log_dm(dm_type, username, session_date, None, False,
                        error="no telegram_user_id (user has not /started bot)")
                failed += 1
                continue
            text = template.format(
                name=name, session_date=session_date, session_day=session_day, theme=theme,
            )
            try:
                bot.send_dm(user_id, text)
                _log_dm(dm_type, username, session_date, user_id, True)
                if dm_type == "2wk":
                    sent_2wk += 1
                else:
                    sent_3day += 1
            except TelegramAPIError as e:
                _log_dm(dm_type, username, session_date, user_id, False, error=str(e))
                failed += 1

    print(f"{GREEN}speaker-dms{RESET}: 2wk={sent_2wk} 3day={sent_3day} skipped={skipped} failed={failed}")
    hc_ping("FIRESIDE_HC_SPEAKER_DMS")


# ============================================================
# Subcommand: sunday-preview (Phase 3 task 3.3)
# ============================================================

def cmd_sunday_preview(args) -> None:
    """Post weekly preview to 31C Tribe group + pin. Cron: Sunday 18:00 local time."""
    schedule = load_state(SCHEDULE) or []
    helmsmen = load_state(HELMSMEN) or {}

    today = _today_local_date()
    week_num = _current_or_upcoming_week(schedule, today)
    if week_num is None:
        print(f"{YELLOW}sunday-preview: no upcoming sessions in schedule{RESET}")
        return

    mon = _week_speakers(schedule, week_num, "Mon")
    wed = _week_speakers(schedule, week_num, "Wed")
    if not mon or not wed:
        print(f"{RED}sunday-preview: incomplete week {week_num} in schedule{RESET}", file=sys.stderr)
        return

    week_start = mon[0]["session_date"]
    helmsman_entry = helmsmen.get(week_start, {})
    helmsman_name = helmsman_entry.get("name", "[Helmsman not yet assigned - please pick one]")

    monday_speakers = " · ".join(s["speaker_name"] for s in mon)
    wednesday_speakers = " · ".join(s["speaker_name"] for s in wed)

    text = SUNDAY_PREVIEW.format(
        theme=mon[0]["theme"],
        monday_date=mon[0]["session_date"],
        wednesday_date=wed[0]["session_date"],
        monday_speakers=monday_speakers,
        wednesday_speakers=wednesday_speakers,
        helmsman_name=helmsman_name,
        zoom_link=_zoom_url(),
    )

    if getattr(args, "dry_run", False):
        print(f"{CYAN}--- Sunday preview (DRY RUN, would post to chat_id={os.environ.get('FIRESIDE_TRIBE_CHAT_ID')}) ---{RESET}")
        print(text)
        print(f"{CYAN}--- end ---{RESET}")
        return

    bot = get_bot()
    chat_id = int(os.environ["FIRESIDE_TRIBE_CHAT_ID"])
    # The SEND and the PIN are separate outcomes. They used to share one `try`,
    # so a pin that failed after a successful post left `LAST_PINNED` unwritten
    # and reported the whole command as failed. A rerun then posted the preview
    # to the group a SECOND time, and `unpin-weekly` had no message id to unpin.
    # Record the post the moment it lands; treat the pin as a separate step that
    # can fail on its own.
    try:
        result = bot.send_message(chat_id, text)
    except TelegramAPIError as e:
        print(f"{RED}sunday-preview failed to post: {e}{RESET}", file=sys.stderr)
        return
    msg_id = result.get("message_id")
    save_state(LAST_PINNED, {"message_id": msg_id, "week": week_num,
                             "posted_at": local_now().isoformat()})
    _log_event("sunday_preview_posted", week=week_num, message_id=msg_id)
    pinned = True
    try:
        bot.pin_chat_message(chat_id, msg_id, disable_notification=True)
    except TelegramAPIError as e:
        pinned = False
        print(f"{YELLOW}sunday-preview: posted week {week_num} "
              f"(message_id={msg_id}) but the pin failed: {e}. Do NOT rerun -- "
              f"the message is already in the group; pin it by hand.{RESET}",
              file=sys.stderr)
    if pinned:
        print(f"{GREEN}sunday-preview{RESET}: posted week {week_num}, message_id={msg_id}, pinned")
    hc_ping("FIRESIDE_HC_SUNDAY_PREVIEW")


# ============================================================
# Subcommand: topic-nudge (weekly topic-collection invite)
# ============================================================

def cmd_topic_nudge(args) -> None:
    """Post the weekly 'topic box is open' invite to the Tribe group.

    Auto-send (same trust class as sunday-preview). Not pinned. Guards: silent
    no-op outside an active cycle, and silent no-op during the final week (the
    CEO-approved cycle-end invite owns that week). Cron: Saturday 12:00 local.
    """
    schedule = load_state(SCHEDULE) or []
    today = _today_local_date()
    if ft._upcoming_week(schedule, today) is None:
        print(f"{GRAY}topic-nudge: no active cycle; skip{RESET}")
        return
    if ft.is_final_week(schedule, today):
        print(f"{GRAY}topic-nudge: final week owned by cycle-end invite; skip{RESET}")
        return

    text = ft.render_nudge()
    if getattr(args, "dry_run", False):
        print(f"{CYAN}--- topic-nudge (DRY RUN, chat_id={os.environ.get('FIRESIDE_TRIBE_CHAT_ID')}) ---{RESET}")
        print(text)
        print(f"{CYAN}--- end ---{RESET}")
        return

    bot = get_bot()
    chat_id = int(os.environ["FIRESIDE_TRIBE_CHAT_ID"])
    try:
        result = bot.send_message(chat_id, text)
        _log_event("topic_nudge_posted", message_id=result.get("message_id"))
        print(f"{GREEN}topic-nudge{RESET}: posted, message_id={result.get('message_id')}")
    except TelegramAPIError as e:
        print(f"{RED}topic-nudge failed: {e}{RESET}", file=sys.stderr)


# ============================================================
# Subcommand: topic-digest (weekly CEO digest of new ideas)
# ============================================================

def cmd_topic_digest(args) -> None:
    """DM the CEO any topic ideas submitted since the last digest. Cron: Sun 09:00.

    Silent no-op when there are no new ideas. Advances the digest cursor only
    after a successful send so a failed DM is retried next run.
    """
    state = ft.load_topic_state(state_dir())
    cursor = state.get("last_digest_idea_id")
    new, new_cursor = ft.new_ideas_since(state_dir(), cursor)
    if not new:
        print(f"{GRAY}topic-digest: no new ideas since last digest{RESET}")
        return

    text = ft.render_digest(new)
    if getattr(args, "dry_run", False):
        print(f"{CYAN}--- topic-digest (DRY RUN, {len(new)} new) ---{RESET}")
        print(text)
        print(f"{CYAN}--- end ---{RESET}")
        return

    misha_id = misha_user_id()
    if not misha_id:
        print(f"{RED}topic-digest: MISHA_TELEGRAM_USER_ID not set{RESET}", file=sys.stderr)
        return

    bot = get_bot()
    try:
        bot.send_message(misha_id, text, parse_mode="")
    except TelegramAPIError as e:
        print(f"{RED}topic-digest DM failed: {e}{RESET}", file=sys.stderr)
        return
    state["last_digest_idea_id"] = new_cursor
    ft.save_topic_state(state_dir(), state)
    _log_event("topic_digest_sent", count=len(new), cursor=new_cursor)
    print(f"{GREEN}topic-digest{RESET}: DMed {len(new)} new idea(s) to CEO")


# ============================================================
# Subcommand: cycle-end-invite (CEO-approved end-of-cycle invite)
# ============================================================

def cmd_cycle_end_invite(args) -> None:
    """On the final-week Sunday, DRAFT the cycle-end topic invite to the CEO.

    Sends the CEO the warm invite + the full backlog summary with inline buttons
    [Send to Tribe] / [Cancel]. Posting to the group happens only on the CEO's
    tap (_handle_callback_query, namespace cycle_invite:*). Daily cron, but
    cycle_end_trigger_today() makes every non-trigger day a no-op. Idempotent:
    a pending draft for the current cycle is not re-drafted.
    """
    schedule = load_state(SCHEDULE) or []
    today = _today_local_date()
    if not ft.cycle_end_trigger_today(schedule, today):
        print(f"{GRAY}cycle-end-invite: not the final-week Sunday; skip{RESET}")
        return

    cycle = ft.current_cycle(schedule, today)
    state = ft.load_topic_state(state_dir())
    pending = state.get("pending_cycle_invite")
    if pending and pending.get("cycle") == cycle:
        print(f"{GRAY}cycle-end-invite: draft already pending for cycle {cycle}; skip{RESET}")
        return

    invite = ft.render_cycle_end_invite()
    backlog = ft.render_backlog_summary(ft.load_ideas(state_dir(), cycle=cycle))
    ceo_text = (
        "*Draft — cycle-end topic invite (your approval needed before it posts to the Tribe)*\n\n"
        "————— message preview —————\n"
        f"{invite}\n"
        "———————————————————\n\n"
        f"{backlog}"
    )
    markup = {"inline_keyboard": [[
        {"text": "✅ Send to Tribe", "callback_data": "cycle_invite:send"},
        {"text": "✖️ Cancel", "callback_data": "cycle_invite:cancel"},
    ]]}

    if getattr(args, "dry_run", False):
        print(f"{CYAN}--- cycle-end-invite (DRY RUN, draft to CEO) ---{RESET}")
        print(ceo_text)
        print(f"{CYAN}--- end ---{RESET}")
        return

    misha_id = misha_user_id()
    if not misha_id:
        print(f"{RED}cycle-end-invite: MISHA_TELEGRAM_USER_ID not set{RESET}", file=sys.stderr)
        return

    bot = get_bot()
    try:
        result = bot.send_message(misha_id, ceo_text, reply_markup=markup, parse_mode="")
    except TelegramAPIError as e:
        print(f"{RED}cycle-end-invite draft DM failed: {e}{RESET}", file=sys.stderr)
        return

    # Disarm the card this draft replaces. `pending` here is a draft for an
    # EARLIER cycle that the CEO never answered: the guard above only skipped on
    # a match, so an unanswered cycle-1 draft was silently discarded while its
    # message kept live buttons in the DM history. Best effort, and reported: if
    # the edit fails the stale card stays tappable, and the tap handler refuses
    # it on the message id, so nothing can be posted under the wrong text.
    if pending and pending.get("approval_msg_id"):
        _log_event("cycle_end_invite_superseded",
                   cycle=pending.get("cycle"), new_cycle=cycle,
                   approval_msg_id=pending.get("approval_msg_id"))
        try:
            bot.edit_message_text(
                misha_id, pending["approval_msg_id"],
                f"⏳ Superseded by the cycle {cycle} draft below. Nothing was sent.",
                parse_mode="", reply_markup=None)
        except TelegramAPIError as e:
            log_error(f"cycle-end-invite: could not retire the cycle "
                      f"{pending.get('cycle')} approval card: {e}")

    state["pending_cycle_invite"] = {
        "text": invite,                      # the exact text posted on approval
        "approval_msg_id": result.get("message_id"),
        "drafted_at": local_now().isoformat(),
        "cycle": cycle,
    }
    ft.save_topic_state(state_dir(), state)
    _log_event("cycle_end_invite_drafted", cycle=cycle,
               approval_msg_id=result.get("message_id"))
    print(f"{GREEN}cycle-end-invite{RESET}: drafted to CEO for approval (cycle {cycle})")


def cmd_cycle_rollover(args) -> None:
    """Rebuild schedule.json when a new cycle config lands and the old cycle ended.

    Daily cron. Reads the cycle config FRESH from disk (not the frozen import
    constants) and, if ft.cycle_rollover_needed() says the live cycle is over and
    the config describes a newer Week-1 Monday, backs up the outgoing schedule,
    rebuilds from the fresh config against the existing roster, saves, logs, and
    DMs the CEO a heads-up. Every non-rollover day is a no-op. Idempotent: once
    rebuilt the live Week-1 Monday equals the config's, so it stops firing.
    Never resets helmsmen or roster - those stay the CEO's call.
    """
    schedule = load_state(SCHEDULE) or []
    try:
        start_monday, weeks = _load_fireside_config_fresh()
    # TypeError too, matching the import-time guard over the same loader. A
    # config that is valid JSON but the wrong SHAPE - a top-level list, or
    # `weeks` as a dict - raises TypeError on `cfg["cycle_1_start_monday"]`,
    # which escaped this branch and surfaced through the main() wrapper as an
    # "uncaught exception" instead of the line naming the file to fix.
    except (OSError, ValueError, KeyError, TypeError) as e:
        print(f"{RED}cycle-rollover: cannot read cycle config: {e}{RESET}", file=sys.stderr)
        return

    today = _today_local_date()
    if not ft.cycle_rollover_needed(schedule, start_monday, today):
        print(f"{GRAY}cycle-rollover: no rollover due (cycle active or already built){RESET}")
        return

    # `or {}`, not `or []`: build_roster_by_name calls .items() on this, so an
    # absent roster used to raise AttributeError instead of rebuilding empty.
    roster_by_name = build_roster_by_name(load_state(TRIBE_ROSTER) or {})
    cycle = next_cycle_number(schedule)
    entries, missing = build_schedule(roster_by_name, start_monday=start_monday,
                                      weeks=weeks, cycle=cycle)
    if not entries:
        print(f"{RED}cycle-rollover: rebuilt schedule is empty; aborting{RESET}", file=sys.stderr)
        return
    dates = sorted({e["session_date"] for e in entries})
    unresolved = sorted(set(missing))
    # The inverse of `unresolved`: members the new cycle forgot entirely.
    unslotted = speaker_gaps(load_state(TRIBE_ROSTER) or {}, entries)

    if getattr(args, "dry_run", False):
        print(f"{CYAN}--- cycle-rollover (DRY RUN) ---{RESET}")
        print(f"would rebuild {len(entries)} entries, {dates[0]} -> {dates[-1]}")
        if unresolved:
            print(f"unresolved speakers: {unresolved}")
        if unslotted:
            print(f"members with no slot in the new cycle: {unslotted}")
        return

    # Back up the outgoing cycle before overwriting (one backup per rollover day).
    old = state_path(SCHEDULE)
    if old.exists():
        backup = state_path(f"schedule.pre-{today.isoformat()}.bak.json")
        if not backup.exists():
            backup.write_text(old.read_text(encoding="utf-8"), encoding="utf-8")

    save_state(SCHEDULE, entries)
    _log_event("cycle_rollover", week1=dates[0], last=dates[-1],
               entries=len(entries), unresolved=unresolved, unslotted=unslotted)
    print(f"{GREEN}cycle-rollover{RESET}: rebuilt {len(entries)} entries "
          f"({dates[0]} -> {dates[-1]}), unresolved={len(unresolved)}, "
          f"unslotted={len(unslotted)}")

    # Heads-up DM to the CEO (best-effort; never fail the job on a send error).
    misha_id = misha_user_id()
    if misha_id:
        note = ("Fireside cycle rolled over automatically.\n"
                f"New schedule: {dates[0]} -> {dates[-1]} ({len(entries)} slots).\n"
                "Helmsman for week 1 is not set yet - pick one when ready.")
        if unresolved:
            note += ("\nSpeakers with no Telegram match (roster refresh needed): "
                     + ", ".join(unresolved))
        if unslotted:
            note += ("\nIn the Tribe but in no week of the new cycle: "
                     + ", ".join(unslotted))
        try:
            get_bot().send_message(misha_id, note, parse_mode="")
        except TelegramAPIError as e:
            print(f"{YELLOW}cycle-rollover: heads-up DM failed: {e}{RESET}", file=sys.stderr)


# ============================================================
# Subcommand: topic-ideas (console-first backlog reader)
# ============================================================

def cmd_topic_ideas(args) -> None:
    """Print the topic backlog to the terminal. No Telegram dependency.

    --cycle N  : only that cycle.  --new : only ideas since the last digest.
    """
    cycle = getattr(args, "cycle", None)
    since = None
    if getattr(args, "new", False):
        since = ft.load_topic_state(state_dir()).get("last_digest_idea_id")
    ideas = ft.load_ideas(state_dir(), cycle=cycle, since_id=since)
    if not ideas:
        print(f"{GRAY}No topic ideas{' since last digest' if since else ''}.{RESET}")
        return
    print(f"{BOLD}{len(ideas)} topic idea(s):{RESET}")
    for n, i in enumerate(ideas, 1):
        who = i.get("name") or i.get("username") or "unknown"
        when = (i.get("ts") or "")[:10]
        print(f"  {n}. {i.get('text','').strip()}")
        print(f"       {GRAY}— {who}, {when}, cycle {i.get('cycle')}{RESET}")


# ============================================================
# Subcommand: dayof-reminders (Phase 3 task 3.4)
# ============================================================

def cmd_dayof_reminders(args) -> None:
    """DM today's 3 speakers their Zoom link. Cron: Mon + Wed 15:30 local (3h before 18:30)."""
    schedule = load_state(SCHEDULE) or []
    helmsmen = load_state(HELMSMEN) or {}
    roster = load_state(TRIBE_ROSTER) or {}
    today_iso = _today_local_date().isoformat()

    bot = get_bot()
    today_entries = [s for s in schedule if s["session_date"] == today_iso]
    if not today_entries:
        print(f"{GRAY}dayof-reminders: no sessions today ({today_iso}){RESET}")
        return

    week_num = today_entries[0]["week"]
    mon_entry = next((s for s in schedule if s["week"] == week_num and s["day"] == "Mon"), None)
    week_start = mon_entry["session_date"] if mon_entry else today_iso
    helmsman_name = (helmsmen.get(week_start, {})).get("name", "[Helmsman TBD]")
    zoom = _zoom_url()

    sent = 0
    failed = 0
    skipped = 0
    for entry in sorted(today_entries, key=lambda s: s["slot"]):
        username = entry.get("speaker_username")
        if not username:
            continue
        user_id = _resolve_speaker_user_id(roster, username)
        if user_id is None:
            _log_dm("dayof", username, today_iso, None, False,
                    error="no telegram_user_id")
            failed += 1
            continue
        # Guarded like the other two send loops. This was the only one of the
        # three with no `_dm_already_sent` check: it WROTE "dayof" rows and read
        # none of them, so a cron double-fire or a manual rerun on session day
        # sent every speaker the same Zoom link twice. That is the exact shape
        # `_dm_already_sent` exists for - the rows to prevent it were already
        # being written. Keyed on the session date, which for this job is today.
        if _dm_already_sent(state_path(DM_LOG), username, "dayof", today_iso):
            skipped += 1
            continue
        # Guarded like the other two send loops in this file: a blank
        # speaker_name raised IndexError mid-loop, and this is the day-of DM
        # carrying the Zoom link, so the speakers after the blank row lost the
        # only message that tells them where to be.
        name = _first_name(entry.get("speaker_name"), f"@{username}")
        text = SPEAKER_DM_DAYOF.format(name=name, zoom_link=zoom, helmsman_name=helmsman_name)
        try:
            bot.send_dm(user_id, text)
            _log_dm("dayof", username, today_iso, user_id, True)
            sent += 1
        except TelegramAPIError as e:
            _log_dm("dayof", username, today_iso, user_id, False, error=str(e))
            failed += 1
    line = f"{GREEN}dayof-reminders{RESET}: sent={sent} failed={failed}"
    if skipped:
        line += f" {GRAY}already-sent-today={skipped}{RESET}"
    print(line)
    hc_ping("FIRESIDE_HC_DAYOF_REMINDERS")


# ============================================================
# Helmsman coverage (pure helpers - no Telegram, no disk)
# ============================================================

def helmsman_gaps(schedule: list, helmsmen: dict,
                  on_or_after: Optional[Any] = None) -> list:
    """Monday keys in `schedule` that have no Helmsman assigned, sorted.

    Assignment has always been a manual CEO call - no code path writes a name
    into helmsmen.json - so a rolled-over cycle starts with every week empty and
    every renderer silently falls back to "TBD". This makes that gap queryable.

    on_or_after: if given, only weeks whose Monday is on or after this date.
    """
    from datetime import date as _date

    mondays = sorted({e["session_date"] for e in schedule if e.get("day") == "Mon"})
    out = []
    for m in mondays:
        if helmsmen.get(m, {}).get("name"):
            continue
        if on_or_after is not None and _date.fromisoformat(m) < on_or_after:
            continue
        out.append(m)
    return out


def speaker_gaps(roster: Optional[dict], schedule: list) -> list:
    """Active, non-excluded roster members who hold no slot in `schedule`, sorted.

    The inverse of the discrepancy report's "names in the schedule with no roster
    match". Membership is rebuilt from Telegram + xlsx by bootstrap, but the weeks
    a cycle runs are hand-authored in the cycle config, and nothing joined the two
    in this direction: a member who joins mid-cycle is in the roster, absent from
    every week, and reported nowhere. Returned as "Name (@username)" strings.

    An empty schedule makes everyone a gap - callers check the schedule first.

    A slot is credited to the member who HOLDS it, by username, whenever the row
    is bound. Matching on display name alone reopened the exact hole this
    function exists to close: two active members sharing a display name both read
    as covered while only one of them had a slot, so the other sat out the whole
    cycle and nothing anywhere said so. `build_roster_by_name` refuses to guess
    between such members and logs the ambiguity; this counted it as resolved.
    Only a row `build_schedule` could not bind (a speaker name with no roster
    match, `speaker_username` None) still falls back to the name, because a name
    is the only thing that row carries.
    """
    scheduled_handles = {str(e.get("speaker_username") or "").strip().lower()
                         for e in schedule if e.get("speaker_username")}
    unbound_names = {str(e.get("speaker_name") or "").strip()
                     for e in schedule if not e.get("speaker_username")}
    out = []
    for username, rec in (roster or {}).items():
        if not rec.get("active", True) or rec.get("excluded_from_fireside"):
            continue
        name = str(rec.get("name") or "").strip()
        if not name:
            continue
        if str(username).strip().lower() in scheduled_handles:
            continue
        if name in unbound_names:
            continue
        out.append(f"{name} (@{username})")
    return sorted(out)


def helmsman_brief_candidates(helmsmen: dict, today: Any,
                              horizon_days: int = 7) -> list:
    """Unbriefed Helmsmen whose week is still live and starts within the horizon.

    Returns [(monday_date, monday_key, entry)] sorted soonest-first.

    The week stays briefable through its Wednesday session (Monday + 2). The
    earlier `today < key_date` rule silently excluded any entry created on or
    after its own Monday, so a late assignment could never be briefed at all -
    which is exactly what happened to the 2026-07-13 week.
    """
    from datetime import date as _date, timedelta

    horizon = today + timedelta(days=horizon_days)
    candidates = []
    for key, entry in helmsmen.items():
        if entry.get("briefed"):
            continue
        try:
            key_date = _date.fromisoformat(key)
        except ValueError:
            continue
        week_ends = key_date + timedelta(days=2)  # Wednesday session
        if week_ends >= today and key_date <= horizon:
            candidates.append((key_date, key, entry))
    candidates.sort(key=lambda t: t[0])
    return candidates


def _nudge_ceo_on_helmsman_gaps(schedule: list, helmsmen: dict, today: Any,
                                lookahead_days: int = 14) -> None:
    """DM the CEO when a near-term week still has no Helmsman. Best-effort.

    Only weeks inside the lookahead window are nudged, so a freshly rolled-over
    cycle does not dump all ten weeks into one message.

    A week stays nudgeable through its Wednesday session (Monday + 2), matching
    `helmsman_brief_candidates` above. This passed `on_or_after=today`, which
    drops a week the moment its Monday is in the past - so the nudge went out on
    Monday and then fell silent on Tuesday and Wednesday, the last two days when
    the CEO could still have assigned someone. The Monday + 2 rule was written
    for exactly this after the 2026-07-13 week, and it landed in one of the two
    functions that needed it. The Wednesday session then ran with "[Helmsman
    TBD]" in every speaker DM and in the pinned preview.
    """
    from datetime import date as _date, timedelta

    gaps = helmsman_gaps(schedule, helmsmen, on_or_after=today - timedelta(days=2))
    near = [g for g in gaps if _date.fromisoformat(g) <= today + timedelta(days=lookahead_days)]
    if not near:
        return

    misha_id = misha_user_id()
    if not misha_id:
        print(f"{YELLOW}helmsman gap: {', '.join(near)} (no MISHA_TELEGRAM_USER_ID to nudge){RESET}")
        return

    weeks = "\n".join(f"  - week starting {g}" for g in near)
    remaining = len(gaps) - len(near)
    text = ("Fireside: no Helmsman assigned for:\n" + weeks
            + (f"\n(+{remaining} later week(s) also unassigned)" if remaining else "")
            + "\n\nSpeakers and the pinned preview will read 'TBD' until one is set."
              "\nAssign: python scripts/fireside-bot.py helmsman set --week <YYYY-MM-DD> --username <handle>")
    # Strictly best-effort: this runs BEFORE the actual brief, so a nudge that
    # blows up must never cost the Helmsman their briefing. Logged, not swallowed.
    try:
        get_bot().send_message(misha_id, text, parse_mode="")
        _log_event("helmsman_gap_nudge", weeks=near)
    except Exception as e:
        log_error("helmsman gap nudge failed", e)
        print(f"{YELLOW}helmsman gap nudge failed: {e}{RESET}", file=sys.stderr)


# ============================================================
# Subcommand: helmsman-brief (Phase 3 task 3.5)
# ============================================================

def cmd_helmsman_brief(args) -> None:
    """Brief the closest unbrief'd Helmsman whose week starts within 7 days. Cron: daily 10:00 local time.

    The earlier tight `today + 7 days == key` rule meant any missed run (bot down, scheduler
    failure) silently skipped the brief forever. Window-based + idempotent via `briefed` flag
    means a missed day catches up on the next run.

    ONE PER RUN, and "catches up" means one per DAY, not all at once. That is a
    deliberate rate limit on DMs to real people, not an oversight -- but it does
    mean a backlog of N pending Helmsmen takes N days to clear, and one whose
    week starts sooner than that is never briefed in time. The count is printed
    when more than one is waiting, so the backlog is visible rather than
    implied; clearing it faster is the operator's call, since it is more DMs in
    one run than this job has ever sent.
    """
    schedule = load_state(SCHEDULE) or []
    # `helmsmen` is loaded and saved much further down, with a Telegram DM in
    # between. It is NOT held under a lock for that whole span on purpose: the
    # send can take seconds and a lock across a network call is how a bot
    # deadlocks itself. The write is instead re-read under the lock at the end,
    # so a concurrent `helmsman set` is merged rather than overwritten.
    helmsmen = load_state(HELMSMEN) or {}
    opt_ins = load_state(OPT_INS) or {"helmsman": [], "wildcard": []}
    roster = load_state(TRIBE_ROSTER) or {}
    today = _today_local_date()

    # An unassigned week is silent otherwise: this job reports "nothing to brief"
    # and pings the healthcheck green, so a gap can run for weeks unnoticed.
    _nudge_ceo_on_helmsman_gaps(schedule, helmsmen, today)

    candidates = helmsman_brief_candidates(helmsmen, today)
    if not candidates:
        print(f"{GRAY}helmsman-brief: no pending Helmsman briefs within 7 days{RESET}")
        hc_ping("FIRESIDE_HC_HELMSMAN_BRIEF")
        return

    if len(candidates) > 1:
        later = ", ".join(str(c[1]) for c in candidates[1:])
        print(f"{YELLOW}helmsman-brief: {len(candidates)} Helmsmen pending; this run "
              f"briefs the closest only. Still waiting: {later}{RESET}",
              file=sys.stderr)
    target_date, target_week_start, helmsman_entry = candidates[0]

    user_id = helmsman_entry.get("user_id")
    if not user_id:
        # Resolve via roster username
        username = helmsman_entry.get("username")
        user_id = _resolve_speaker_user_id(roster, username) if username else None

    if not user_id:
        print(f"{RED}helmsman-brief: no user_id for Helmsman {helmsman_entry}{RESET}", file=sys.stderr)
        return

    week_num = next((s["week"] for s in schedule if s["session_date"] == target_week_start), None)
    if week_num is None:
        print(f"{RED}helmsman-brief: no schedule for week starting {target_week_start}{RESET}", file=sys.stderr)
        return

    mon = _week_speakers(schedule, week_num, "Mon")
    wed = _week_speakers(schedule, week_num, "Wed")
    # Guarded, the way `sunday-preview` already guards. `mon[0]` and `wed[0]`
    # below assume both days exist; a week holding only one of them raised
    # IndexError BEFORE the entry was marked `briefed`, so the daily job hit the
    # same candidate again the next day, and the next -- a permanent failure
    # loop over one malformed week, and no Helmsman ever briefed.
    if not mon or not wed:
        print(f"{RED}helmsman-brief: week {week_num} starting {target_week_start} "
              f"has {len(mon)} Mon and {len(wed)} Wed session(s); both are needed. "
              f"Fix the schedule -- this candidate is skipped, not marked briefed."
              f"{RESET}", file=sys.stderr)
        return
    monday_speakers = ", ".join(s["speaker_name"] for s in mon)
    wednesday_speakers = ", ".join(s["speaker_name"] for s in wed)
    wildcard_lines = "\n".join(
        f"  - @{w['username']}" for w in opt_ins.get("wildcard", [])
    ) or "  (none yet - opt-ins are still open in the 31C Tribe group)"

    # Guarded like the three send loops, through the one helper they now share.
    # A hand-edited helmsmen.json with an empty or null `name` used to raise out
    # of this daily job before the entry was stamped `briefed`.
    handle = helmsman_entry.get("username")
    name = _first_name(helmsman_entry.get("name"),
                       f"@{handle}" if handle else "Helmsman")
    text = HELMSMAN_BRIEF.format(
        name=name,
        week_starting=target_week_start,
        monday_date=mon[0]["session_date"],
        wednesday_date=wed[0]["session_date"],
        monday_speakers=monday_speakers,
        wednesday_speakers=wednesday_speakers,
        theme=mon[0]["theme"],
        wildcard_list=wildcard_lines,
    )

    bot = get_bot()
    try:
        bot.send_dm(user_id, text)
        # Re-read under the lock and stamp only THIS entry. Saving the dict
        # loaded before the DM overwrote whatever `helmsman set` wrote while the
        # send was in flight -- an assignment the operator had just made,
        # silently gone.
        with locked_state(HELMSMEN, {}) as fresh:
            row = fresh.get(target_week_start)
            if isinstance(row, dict):
                row["briefed"] = True
                row["briefed_at"] = local_now().isoformat()
            else:
                helmsman_entry["briefed"] = True
                helmsman_entry["briefed_at"] = local_now().isoformat()
                fresh[target_week_start] = helmsman_entry
        _log_event("helmsman_briefed", week=week_num, user_id=user_id)
        print(f"{GREEN}helmsman-brief{RESET}: sent to {name} (user_id={user_id}) for week {week_num}")
        hc_ping("FIRESIDE_HC_HELMSMAN_BRIEF")
    except TelegramAPIError as e:
        print(f"{RED}helmsman-brief failed: {e}{RESET}", file=sys.stderr)


# ============================================================
# Subcommand: helmsman (console-first assignment: set | list | gaps)
# ============================================================

def cmd_helmsman(args) -> Optional[int]:
    """Assign, list, and audit Helmsman coverage from the terminal.

    Assignment was previously a hand-edit of helmsmen.json over SSH, which is
    how cycle 2 shipped with nine empty weeks. `gaps` exits 1 when any week in
    the live schedule is unassigned, so it works as a check in a pipeline.
    """
    from datetime import date as _date

    schedule = load_state(SCHEDULE) or []
    helmsmen = load_state(HELMSMEN) or {}
    action = args.action

    if action == "gaps":
        gaps = helmsman_gaps(schedule, helmsmen)
        if not gaps:
            print(f"{GREEN}helmsman gaps{RESET}: none - every week in the schedule is assigned")
            return 0
        print(f"{YELLOW}helmsman gaps{RESET}: {len(gaps)} week(s) unassigned")
        for g in gaps:
            print(f"  {g}")
        return 1

    if action == "list":
        if not helmsmen:
            print(f"{GRAY}No Helmsmen assigned.{RESET}")
            return 0
        mondays = sorted({e["session_date"] for e in schedule if e.get("day") == "Mon"})
        for m in mondays:
            entry = helmsmen.get(m, {})
            name = entry.get("name") or f"{YELLOW}-- unassigned --{RESET}"
            flag = "briefed" if entry.get("briefed") else "not briefed"
            print(f"  {m}  {name}" + (f"  ({flag})" if entry.get("name") else ""))
        return 0

    # action == "set"
    week = args.week
    try:
        week_date = _date.fromisoformat(week)
    except ValueError:
        print(f"{RED}helmsman set: --week must be YYYY-MM-DD{RESET}", file=sys.stderr)
        return 1
    if week_date.weekday() != 0:
        print(f"{RED}helmsman set: {week} is not a Monday{RESET}", file=sys.stderr)
        return 1
    known = {e["session_date"] for e in schedule if e.get("day") == "Mon"}
    if week not in known:
        print(f"{RED}helmsman set: {week} is not a week in the current schedule{RESET}",
              file=sys.stderr)
        return 1

    roster = load_state(TRIBE_ROSTER) or {}
    username = args.username.lstrip("@")
    entry = _roster_entry(roster, username)
    if entry is None:
        print(f"{RED}helmsman set: @{username} is not in the roster{RESET}", file=sys.stderr)
        return 1
    if not entry.get("active", True) or entry.get("excluded_from_fireside"):
        print(f"{RED}helmsman set: @{username} is inactive or excluded from fireside{RESET}",
              file=sys.stderr)
        return 1
    user_id = entry.get("telegram_user_id")
    if not user_id:
        print(f"{RED}helmsman set: @{username} has no telegram_user_id (run bootstrap){RESET}",
              file=sys.stderr)
        return 1

    # Re-read under the lock rather than saving the dict loaded at the top of
    # this command: `helmsman-brief` runs daily and rewrites the same file, and
    # whichever of the two saved second used to erase the other's change.
    with locked_state(HELMSMEN, {}) as fresh:
        previous = fresh.get(week)
        record = {
            "name": entry.get("name"),
            "username": username,
            "user_id": user_id,
            "briefed": False,
            "assigned_at": _today_local_date().isoformat(),
        }
        if args.note:
            record["note"] = args.note
        if previous:
            record["previous_helmsman_for_week"] = previous
        fresh[week] = record
    _log_event("helmsman_assigned", week=week, user_id=user_id)
    verb = "reassigned" if previous else "assigned"
    print(f"{GREEN}helmsman set{RESET}: {week} {verb} to {record['name']} (@{username})")
    return 0


# ============================================================
# Subcommand: speaker-gaps (console-first coverage check)
# ============================================================

def cmd_speaker_gaps(args) -> Optional[int]:
    """List roster members with no slot in the live schedule. Exits 1 if any.

    Same shape as `helmsman gaps`, so it works as a check in a pipeline before
    a new cycle config is declared finished.
    """
    schedule = load_state(SCHEDULE) or []
    if not schedule:
        print(f"{YELLOW}speaker-gaps{RESET}: schedule.json is empty - nothing to check "
              f"(run bootstrap or cycle-rollover first)")
        return 1

    gaps = speaker_gaps(load_state(TRIBE_ROSTER) or {}, schedule)
    if not gaps:
        print(f"{GREEN}speaker-gaps{RESET}: none - every active member holds a slot")
        return 0
    print(f"{YELLOW}speaker-gaps{RESET}: {len(gaps)} active member(s) with no slot this cycle")
    for g in gaps:
        print(f"  {g}")
    return 1


# ============================================================
# Subcommand: weekly-discrepancy-report (Phase 3 task 3.6)
# ============================================================

def cmd_weekly_discrepancy_report(args) -> None:
    """Re-run cross-reference; DM Misha if discrepancies found. Cron: Sunday 17:00 local time."""
    import asyncio
    try:
        bootstrap_result = asyncio.run(_bootstrap_async())
    except Exception as e:
        print(f"{RED}weekly-discrepancy-report: Telethon failed: {e}{RESET}", file=sys.stderr)
        log_error("weekly-discrepancy-report Telethon failed", e)
        return

    try:
        xlsx_roster = load_tribe_metadata()
    except (FileNotFoundError, *_UNREADABLE_SHEET) as e:
        print(f"{RED}weekly-discrepancy-report: xlsx load failed: {e}{RESET}", file=sys.stderr)
        return

    roster, discrepancy = cross_reference(xlsx_roster, bootstrap_result["telegram_members"])
    in_tg = discrepancy["in_telegram_not_in_xlsx"]
    in_xlsx = discrepancy["in_xlsx_not_in_telegram"]
    no_un = discrepancy["no_username_in_telegram"]
    # Members who are in the Tribe but in no week of the running cycle. Reported
    # here rather than gating the job: an unslotted member is a CEO to-do for the
    # next cycle config, not a daemon failure.
    schedule = load_state(SCHEDULE) or []
    unslotted = speaker_gaps(roster, schedule) if schedule else []

    if not in_tg and not in_xlsx and not no_un and not unslotted:
        print(f"{GREEN}weekly-discrepancy-report: no discrepancies{RESET}")
        return

    lines = ["**Weekly Tribe roster discrepancy report**", ""]
    if in_tg:
        lines.append(f"In Telegram, missing from xlsx ({len(in_tg)}):")
        for r in in_tg:
            lines.append(f"  - @{r['username']}: {r['full_name']}")
        lines.append("")
    if in_xlsx:
        lines.append(f"In xlsx, missing from Telegram ({len(in_xlsx)}):")
        for r in in_xlsx:
            lines.append(f"  - @{r['username']}: {r['name']}")
        lines.append("")
    if no_un:
        lines.append(f"In Telegram, no username set ({len(no_un)}):")
        for r in no_un:
            lines.append(f"  - {r['full_name']} (id={r['user_id']})")
        lines.append("")
    if unslotted:
        lines.append(f"In the Tribe, no slot in this cycle ({len(unslotted)}) "
                     f"- add to the next cycle config:")
        for g in unslotted:
            lines.append(f"  - {g}")
    text = "\n".join(lines).rstrip()

    misha_id = misha_user_id()
    if not misha_id:
        print(text)
        return
    bot = get_bot()
    try:
        # parse_mode="", like every other send in this file. Legacy Markdown
        # here interpolated raw @handles, full names and roster data, so a
        # single `_` in a username made Telegram reject the message with
        # "can't parse entities" -- the discrepancy report failed precisely when
        # the data was messy, which is the only time it has anything to say.
        bot.send_message(misha_id, text, parse_mode="")
        print(f"{GREEN}weekly-discrepancy-report{RESET}: DM sent to Misha")
    except TelegramAPIError as e:
        print(f"{RED}weekly-discrepancy-report DM failed: {e}{RESET}", file=sys.stderr)


# ============================================================
# Subcommand: email-backup (Phase 3 task 3.7)
# ============================================================

def cmd_email_backup(args) -> None:
    """Email speakers who haven't responded to bot DMs. Cron: Sunday 19:00 local time."""
    import subprocess
    from datetime import date as _date, timedelta

    schedule = load_state(SCHEDULE) or []
    roster = load_state(TRIBE_ROSTER) or {}
    today = _today_local_date()

    # Engagement lives in sessions.jsonl, NOT dm-log.jsonl. `_log_event` writes
    # `start_received` / `swap_requested` with an `event_type` key to
    # SESSIONS_LOG; dm-log rows carry `dm_type` and have no `event_type` at all.
    # Scanning the wrong file meant `responded_user_ids` was ALWAYS empty, so a
    # member who had answered the bot was never recognised as responsive and
    # stayed on the backup-email list.
    #
    # All four engagement events, not two. `idea_submitted` and
    # `opt_in_changed` are written by `_log_event` with a `user_id` and each is
    # unambiguous proof the member is alive and using the bot, and both were
    # missing from this set. A member pre-bound by `bootstrap` (so never needing
    # `/start`) who sent an `/idea` last week was still classed unresponsive and
    # mailed "I've sent you a few Telegram DMs... but haven't seen a response" -
    # false, and false to exactly the members who are engaged.
    ENGAGEMENT_EVENTS = ("start_received", "swap_requested", "idea_submitted",
                         "opt_in_changed")
    responded_user_ids: set[int] = set()
    for e in _read_jsonl_rows(state_path(SESSIONS_LOG)):
        if e.get("event_type") in ENGAGEMENT_EVENTS and e.get("user_id"):
            try:
                responded_user_ids.add(int(e["user_id"]))
            except (TypeError, ValueError):
                continue

    sent = 0
    skipped = 0
    already = 0
    # Two silent drops used to leave no trace at all. The summary line reported
    # `sent` and `skipped`, where `skipped` counts ONLY the members who had
    # already engaged with the bot -- so a speaker with no e-mail in the sheet,
    # or one whose schedule row names a handle the roster does not carry, fell
    # out of the loop and the line read as though the whole window was handled.
    # They are exactly the people this command exists for. Counted and named now.
    no_email: list[str] = []
    not_in_roster: list[str] = []
    failed: list[str] = []
    for entry in schedule:
        username = entry.get("speaker_username")
        if not username:
            continue
        d = _date.fromisoformat(entry["session_date"])
        days_until = (d - today).days
        if not (1 <= days_until <= 14):
            continue  # only current 2-week window
        roster_entry = _roster_entry(roster, username)
        if not roster_entry:
            not_in_roster.append(str(username))
            continue
        user_id = roster_entry.get("telegram_user_id")
        email = roster_entry.get("email")
        if not email:
            no_email.append(str(username))
            continue
        if user_id and user_id in responded_user_ids:
            skipped += 1
            continue  # they've engaged via bot, no need for email
        # Already mailed TODAY, not already mailed at all: the two-Sundays
        # behaviour below is deliberate. A rerun on the same Sunday, or a cron
        # double-fire, used to send real people a second identical email.
        if _dm_already_sent(state_path(DM_LOG), username, "email-backup",
                            entry["session_date"], on_date=today.isoformat()):
            already += 1
            continue

        # Silence is the trigger, not a failed DM. Until 2026-08-24 this line
        # also required a dm-log row with `dm_type in ("2wk", "3day")` and
        # `not delivered`, so only a member whose DM had FAILED got mail. A
        # member whose DM arrived and who then said nothing -- the exact
        # "unresponsive Tribe member" the command is named for -- got none, and
        # neither did a member with no Telegram account at all. Operator ruled
        # on it: "шли всегда". Everyone in the 1..14 day window who has an email
        # and has not engaged with the bot is mailed now.
        #
        # This mails on both Sundays inside the window, mirroring the two DM
        # nudges (2wk and 3day). It is two emails per session, not one.
        # Guarded: `roster_entry["name"].split()[0]` raised KeyError on a record
        # with no Name and IndexError on a blank one, out of the middle of a send
        # loop. Earlier addressees had their mail, later ones never would, and the
        # summary line below was never reached -- a partial send that reported
        # nothing at all. A handle is a worse greeting than a first name and a
        # better one than a dead job.
        name = _first_name(roster_entry.get("name"), f"@{username}")
        subject = EMAIL_BACKUP_SUBJECT.format(session_date=entry["session_date"])
        # The "reach a human" line used to be a tenant mailbox literal, in a
        # template that goes OUT as email. On any other deployment that invites
        # a stranger's Tribe to write to this operator. Resolved from the
        # operator seam instead, and OMITTED ENTIRELY when unconfigured: an
        # empty address in an outbound message is worse than no offer of help,
        # because the reader tries it.
        #
        # `admin_email` first, because the human who fields "my fireside slot
        # is wrong" is whoever administers the fleet, not necessarily whoever
        # the daemon runs as. `email` is the fallback rather than the primary
        # for the same reason. On a single-operator workspace they are the same
        # person and the order never shows; on an exec workspace it does.
        _human = admin_email() or (get_operator().get("email") or "").strip()
        body_text = EMAIL_BACKUP_BODY.format(
            name=name,
            session_date=entry["session_date"],
            session_day=d.strftime("%A"),
            theme=entry["theme"],
            human_contact=(f"\n(via {_human} if you need to reach a human)"
                           if _human else ""),
        )
        body_html = "<p>" + body_text.replace("\n\n", "</p><p>").replace("\n", "<br/>") + "</p>"

        # sys.executable, never a bare "python": the service host has no
        # `python` on PATH (only python3 + the venv), so every backup email
        # raised FileNotFoundError and this job reported sent=0 while its
        # healthcheck stayed green. sys.executable is the interpreter already
        # running the daemon, so it carries the pinned dependencies.
        cmd = [
            sys.executable, "scripts/send-email.py",
            "--to", email,
            "--subject", subject,
            "--body", body_html,
        ]
        err = None
        try:
            r = subprocess.run(cmd, capture_output=True, text=True, timeout=60,
                               cwd=str(WORKSPACE_ROOT))
            ok = (r.returncode == 0)
            if not ok:
                err = (r.stderr or "")[:200]
                # A non-zero exit reached dm-log.jsonl and nothing else. The
                # exception branch below has always called `log_error`; the
                # ordinary failure - an expired Exchange password, a rejected
                # recipient - did not, so the one class that happens routinely
                # was the one that left no line in errors.log.
                log_error(f"email-backup send failed for {email} "
                          f"(exit {r.returncode}): {err}")
        except Exception as e:
            log_error(f"email-backup subprocess failed for {email}", e)
            ok = False
            err = str(e)[:200]
        _log_dm("email-backup", username, entry["session_date"], user_id, ok, error=err)
        if ok:
            sent += 1
        else:
            failed.append(str(username))
    # `failed` too. The comment above says every drop class is counted and named
    # now, and it enumerated the two that skip the send. A send that RAN and
    # failed incremented nothing, so eight bounced emails printed
    # "sent=0 skipped=0" - byte-identical to the healthy line for a Sunday when
    # nobody was due. Both sibling send loops in this file print `failed=`.
    line = f"{GREEN}email-backup{RESET}: sent={sent} skipped={skipped}"
    if failed:
        line += (f" {RED}failed={len(failed)}{RESET} ("
                 + ", ".join("@" + u for u in sorted(set(failed))) + ")")
    if already:
        line += f" {GRAY}already-mailed-today={already}{RESET}"
    if no_email:
        line += (f" {YELLOW}no-email={len(no_email)}{RESET} ("
                 + ", ".join("@" + u for u in sorted(set(no_email))) + ")")
    if not_in_roster:
        line += (f" {YELLOW}not-in-roster={len(not_in_roster)}{RESET} ("
                 + ", ".join("@" + u for u in sorted(set(not_in_roster))) + ")")
    print(line)


# ============================================================
# Subcommand: stats (Phase 3 task 3.8)
# ============================================================

def cmd_stats(args) -> None:
    """Generate markdown stats report from dm-log + sessions logs. On-demand."""
    schedule = load_state(SCHEDULE) or []
    roster = load_state(TRIBE_ROSTER) or {}
    opt_ins = load_state(OPT_INS) or {"helmsman": [], "wildcard": []}
    helmsmen = load_state(HELMSMEN) or {}
    today = _today_local_date()

    spoken_users: set[str] = set()
    no_show_count: dict[str, int] = {}
    swap_count = 0

    # `_read_jsonl_rows`, not another copy of the walk. `except (json.JSONDecodeError, ValueError)`
    # here LOOKED like it covered the decode -- `UnicodeDecodeError` is a
    # `ValueError` -- but the decode happens in `for line in f`, one frame
    # outside the try, so it never reached this clause and the report died on a
    # torn append instead of printing.
    for e in _read_jsonl_rows(state_path(SESSIONS_LOG)):
        t = e.get("event_type")
        if t == "session_logged":
            for u in (e.get("shared") or "").split(","):
                if u.strip():
                    spoken_users.add(u.strip())
            for u in (e.get("no_shows") or "").split(","):
                if u.strip():
                    no_show_count[u.strip()] = no_show_count.get(u.strip(), 0) + 1
        elif t == "swap_requested":
            swap_count += 1

    delivered = total = 0
    for e in _read_jsonl_rows(state_path(DM_LOG)):
        # No "helmsman_brief" here: nothing writes a dm-log row with
        # that type. `cmd_helmsman_brief` records its send through
        # `_log_event("helmsman_briefed", ...)`, so the filter named
        # five categories and counted four, and the percentage beneath
        # spoke for a set it never contained. Either the writer or the
        # reader had to change; the brief is already covered by the
        # "briefed" flag rendered under "Helmsman schedule" below, so
        # the reader is the one that was lying.
        if e.get("dm_type") in ("2wk", "3day", "dayof", "email-backup"):
            total += 1
            if e.get("delivered"):
                delivered += 1

    completed = sum(1 for s in schedule if s.get("completed"))
    sessions_total = len(set((s["session_date"]) for s in schedule))
    completed_sessions = len(set(s["session_date"] for s in schedule if s.get("completed")))
    # `or 1` turned "no current or upcoming session" into "week 1 of 9", so a
    # finished cycle reported itself as just starting. None means the cycle is
    # over (or has not begun); the line below says that instead of inventing a
    # week the schedule does not contain.
    #
    # The denominator was the literal 9 while the numerator is measured from
    # the schedule. Cycle length is DATA - the `weeks` array of
    # fireside-schedule.json, which `cmd_cycle_rollover` re-reads fresh and
    # rebuilds from - so a cycle configured with ten weeks reported "week 10 of
    # 9" and one with eight reported "week 8 of 9" as though a week were still
    # to come. Read it from the same schedule the numerator came from.
    current_week = _current_or_upcoming_week(schedule, today)
    total_weeks = max((s.get("week") or 0) for s in schedule) if schedule else 0
    current_week_label = (f"**{current_week}** of {total_weeks or '?'}"
                          if current_week
                          else "**no active week** (cycle complete or not started)")

    # Tribe rotation health
    all_speakers = sorted({s["speaker_name"] for s in schedule})
    spoken_names = sorted(spoken_users)
    unspoken = [n for n in all_speakers if n not in spoken_users]

    lines = [
        f"# Tribe Fireside Stats — {today.isoformat()}",
        "",
        f"## Cycle progress",
        f"- Current week: {current_week_label}",
        f"- Sessions completed: **{completed_sessions}** of {sessions_total}",
        f"- Speaker entries completed: **{completed}** of {len(schedule)}",
        "",
        f"## Speaker rotation",
        f"- Spoken so far ({len(spoken_names)}): {', '.join(spoken_names) or '(none yet)'}",
        f"- Not spoken yet ({len(unspoken)}): {', '.join(unspoken[:10])}{'...' if len(unspoken) > 10 else ''}",
        "",
        f"## No-show counts",
    ]
    if no_show_count:
        for name, count in sorted(no_show_count.items(), key=lambda x: -x[1]):
            lines.append(f"- {name}: {count}")
    else:
        lines.append("- No no-shows recorded")
    lines.extend([
        "",
        f"## Swap requests",
        f"- Total: **{swap_count}**",
        "",
        f"## Opt-in rosters",
        f"- 🧭 Helmsmen: **{len(opt_ins['helmsman'])}** opted in",
        f"- 🌟 Wildcards: **{len(opt_ins['wildcard'])}** opted in",
        "",
        f"## DM delivery",
        f"- Delivered: {delivered} / {total} ({(100*delivered/total) if total else 0:.0f}%)",
        "",
        f"## Roster health",
        f"- Active members in roster: **{sum(1 for r in roster.values() if r.get('active', True))}**",
        f"- Members with telegram_user_id (have /started bot or pre-populated): "
        f"**{sum(1 for r in roster.values() if r.get('telegram_user_id'))}**",
        f"- Members without telegram_user_id (won't receive DMs until they /start): "
        f"**{sum(1 for r in roster.values() if not r.get('telegram_user_id'))}**",
        "",
        f"## Helmsman schedule",
    ])
    if helmsmen:
        for week_start, h in sorted(helmsmen.items()):
            briefed = "✓ briefed" if h.get("briefed") else "pending brief"
            lines.append(f"- {week_start}: {h.get('name', '?')} ({briefed})")
    else:
        lines.append("- No Helmsmen assigned yet")

    report = "\n".join(lines) + "\n"
    out_path = require_writable_stats_dir() / f"{today.isoformat()}_stats.md"
    out_path.write_text(report, encoding="utf-8")

    print(f"{GREEN}stats{RESET}: written to {out_path}")
    if getattr(args, "show", False):
        print()
        print(report)


# ============================================================
# Subcommand: health-check (Phase 3 task 3.9)
# ============================================================

def cmd_health_check(args) -> None:
    """Alert Misha if no liveness tick (poll-tick or heartbeat-tick) in 30 min.

    A tick means "a process reached Telegram", nothing more. In webhook mode
    that is a DIFFERENT process from the one handling updates, so this command
    reports the webhook's pending count and last error alongside the tick age
    rather than implying it has checked the handler. Cron: every 30 min.
    """
    from datetime import timedelta

    # An absent dm-log.jsonl is the STRONGEST case of "no liveness tick": the
    # daemon has not written a single one since the file was last there. This
    # printed a line to stdout that only a person watching cron output would
    # ever see, and returned before the alert path - while the strictly weaker
    # case, a file that exists and holds no tick, alerted. The monitor went
    # quiet exactly where the evidence was worst. Fall through with no ticks
    # instead, and let the `last_tick_ts is None` branch below say so.
    #
    # The read goes through `_read_jsonl_rows`, which decodes a line at a time.
    # `read_text(...).splitlines()` decoded the WHOLE file with no handler at
    # all, and the `except (json.JSONDecodeError, ValueError)` two lines further down could not
    # cover it because the decode had already happened outside the try. So this
    # monitor - the one job whose purpose is to notice that the bot stopped
    # writing ticks - died on a torn append to the very file it reads, and
    # `dm-log.jsonl` is append-only with nothing pruning it, so it stayed dead.
    dm_log_path = state_path(DM_LOG)
    if not dm_log_path.exists():
        print(f"{YELLOW}health-check: dm-log.jsonl missing{RESET}")

    last_tick_ts = None
    last_tick: dict = {}
    for e in _read_jsonl_rows(dm_log_path):
        # Either tick type counts as proof of life, and since 2026-08-25 the two
        # certify the same thing: poll-tick is written after a successful
        # getUpdates, heartbeat-tick after a successful getWebhookInfo. Both mean
        # "the process that stamped this reached Telegram". heartbeat-tick used
        # to mean only "cron fired and a file append worked", so this comment's
        # "either is enough" contradicted poll-tick's own definition.
        if e.get("dm_type") in ("poll-tick", "heartbeat-tick"):
            last_tick_ts = e.get("ts")
            last_tick = e

    now = local_now()
    last_dt = None
    if last_tick_ts is not None:
        # Guarded. An unparseable or naive `ts` raised ValueError/TypeError out
        # of the one command whose job is to NOTICE that something is wrong --
        # health-check crashed instead of alerting, which reads as a silent
        # daemon either way.
        try:
            last_dt = datetime.fromisoformat(last_tick_ts)
            if last_dt.tzinfo is None:
                last_dt = last_dt.replace(tzinfo=now.tzinfo)
        except (TypeError, ValueError):
            last_dt = None

    if last_tick_ts is None:
        msg = "⚠️ Fireside Bot health-check: no liveness tick ever recorded. Bot may not be running."
    elif last_dt is None:
        msg = (f"⚠️ Fireside Bot health-check: the newest liveness tick has an "
               f"unreadable timestamp ({last_tick_ts!r}). Liveness is UNKNOWN, "
               f"which is not the same as healthy.")
    else:
        age = now - last_dt
        if age > timedelta(minutes=30):
            mins = int(age.total_seconds() // 60)
            msg = (f"⚠️ Fireside Bot health-check: last liveness tick was {mins} min ago "
                   f"(threshold 30 min). Check daemon: "
                   f"`systemctl status <fireside-unit>` on the service host.")
        else:
            # What a fresh tick establishes, and no more: a process reached
            # Telegram. In webhook mode it does NOT establish that the webhook
            # HANDLER is consuming updates - that is a separate process, and the
            # heartbeat cron can be perfectly healthy above a dead one. The two
            # fields below are the evidence available from here, so they are
            # printed rather than hidden. They raise no alert: the pending count
            # at which "the handler is dead" beats "a session is busy" is a
            # threshold the operator owns.
            print(f"{GREEN}health-check{RESET}: last tick "
                  f"{int(age.total_seconds())}s ago, reached Telegram")
            print(f"  webhook: pending_update_count="
                  f"{last_tick.get('pending_update_count', 'n/a')} "
                  f"last_error={last_tick.get('webhook_last_error') or 'none'}")
            print(f"{GRAY}  a fresh tick does not prove the webhook handler is "
                  f"alive; a climbing pending count is the signal to check it"
                  f"{RESET}")
            return

    misha_id = misha_user_id()
    if not misha_id:
        print(msg)
        return
    bot = get_bot()
    try:
        bot.send_message(misha_id, msg)
        print(f"{YELLOW}health-check{RESET}: alert DM sent to Misha")
    except TelegramAPIError as e:
        print(f"{RED}health-check alert failed: {e}{RESET}", file=sys.stderr)


# ============================================================
# Subcommand: unpin-weekly (Phase 3 task 3.10)
# ============================================================

def cmd_unpin_weekly(args) -> None:
    """Unpin the Sunday preview message. Runs AFTER the Wednesday session.

    The docstring used to add "Cron: Wed 16:00 local (after Wed session)", and
    the module header repeated it. Sessions are at 18:30, so 16:00 is two and a
    half hours BEFORE the Wednesday session, not after it - and the pinned
    message is the one carrying that day's Zoom link. The two halves of that
    sentence cannot both be true. The real cron lives outside this file, so this
    docstring no longer states a time it cannot verify; it states the intent,
    and the operator's schedule is the thing to check against it.
    """
    last = load_state(LAST_PINNED) or {}
    msg_id = last.get("message_id")
    if not msg_id:
        print(f"{GRAY}unpin-weekly: no pinned message recorded; nothing to unpin{RESET}")
        return
    bot = get_bot()
    chat_id = int(os.environ["FIRESIDE_TRIBE_CHAT_ID"])
    try:
        bot.unpin_chat_message(chat_id, msg_id)
        save_state(LAST_PINNED, {"message_id": None})
        print(f"{GREEN}unpin-weekly{RESET}: unpinned message_id={msg_id}")
    except TelegramAPIError as e:
        # A 4xx means Telegram has considered the request and refused it: the
        # message was unpinned by hand, or deleted, or was never pinned because
        # the pin failed after posting (a case `cmd_sunday_preview` now
        # survives). Nothing about that recovers on a retry, and the recorded id
        # used to be kept, so the job printed this same warning every Wednesday
        # forever with no way out short of hand-editing last-pinned.json.
        #
        # A 5xx or a transport failure (status_code None) IS transient, and the
        # id is kept for the next run: dropping it there would leave a real
        # pinned message pinned for good.
        status = getattr(e, "status_code", None)
        if status is not None and 400 <= status < 500:
            save_state(LAST_PINNED, {"message_id": None})
            print(f"{YELLOW}unpin-weekly: Telegram refused message_id={msg_id} "
                  f"({e}); clearing the recorded id - it is already unpinned or "
                  f"gone{RESET}", file=sys.stderr)
            return
        print(f"{YELLOW}unpin-weekly: {e}; keeping message_id={msg_id} for the "
              f"next run{RESET}", file=sys.stderr)


# ============================================================
# Subcommand: log-session (Phase 3 task 3.11)
# ============================================================

def cmd_log_session(args) -> None:
    """Manually log a session result. Run after each Mon/Wed session.

    `--shared` and `--no-shows` take DISPLAY NAMES, matched against each
    schedule row's `speaker_name`. Not handles: the example here passed
    lower-case handles, which match no row, so the documented invocation logged
    nothing. Since the zero-match guard below it now exits 1, following the
    docstring produced an error rather than a silent no-op -- but the error names
    the date, not the thing that was actually wrong.

    The date was a second reason the same example could never work: 2026-05-12
    is a Tuesday, and a schedule only ever holds Mon and Wed rows, so the zero
    match was guaranteed whatever names were passed. It is a Monday now.

    CLI: python scripts/fireside-bot.py log-session --date 2026-05-11 \
                                                    --shared "Vesper Lynd,Felix Leiter" \
                                                    --no-shows ""
    """
    if not args.date:
        print(f"{RED}log-session: --date YYYY-MM-DD required{RESET}", file=sys.stderr)
        sys.exit(1)
    if not args.shared:
        print(f"{RED}log-session: --shared required (comma-separated speaker names){RESET}", file=sys.stderr)
        sys.exit(1)

    shared_names = [s.strip() for s in args.shared.split(",") if s.strip()]
    no_show_names = [s.strip() for s in (args.no_shows or "").split(",") if s.strip()]

    # Under the lock, like the two swap writers. This was the third
    # load-modify-save of schedule.json and the only one outside it: between the
    # read and the write, an accepted /swap took the lock in the webhook daemon,
    # wrote the swapped schedule, and this save replaced it with the pre-swap
    # copy. Both members had already been told the swap went through, and
    # swap-requests.jsonl still records it. `locked_state` is cross-process, so
    # a lock the daemon holds does nothing for a lock this process never asks
    # for. The guard test that should have caught this forbade one variable name
    # rather than the call, and this writer used a different one.
    updated = 0
    with locked_state(SCHEDULE, []) as schedule:
        # Mark schedule entries as completed
        for entry in schedule:
            if entry["session_date"] != args.date:
                continue
            if entry["speaker_name"] in no_show_names:
                entry["no_show"] = True
                entry["completed"] = True
                updated += 1
            elif entry["speaker_name"] in shared_names:
                entry["completed"] = True
                updated += 1

        # A date that matched NO schedule entry is a typo, not a session. The
        # event used to be logged and success printed regardless, so
        # `--date 2026-99-99` (or a real date off by one) wrote a
        # `session_logged` row into the stats corpus while the schedule went
        # untouched, and the line beneath said "schedule entries updated=0" as
        # though that were a normal outcome.
        #
        # Raising leaves the file alone: `locked_state` writes nothing when the
        # block raises, so the refusal cannot save the half-marked list it was
        # about to reject.
        if updated == 0:
            dates = sorted({e["session_date"] for e in schedule})
            near = [d for d in dates if d[:7] == args.date[:7]] or dates[-5:]
            print(f"{RED}log-session: no schedule entry matches --date {args.date} "
                  f"and the named speakers; nothing logged. Sessions this month: "
                  f"{', '.join(near) or '(none)'}{RESET}", file=sys.stderr)
            sys.exit(1)

    _log_event(
        "session_logged",
        date=args.date,
        shared=args.shared,
        no_shows=args.no_shows or "",
        swaps=getattr(args, "swaps", "") or "",
    )
    print(f"{GREEN}log-session{RESET}: {args.date} - shared={len(shared_names)}, "
          f"no-shows={len(no_show_names)}, schedule entries updated={updated}")


# ============================================================
# Phase 2/3 subcommand stubs (anything not yet wired)
# ============================================================

# ============================================================
# Webhook subcommands (Phase 4 — real-time delivery via setWebhook)
# ============================================================

def cmd_heartbeat(args) -> None:
    """Daemon liveness signal in webhook mode.

    Polls no longer run when FIRESIDE_WEBHOOK_ENABLED=true, so the per-poll
    side effects vanish: nothing pings FIRESIDE_HC_POLL (healthchecks.io flags
    DOWN), and nothing appends poll-tick to dm-log (cmd_health_check DMs Misha
    about stale polls). This heartbeat reinstates both signals from a 1-min
    cron job.

    It CALLS TELEGRAM before it stamps anything. It used to ping and append
    unconditionally, which proved only that cron fired and a JSONL append
    worked, while `poll-tick` - the marker `cmd_health_check` treats as
    equivalent - is written only after a successful `getUpdates`. So the two
    tick types certified different things under one name, and this one
    certified almost nothing: with Telegram unreachable from the host, the
    heartbeat kept stamping fresh ticks and health-check kept printing
    "healthy".

    `getWebhookInfo` rather than `getMe`, because it costs the same round trip
    and carries the only evidence a non-handler process has about the handler:
    `pending_update_count` (updates Telegram is holding because nothing consumed
    them) and `last_error_message`. Both are recorded on the tick row and
    reported by health-check. They do NOT raise an alert on their own: the
    threshold at which a pending count means "the handler is dead" rather than
    "a session is busy" is the operator's call, and inventing one here would
    trade a blind monitor for a noisy one. See the note in `cmd_health_check`.
    """
    row: dict = {"ts": local_now().isoformat(), "dm_type": "heartbeat-tick"}
    try:
        info = get_bot().get_webhook_info() or {}
    except TelegramAPIError as e:
        # No tick and no ping. This host could not reach Telegram, which is
        # exactly what the monitor exists to notice; stamping anyway would
        # report the outage as health.
        log_error(f"heartbeat: Telegram unreachable, no tick stamped: {e}")
        print(f"{RED}heartbeat: Telegram unreachable ({e}); no liveness tick "
              f"recorded{RESET}", file=sys.stderr)
        return
    row["pending_update_count"] = info.get("pending_update_count")
    row["webhook_last_error"] = info.get("last_error_message")
    hc_ping("FIRESIDE_HC_POLL")
    append_jsonl(DM_LOG, row)


def cmd_set_webhook(args) -> None:
    """Register the bot's webhook URL with Telegram and upload the self-signed cert.

    Reads FIRESIDE_WEBHOOK_PUBLIC_URL, FIRESIDE_WEBHOOK_SECRET, FIRESIDE_WEBHOOK_CERT
    from .env. Telegram will POST every future update to PUBLIC_URL with the
    SECRET in the X-Telegram-Bot-Api-Secret-Token header.
    """
    bot = get_bot()
    url = os.environ.get("FIRESIDE_WEBHOOK_PUBLIC_URL")
    secret = os.environ.get("FIRESIDE_WEBHOOK_SECRET")
    cert_path = os.environ.get("FIRESIDE_WEBHOOK_CERT")
    if not (url and secret and cert_path):
        print(f"{RED}Missing one of FIRESIDE_WEBHOOK_PUBLIC_URL / SECRET / CERT in .env{RESET}",
              file=sys.stderr)
        sys.exit(1)
    if not Path(cert_path).exists():
        print(f"{RED}Cert file not found: {cert_path}{RESET}", file=sys.stderr)
        sys.exit(1)

    data = {
        "url": url,
        "secret_token": secret,
        "allowed_updates": json.dumps([
            "message", "message_reaction", "message_reaction_count",
            "chat_member", "my_chat_member", "callback_query",
        ]),
        "drop_pending_updates": "false",
    }
    # Through the bot, not around it. This built its own
    # `.../bot{token}/setWebhook` URL and called `requests.post` on it, which is
    # the one Telegram call in this file that skips `TelegramBot._call` and the
    # token redaction its docstring promises. A connection reset, an expired TLS
    # chain or a DNS failure raises a `requests` exception whose message quotes
    # the URL, so the bot token landed in the terminal and in the traceback. The
    # bare `r.json()` had the same shape: a proxy or captive portal answering
    # with HTML raised JSONDecodeError out of a command whose only error path is
    # the `ok` check below.
    with open(cert_path, "rb") as f:
        files = {"certificate": (Path(cert_path).name, f, "application/x-pem-file")}
        try:
            result = bot.call_multipart("setWebhook", data=data, files=files)
        except TelegramAPIError as e:
            print(f"{RED}setWebhook failed: {e}{RESET}", file=sys.stderr)
            sys.exit(1)
    print(f"{GREEN}OK{RESET}  webhook set to {url}")
    print(f"     result: {result}")


def cmd_delete_webhook(args) -> None:
    """Clear the bot's webhook. Polling becomes possible again immediately."""
    bot = get_bot()
    result = bot._call("deleteWebhook", drop_pending_updates=False)
    print(f"{GREEN}OK{RESET}  webhook deleted (result={result})")


def cmd_webhook_info(args) -> None:
    """Print the current webhook registration as Telegram sees it."""
    bot = get_bot()
    info = bot._call("getWebhookInfo")
    print(json.dumps(info, indent=2, ensure_ascii=False))


# ============================================================
# Main
# ============================================================

def main() -> int:
    parser = argparse.ArgumentParser(
        prog="fireside-bot",
        description="Tribe Fireside Bot - coordinates Mon + Wed firesides via Telegram.",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="See docs/superpowers/specs/2026-05-03-tribe-fireside-bot-design.md for details.",
    )
    sub = parser.add_subparsers(dest="cmd", required=True, metavar="<subcommand>")

    # Phase 1 - implemented
    sub.add_parser("test-telegram", help="Smoke test: send DM to Misha (Phase 1 DoD)")
    sub.add_parser("xlsx-check", help="Print xlsx loader summary (Phase 1 helper)")
    sub.add_parser("init-state", help="Initialise state directory + files (Phase 1 helper)")

    # Phase 2+ - stubs registered so --help shows full surface
    sub.add_parser("bootstrap", help="One-time: enumerate Telegram group + build initial roster (Phase 2)")
    sub.add_parser("poll", help="Process Telegram updates, every 5 min (Phase 3)")
    sub.add_parser("speaker-dms", help="Send 2-week + 3-day speaker reminders (Phase 3)")
    sunday_preview = sub.add_parser("sunday-preview", help="Post pinned weekly preview (Phase 3)")
    sunday_preview.add_argument("--dry-run", action="store_true",
                                help="Print rendered preview without posting to group")
    sub.add_parser("dayof-reminders", help="DM speakers Zoom link 3h before session (Phase 3)")
    sub.add_parser("helmsman-brief", help="Brief next week's Helmsman 7 days ahead (Phase 3)")
    helmsman = sub.add_parser("helmsman", help="Assign / list / audit Helmsman coverage")
    helmsman_sub = helmsman.add_subparsers(dest="action", required=True, metavar="<action>")
    helmsman_set = helmsman_sub.add_parser("set", help="Assign a Helmsman to one week")
    helmsman_set.add_argument("--week", required=True, help="Week-starting Monday, YYYY-MM-DD")
    helmsman_set.add_argument("--username", required=True, help="Telegram handle from the roster")
    helmsman_set.add_argument("--note", default="", help="Optional note stored on the record")
    helmsman_sub.add_parser("list", help="Show every scheduled week and its Helmsman")
    helmsman_sub.add_parser("gaps", help="List unassigned weeks; exits 1 if any")
    sub.add_parser("speaker-gaps", help="List members with no slot this cycle; exits 1 if any")
    sub.add_parser("weekly-discrepancy-report", help="Report Telegram-vs-xlsx mismatches (Phase 3)")
    sub.add_parser("email-backup", help="Email reminder for unresponsive Tribe (Phase 3)")
    stats = sub.add_parser("stats", help="Generate stats markdown report (Phase 3)")
    stats.add_argument("--show", action="store_true", help="Print the report after writing")
    sub.add_parser("health-check", help="Alert if poll hasn't run in 30 min (Phase 3)")
    sub.add_parser("unpin-weekly", help="Unpin Sunday preview after Wed session (Phase 3)")
    topic_nudge = sub.add_parser("topic-nudge", help="Post weekly topic-collection invite to Tribe")
    topic_nudge.add_argument("--dry-run", action="store_true", help="Print without posting")
    topic_digest = sub.add_parser("topic-digest", help="DM CEO new topic ideas since last digest")
    topic_digest.add_argument("--dry-run", action="store_true", help="Print without sending")
    cycle_end = sub.add_parser("cycle-end-invite", help="Draft end-of-cycle topic invite to CEO for approval")
    cycle_end.add_argument("--dry-run", action="store_true", help="Print draft without DMing CEO")
    cycle_rollover = sub.add_parser("cycle-rollover", help="Rebuild schedule.json from fresh config when the cycle rolls over")
    cycle_rollover.add_argument("--dry-run", action="store_true", help="Show what would rebuild without writing")
    topic_ideas = sub.add_parser("topic-ideas", help="List the topic backlog (terminal)")
    topic_ideas.add_argument("--cycle", type=int, default=None, help="Filter to one cycle")
    topic_ideas.add_argument("--new", action="store_true", help="Only ideas since last digest")

    log_session = sub.add_parser("log-session", help="Log session result, manual (Phase 3)")
    log_session.add_argument("--date", help="Session date YYYY-MM-DD")
    log_session.add_argument("--shared",
                             help="Comma-separated speaker DISPLAY NAMES who shared (not handles)")
    log_session.add_argument("--no-shows", default="",
                             help="Comma-separated speaker DISPLAY NAMES who no-showed")
    log_session.add_argument("--swaps", default="", help="Comma-separated swap notes")

    # Phase 4 - webhook subcommands
    sub.add_parser("set-webhook", help="Register webhook URL with Telegram (Phase 4)")
    sub.add_parser("delete-webhook", help="Clear webhook so polling can resume (Phase 4)")
    sub.add_parser("webhook-info", help="Show current Telegram webhook registration (Phase 4)")
    sub.add_parser("heartbeat", help="Ping FIRESIDE_HC_POLL — alive signal in webhook mode (Phase 4)")

    args = parser.parse_args()

    handlers = {
        "test-telegram": cmd_test_telegram,
        "xlsx-check": cmd_xlsx_check,
        "init-state": cmd_init_state,
        "bootstrap": cmd_bootstrap,
        "poll": cmd_poll,
        "speaker-dms": cmd_speaker_dms,
        "sunday-preview": cmd_sunday_preview,
        "dayof-reminders": cmd_dayof_reminders,
        "helmsman-brief": cmd_helmsman_brief,
        "helmsman": cmd_helmsman,
        "speaker-gaps": cmd_speaker_gaps,
        "weekly-discrepancy-report": cmd_weekly_discrepancy_report,
        "email-backup": cmd_email_backup,
        "stats": cmd_stats,
        "health-check": cmd_health_check,
        "unpin-weekly": cmd_unpin_weekly,
        "topic-nudge": cmd_topic_nudge,
        "topic-digest": cmd_topic_digest,
        "cycle-end-invite": cmd_cycle_end_invite,
        "cycle-rollover": cmd_cycle_rollover,
        "topic-ideas": cmd_topic_ideas,
        "log-session": cmd_log_session,
        "set-webhook": cmd_set_webhook,
        "delete-webhook": cmd_delete_webhook,
        "webhook-info": cmd_webhook_info,
        "heartbeat": cmd_heartbeat,
    }
    handler = handlers.get(args.cmd)
    if handler is None:
        parser.error(f"Unknown subcommand: {args.cmd}")
    # Handlers return None (success) or an explicit exit code - `helmsman gaps`
    # exits 1 on an unassigned week so it can gate a pipeline.
    return handler(args) or 0


if __name__ == "__main__":
    # Wrap main() so uncaught exceptions land in errors.log even when the script
    # is launched via pythonw (no console, stderr discarded). Cron-fired runs
    # would otherwise fail silently.
    try:
        sys.exit(main())
    except SystemExit:
        raise
    except BaseException as _exc:
        try:
            log_error(f"uncaught exception in main()", _exc)
        except BaseException:  # noqa: S110 - last-resort handler; log_error itself failed and we are already exiting non-zero, so there is nothing safe left to do.
            pass
        sys.exit(1)
