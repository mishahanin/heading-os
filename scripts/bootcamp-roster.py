#!/usr/bin/env python3
"""Build a Bootcamp Tribe roster + track recommendations.

Event-specific values (input/output xlsx names, event folder, sheet title) are
instance DATA: they resolve from the private bootcamp-org-chart config in the
data overlay; the engine ships scripts/bootcamp-org-chart.example.json with
generic placeholders.

Reads:
  - outputs/_sync/gal-<domain>.json (from gal-export.py)
  - the org-chart markdown referenced in ops
  - datastore/events/<event>/<prelim>.xlsx   (event + filename from config)

Writes:
  - datastore/events/<event>/<roster>.xlsx   (event + filename from config)

Logic:
  - Filter out Public DLs and shared/system mailboxes
  - Filter out non-Tribe members (resellers, shareholders) listed in the org-chart config
  - Override GAL title with org chart title where the chart is the current revision
  - Tag each Tribe member with: Function, Reports To, In prelim?, Tech track?, Ops/Exec track?, Rationale
"""
from __future__ import annotations

from __future__ import annotations

import json
import os
import sys
from datetime import datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from scripts.utils.venv_guard import ensure_venv  # noqa: E402

ensure_venv()
from scripts.utils.operator_identity import operator_email_domain  # noqa: E402
from scripts.utils.workspace import (
    DataRootError,
    get_datastore_dir,
    get_default_tz,
    get_outputs_dir,
    get_workspace_root,
    resolve_config_with_example,
)

WS = get_workspace_root()

# ============================================================
# Exclusion lists + org chart are per-instance DATA (real Tribe names, titles,
# internal mailboxes). They live in the data overlay at
# <data-root>/config/bootcamp-org-chart.json (resolved via get_data_config_dir());
# the engine ships scripts/bootcamp-org-chart.example.json as the generic fallback.
# Schema: chart = email_local -> {title, function, reports_to}.
# ============================================================
_ORG_CHART_EXAMPLE = WS / "scripts" / "bootcamp-org-chart.example.json"


def _org_chart_file() -> Path:
    """The org-chart path, resolved at call time and never at import.

    Every resolver below reaches `get_data_root()`, which reads HEADING_OS_DATA
    on EVERY call. Asking at module scope asked once, during this module's own
    import, and stored the answer -- so a caller that repointed the root
    afterwards still read the first overlay, and `out_xlsx()` below reaches a
    `mkdir` that `tests/conftest.py` does not wrap.
    """
    try:
        return resolve_config_with_example(
            "bootcamp-org-chart.json", _ORG_CHART_EXAMPLE
        )
    except (DataRootError, ValueError, OSError) as exc:
        # Same call the operator seam makes in operator_identity._resolve_file,
        # and for the same reason. resolve_config_with_example() reaches
        # get_data_root(), which REFUSES when HEADING_OS_DATA names a path that
        # is not a directory -- a guard against a WRITE landing on the live
        # overlay. This is a READ, and the banner above already documents the
        # lower tier: the shipped example. Until 2026-08-30 the refusal arrived
        # at module scope, so importing this module at all died with a traceback
        # and exit 1 on a machine with no overlay. Fall to the example and say
        # so; the paths that genuinely need the overlay refuse separately, below.
        print(f"[bootcamp-roster] the private data overlay could not be resolved "
              f"({type(exc).__name__}: {exc}); reading the org chart from the "
              f"shipped example {_ORG_CHART_EXAMPLE.name} instead. Its names and "
              f"titles are generic placeholders, not your Tribe.", file=sys.stderr)
        return _ORG_CHART_EXAMPLE


# Parsed on first use and cached per data root. The key is the HEADING_OS_DATA
# value itself, so a caller that repoints the root MISSES the cache and reloads
# rather than being served the first overlay's chart for the life of the
# process. A plain module-level parse had no key and could not do that.
_ORG_CACHE: dict[str, dict] = {}


def _org_data() -> dict:
    key = os.environ.get("HEADING_OS_DATA", "")
    data = _ORG_CACHE.get(key)
    if data is None:
        data = _ORG_CACHE[key] = json.loads(
            _org_chart_file().read_text(encoding="utf-8"))
    return data


def shared_mailboxes() -> set[str]:
    return set(_org_data()["shared_mailboxes"])


def non_tribe() -> set[str]:
    # dict email -> reason; membership uses keys
    return set(_org_data()["non_tribe"])


def chart() -> dict:
    return _org_data()["chart"]


def _leader_emails() -> set[str]:
    return set(_org_data()["leader_emails"])


def _gal_aliases() -> dict:
    return _org_data()["aliases"]


def _gal_domain() -> str:
    """The tenant domain the GAL export filename carries.

    `gal-export.py` writes `gal-<domain>.json`. This module hardcoded one
    company's domain until 2026-08-23, so on any other deployment the writer and
    the reader named different files and `build_roster()` died on a missing
    path. The domain is instance data like everything else here;
    `operator_email_domain()` is the fallback so a workspace that never edited
    the org chart still lines up with what `gal-export.py` produced.
    """
    return _org_data().get("gal_domain") or operator_email_domain() or "example.com"


def _event() -> dict:
    # Event-specific paths/title are instance DATA resolved from the (private)
    # config; the engine example ships generic placeholders.
    return _org_data().get("event", {})


def _event_dir() -> str:
    return _event().get("dir", "Example Bootcamp")


# These three name real overlay data: the GAL export this reads, the preliminary
# attendee list, and the roster it writes. Unlike the org chart above there is no
# lower tier to fall to -- an engine example cannot stand in for the operator's
# attendee list -- so an unresolvable overlay is a REFUSAL here, not a
# degradation. What it must not be is a traceback at module scope: fixing the
# org-chart line alone simply moved the same crash from line 55 to this one, and
# a crash during import happens before any argument is parsed or any message is
# chosen. main() asks overlay_error() and states it.
def gal_json() -> Path:
    return get_outputs_dir() / "_sync" / f"gal-{_gal_domain()}.json"


def prelim_xlsx() -> Path:
    return (get_datastore_dir() / "events" / _event_dir()
            / _event().get("prelim_xlsx", "prelim.xlsx"))


def out_xlsx() -> Path:
    return (get_datastore_dir() / "events" / _event_dir()
            / _event().get("out_xlsx", "roster.xlsx"))


def overlay_error() -> str | None:
    """The refusal message for main(), or None when all three paths resolve."""
    try:
        for resolve in (gal_json, prelim_xlsx, out_xlsx):
            resolve()
    except (DataRootError, ValueError, OSError) as exc:
        return f"{type(exc).__name__}: {exc}"
    return None

# ============================================================
# Track recommendation logic
# ============================================================
def recommend_tracks(email_local: str, function: str, title: str) -> tuple[str, str, str]:
    """Return (tech_track, ops_exec_track, rationale)."""
    f = (function or "").lower()
    t = (title or "").lower()

    # CEO + executives that touch both - go to BOTH (data-driven; see _leader_emails)
    if email_local in _leader_emails():
        return "Y", "Y", "Leadership with technical authority — attend both passes"

    # InfoSec sits between - attend both.
    #
    # This ran LAST, after the two substring rules below, and was therefore
    # unreachable for every combined label. `technical_functions` contains
    # "engineering" and `ops_functions` contains "operations", so "InfoSec
    # Engineering" returned Tech-only and "TrustONE Operations" returned
    # Ops-only -- each contradicting this comment, for exactly the population
    # the branch was written for. The specific rule goes before the generic
    # ones; that is the whole fix.
    if "infosec" in f or "trustone" in f:
        return "Y", "Y", "InfoSec/TrustONE — both passes (technical + governance)"

    # Pure technical chain (Engineering, AI Lab researchers, DevOps, QA)
    technical_functions = (
        "engineering", "ai lab", "devops", "qa", "core engine",
        "backend", "frontend", "ai engineering",
    )
    if any(k in f for k in technical_functions):
        return "Y", "N", f"Technical IC ({function}) — Tech pass only"

    # Strategy / Marketing / Product / HR / Finance / Legal / Pre-sales = ops/exec only
    ops_functions = (
        "strategy", "marketing", "hr", "finance", "legal",
        "operations", "product", "pre-sales", "customer alignment",
    )
    if any(k in f for k in ops_functions):
        return "N", "Y", f"{function} — Ops/Executive pass only"

    # InfoSec sits between - attend both
    if "infosec" in f or "trustone" in f:
        return "Y", "Y", "InfoSec/TrustONE — both passes (technical + governance)"

    # Default: title heuristics for unknowns
    if any(k in t for k in ("developer", "devloper", "engineer", "devops", "qa", "architect", "researcher", "ml ", "ai ")):
        return "Y", "N", f"Title indicates IC technical role ({title}) — Tech pass"
    if any(k in t for k in ("analyst",)):
        return "N", "Y", f"Business/data analyst ({title}) — Ops/Exec pass"
    if any(k in t for k in ("manager", "director", "officer", "ceo", "cto", "coo", "cfo", "chief", "vp", "head", "lead")):
        return "N", "Y", f"Leadership/management ({title}) — Ops/Exec pass"

    return "?", "?", "Unknown role — needs CEO confirmation"


# ============================================================
# Preliminary list parser
# ============================================================
# Headers that mark the column holding attendee names. Matched case-folded and
# whitespace-stripped, against the first cell in the sheet that matches.
_NAME_HEADERS = ("name", "names", "full name", "attendee", "attendees",
                 "participant", "participants")


class PrelimUnavailable(RuntimeError):
    """The preliminary list could not be read at all."""


def _prelim_column(ws) -> tuple[int, int] | None:
    """(column index, header row index) of the name column, or None.

    Zero-indexed, both. None means no recognised header anywhere in the sheet.
    """
    for r, row in enumerate(ws.iter_rows(values_only=True)):
        for c, cell in enumerate(row):
            if isinstance(cell, str) and cell.strip().lower() in _NAME_HEADERS:
                return c, r
    return None


def load_prelim() -> set[str]:
    """Return the set of attendee-name strings (lowercased) from the prelim Excel.

    One column, not the whole workbook. This read EVERY non-empty string cell in
    the sheet and excluded only the literal header `"name"`, so job titles,
    departments, locations, emails and free-text notes all landed in the set
    `in_prelim` matches first names, last names and initials against. Any
    employee whose first or last name equalled any word standing alone in any
    cell -- a city in a travel column, a surname in a catering note -- was
    reported `In Prelim List? = Y` without being on the list. The docstring
    described the intended behaviour correctly the whole time; the loop did not.

    When no recognised header is found, the whole-sheet read is kept and said
    out loud, because refusing would break a run over a sheet shape nobody has
    described here, and silently returning nothing would mark every attendee
    absent.

    Raises PrelimUnavailable when the file cannot be opened or parsed. It used
    to return an empty set, which is indistinguishable from "the list is empty":
    the run then completed, exited 0, and wrote a roster whose entire `In Prelim
    List?` column read `N` behind a single WARN line.
    """
    if openpyxl is None:                      # pragma: no cover - CLI calls the guard
        raise PrelimUnavailable(
            "openpyxl is not bound; call _ensure_openpyxl() before build_roster()")
    try:
        wb = openpyxl.load_workbook(prelim_xlsx(), data_only=True)
    except Exception as e:                    # noqa: BLE001 - re-raised, see below
        # Broad on purpose, and RE-RAISED, never swallowed. A .xlsx is a zip
        # container, so a truncated download surfaces as `zipfile.BadZipFile`,
        # which derives from Exception and not from OSError; a wrong extension
        # is openpyxl's own `InvalidFileException`; a damaged sheet can be a
        # KeyError from deep inside the reader. Enumerating that list means
        # missing the next member of it, and the one thing this must never do
        # again is turn an unreadable file into an empty attendee set.
        raise PrelimUnavailable(f"cannot read {prelim_xlsx()}: {e}") from e
    ws = wb.active

    found = _prelim_column(ws)
    if found is None:
        print(f"[WARN] no name column found in {prelim_xlsx()} "
              f"(looked for {', '.join(_NAME_HEADERS)}). Falling back to every "
              f"cell in the sheet, which over-reports: any cell equal to "
              f"someone's name marks them present.")
        return {cell.strip().lower()
                for row in ws.iter_rows(values_only=True)
                for cell in row
                if isinstance(cell, str) and cell.strip()}

    col, header_row = found
    names = set()
    for r, row in enumerate(ws.iter_rows(values_only=True)):
        if r <= header_row or col >= len(row):
            continue
        cell = row[col]
        if isinstance(cell, str) and cell.strip():
            names.add(cell.strip().lower())
    return names


def in_prelim(display_name: str, email_local: str, prelim: set[str]) -> bool:
    """Match prelim short names against full GAL names."""
    # `if not display_name` rejected "" and passed "   " straight through, and
    # "   ".split() is [], so [0] raised IndexError and killed the whole roster
    # build -- after the filtering, so not even a partial file was written.
    parts = display_name.split()
    if not parts:
        return False
    first_name = parts[0].lower()
    last_name = parts[-1].lower() if len(parts) > 1 else ""
    # Exact first-name hit
    if first_name in prelim:
        return True
    # Initials + last-name (e.g., "K. Mertens" matches "Karl Mertens")
    initial_form = f"{first_name[0]}. {last_name}".lower() if last_name else ""
    initial_form_no_space = f"{first_name[0]}.{last_name}".lower() if last_name else ""
    if initial_form in prelim or initial_form_no_space in prelim:
        return True
    # Last name only
    if last_name and last_name in prelim:
        return True
    # Special aliases (data-driven; see _gal_aliases)
    for k, v in _gal_aliases().items():
        if first_name == k and any(a in prelim for a in v):
            return True
    return False


# ============================================================
# Build roster
# ============================================================
def build_roster() -> tuple[list[dict], dict]:
    with gal_json().open(encoding="utf-8") as f:
        gal = json.load(f)
    prelim = load_prelim()

    rows = []
    excluded = {"public_dl": [], "shared_mailbox": [], "non_tribe": []}

    # Resolved once for the whole sweep rather than per row: each of these
    # rebuilds a set from the cached chart, and the chart cannot move mid-loop.
    shared = shared_mailboxes()
    excluded_non_tribe = non_tribe()
    org_chart = chart()

    for r in gal:
        email = r.get("email") or ""
        if not email:
            continue
        if r.get("mailbox_type") == "PublicDL":
            excluded["public_dl"].append(email)
            continue
        if email in shared:
            excluded["shared_mailbox"].append(email)
            continue
        if email in excluded_non_tribe:
            excluded["non_tribe"].append(email)
            continue

        local = email.split("@")[0]
        chart_entry = org_chart.get(local)
        gal_title = r.get("job_title") or ""
        if chart_entry:
            title = chart_entry["title"]
            function = chart_entry["function"]
            reports_to = chart_entry["reports_to"]
            in_chart = "Y"
        else:
            title = gal_title or "TBD"
            function = "TBD - not in the org chart"
            reports_to = "TBD"
            in_chart = "N"

        display_name = r.get("display_name") or r.get("name") or ""
        in_prelim_flag = "Y" if in_prelim(display_name, local, prelim) else "N"

        tech, ops, rationale = recommend_tracks(local, function, title)

        rows.append({
            "name": display_name,
            "email": email,
            "title_chart": title,
            "title_gal": gal_title,
            "function": function,
            "reports_to": reports_to,
            "in_chart": in_chart,
            "in_prelim": in_prelim_flag,
            "tech": tech,
            "ops_exec": ops,
            "rationale": rationale,
        })

    rows.sort(key=lambda r: (r["function"], r["name"]))
    return rows, excluded


# ============================================================
# Excel writer
# ============================================================
# openpyxl names + style constants are bound lazily (F-2.1: import stays pure).
openpyxl = Alignment = Font = PatternFill = get_column_letter = None
HEADER_FILL = HEADER_FONT = TECH_FILL = OPS_FILL = BOTH_FILL = UNKNOWN_FILL = None


def _ensure_openpyxl():
    global openpyxl, Alignment, Font, PatternFill, get_column_letter
    global HEADER_FILL, HEADER_FONT, TECH_FILL, OPS_FILL, BOTH_FILL, UNKNOWN_FILL
    if openpyxl is not None:
        return
    from scripts.utils.optdeps import require
    openpyxl = require("openpyxl", extra="documents")
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    HEADER_FILL = PatternFill(start_color="1A2332", end_color="1A2332", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
    TECH_FILL = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
    OPS_FILL = PatternFill(start_color="FFF4E6", end_color="FFF4E6", fill_type="solid")
    BOTH_FILL = PatternFill(start_color="E8F8E8", end_color="E8F8E8", fill_type="solid")
    UNKNOWN_FILL = PatternFill(start_color="F8E8E8", end_color="F8E8E8", fill_type="solid")


def row_fill(tech: str, ops: str) -> PatternFill | None:
    if tech == "Y" and ops == "Y":
        return BOTH_FILL
    if tech == "Y":
        return TECH_FILL
    if ops == "Y":
        return OPS_FILL
    if tech == "?" or ops == "?":
        return UNKNOWN_FILL
    return None


def write_excel(rows: list[dict], excluded: dict):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tribe Roster"

    # Title row
    ws["A1"] = _event().get("title", "Bootcamp - Tribe Roster & Track Recommendations")
    ws["A1"].font = Font(bold=True, size=14)
    # A:L, not A:K -- `headers` carries 12 entries, so the table runs to
    # column L ("Rationale", the widest at width 50) and the banner used
    # to stop one short of it.
    ws.merge_cells("A1:L1")
    generated = datetime.now(get_default_tz()).date().isoformat()
    ws["A2"] = f"Tribe Roster & Track Recommendations | Generated {generated} from Exchange GAL + org chart"
    ws["A2"].font = Font(italic=True, color="666666")
    ws.merge_cells("A2:L2")

    headers = [
        "#", "Name", "Email", "Title (reconciled)", "GAL Title (raw)",
        "Function / Department", "Reports To",
        "In Org Chart?", "In Prelim List?",
        "Attend Tech Track?", "Attend Ops/Exec Track?", "Rationale",
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for idx, r in enumerate(rows, 1):
        row_num = idx + 4
        cells = [
            idx, r["name"], r["email"], r["title_chart"], r["title_gal"],
            r["function"], r["reports_to"],
            r["in_chart"], r["in_prelim"],
            r["tech"], r["ops_exec"], r["rationale"],
        ]
        for col, v in enumerate(cells, 1):
            c = ws.cell(row=row_num, column=col, value=v)
            c.alignment = Alignment(vertical="top", wrap_text=True)
        fill = row_fill(r["tech"], r["ops_exec"])
        if fill:
            for col in range(1, len(cells) + 1):
                ws.cell(row=row_num, column=col).fill = fill

    # Column widths
    widths = [4, 24, 32, 38, 30, 30, 22, 8, 8, 8, 8, 50]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.row_dimensions[4].height = 32
    ws.freeze_panes = "A5"

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2["A1"] = "Bootcamp Roster Summary"
    ws2["A1"].font = Font(bold=True, size=14)

    tech_count = sum(1 for r in rows if r["tech"] == "Y")
    ops_count = sum(1 for r in rows if r["ops_exec"] == "Y")
    both_count = sum(1 for r in rows if r["tech"] == "Y" and r["ops_exec"] == "Y")
    unknown_count = sum(1 for r in rows if r["tech"] == "?" or r["ops_exec"] == "?")

    rows2 = [
        ("Metric", "Value"),
        ("Total Tribe in GAL (filtered)", len(rows)),
        ("Recommended for Tech track", tech_count),
        ("Recommended for Ops/Exec track", ops_count),
        ("Recommended for BOTH passes", both_count),
        ("Unknown role (needs CEO confirmation)", unknown_count),
        ("In org chart", sum(1 for r in rows if r["in_chart"] == "Y")),
        ("Not in org chart (newer/contractors)", sum(1 for r in rows if r["in_chart"] == "N")),
        ("In the preliminary list", sum(1 for r in rows if r["in_prelim"] == "Y")),
        ("", ""),
        ("Excluded — Public DLs", len(excluded["public_dl"])),
        ("Excluded — Shared/system mailboxes", len(excluded["shared_mailbox"])),
        ("Excluded — Non-Tribe (resellers/shareholders)", len(excluded["non_tribe"])),
    ]
    for i, (k, v) in enumerate(rows2, 3):
        ws2.cell(row=i, column=1, value=k).font = Font(bold=(i == 3))
        ws2.cell(row=i, column=2, value=v)

    ws2.column_dimensions["A"].width = 50
    ws2.column_dimensions["B"].width = 12

    # Excluded detail sheet
    ws3 = wb.create_sheet("Excluded")
    ws3["A1"] = "Excluded entries (for audit)"
    ws3["A1"].font = Font(bold=True, size=14)
    ws3.append([])
    ws3.append(["Reason", "Email"])
    for label, key in [("Public DL", "public_dl"), ("Shared mailbox", "shared_mailbox"), ("Non-Tribe", "non_tribe")]:
        for e in excluded[key]:
            ws3.append([label, e])
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 40

    out_path = out_xlsx()
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"[OK] Wrote {out_path}")
    print(f"     Tribe roster: {len(rows)} | Tech: {tech_count} | Ops/Exec: {ops_count} | Both: {both_count} | Unknown: {unknown_count}")


def main() -> int:
    # Resolve the org chart BEFORE the refusal below, and discard the result.
    # It looks like a pointless call and it is not: on an unresolvable overlay
    # `_org_chart_file()` degrades to the shipped example and SAYS so on stderr,
    # and that notice used to be printed during this module's import, ahead of
    # the refusal. Moving the resolution to call time silently dropped it --
    # `tests/test_three_admin_tools_that_died_one_frame_below_the_seam.py`
    # caught the loss. Asking here keeps both messages and their original order.
    _org_data()
    overlay_refusal = overlay_error()
    if overlay_refusal is not None:
        # Loud, and with the same exit code the other refusal below uses. A
        # partial run is the danger this closes: every path this tool reads and
        # writes lives in the overlay, so without one there is no roster to
        # build and nothing that could look complete.
        print(f"[ERROR] the private data overlay is unreachable "
              f"({overlay_refusal})", file=sys.stderr)
        print("        The GAL export, the preliminary attendee list and the "
              "roster output all live there. Refusing to run.", file=sys.stderr)
        return 1
    _ensure_openpyxl()
    try:
        rows, excluded = build_roster()
    except PrelimUnavailable as e:
        # Refuse the artifact rather than ship a wrong one. Without the prelim
        # list every row's `In Prelim List?` reads `N`, which is a plausible,
        # readable, entirely false attendance column -- and the run exited 0
        # behind one WARN line, so nothing downstream could tell.
        print(f"[ERROR] {e}", file=sys.stderr)
        print("        The roster would mark every attendee as NOT on the "
              "preliminary list. Refusing to write it.", file=sys.stderr)
        return 1
    write_excel(rows, excluded)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
